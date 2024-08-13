# Necessary imports
from datetime import datetime, timedelta
from openpyxl import load_workbook
import mysql.connector
import openpyxl
import os
import pandas as pd
import time

# Function to copy values from source range to destination range
def copy_values(source_sheet, dest_sheet, source_range, dest_range):
    source_start, source_end = source_range
    dest_start, dest_end = dest_range
    
    # Copy values from source to destination
    source_values = []
    for row in source_sheet[source_start:source_end]:
        for cell in row:
            source_values.append(cell.value)
    
    dest_index = 0
    for row in dest_sheet[dest_start:dest_end]:
        for cell in row:
            if dest_index < len(source_values):
                cell.value = source_values[dest_index]
                dest_index += 1
            else:
                break

def initial_function():
    # # Rates - Supplier Rates for the Month_July 2024
    # # Defining the base directory
    # base_directory_1 = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\MORE Energy Sourcing\013. Rate Analysis\005. 2024 Rate Analysis"

    # # Getting the current month and year
    # current_date = datetime.now()
    # current_month_numeric = current_date.strftime("%m")  # Month in numeric format (e.g., '01' for January)
    # current_month_name = current_date.strftime("%B")   # Full month name (e.g., 'January')
    # current_day = current_date.strftime("%d") # Day in numeric format (e.g., '26')
    # current_year = current_date.strftime("%Y")   # Year in 4-digit format (e.g., '2024')

    # #List of all month names
    # months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    # #Find index of current month
    # current_month_index = months.index(current_month_name)

    # # Initialize current_month_numeric_increment
    # current_month_numeric_increment = current_month_numeric # Default to current month if condition not met

    # if (current_day > '25') and (current_date.strftime("%m") == current_month_numeric):
    #     #Increment current month by converting to integer, adding 1, and formatting back to zero-padded string
    #     current_month_numeric_increment = '{:02d}'.format((int(current_month_numeric) % 12) + 1)
    #     #Calculate index of next month
    #     next_month_index = (current_month_index + 1) % 12
    #     current_month_name = months[next_month_index]

    # # Constructing the folder name
    # folder_prefix = current_month_numeric_increment.zfill(3)  # Zero-padded three-digit month number (e.g., '006' for June)
    # folder_name = f"{folder_prefix}. {current_month_name} {current_year}"

    # # Constructing the directory path
    # folder_path = os.path.join(base_directory_1, folder_name)

    # # Additional needed variables for the folder name
    # drs = "Daily Rate Simulation_"
    # extension = ".xlsx"

    # # Folder name (Example: Daily Rate Simulation_06252024) 
    # folder_name_2 = f"{drs}{current_month_numeric}{current_day}{current_year}{extension}"
    # folder_path_2 = os.path.join(folder_path, folder_name_2)

    base_directory = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\01_Energy Sourcing Files\03_Daily Reports\01_Day Ahead Projections"
    current_date = datetime.now()
    current_month_numeric = current_date.strftime("%m")
    current_month_name = current_date.strftime("%B")
    current_day = current_date.strftime("%d")
    current_year = current_date.strftime("%Y")

    folder_path = os.path.join(base_directory, current_year)

    #List of all month names
    months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    #Find index of current month
    current_month_index = months.index(current_month_name)

    if (current_day>'25') and (current_date.strftime("%m")==current_month_numeric):
        #Increment current month by converting to integer, adding 1, and formatting back to zero-padded string
        current_month_numeric_increment = '{:02d}'.format((int(current_month_numeric) % 12) + 1)
        #Calculate index of next month
        next_month_index = (current_month_index + 1) % 12
        current_month_name = months[next_month_index]
    elif(current_day<='25') and (current_date.strftime("%m")==current_month_numeric):
        current_month_numeric_increment = current_month_numeric

    # Constructing the folder name
    folder_prefix = current_month_numeric_increment.zfill(3)  # Zero-padded three-digit month number (e.g., '006' for June)
    folder_name = f"{folder_prefix}. {current_month_name} {current_year}"

    # Constructing the directory path
    folder_path_2 = os.path.join(folder_path, folder_name)

    # Additional needed variables for the folder name
    drs = "Daily Rate Simulation_"
    extension = ".xlsx"

    # Folder Name (Example: Daily Rate Simulation_06252024) 
    folder_name_2 = f"{drs}{current_month_numeric}{current_day}{current_year}{extension}"
    folder_path_3 = os.path.join(folder_path_2, folder_name_2)
    
    # For creating a new Excel file with the needed columns
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'HOUR' 
    sheet['B1'] = 'BCQ_SCPC' 
    sheet['C1'] = 'RATE_SCPC_B1_FIXED_FEE' 
    sheet['D1'] = 'RATE_SCPC_B1_VARIABLE_FEE'
    sheet['E1'] = 'RATE_SCPC_B2_FIXED_FEE' 
    sheet['F1'] = 'RATE_SCPC_B2_VARIABLE_FEE'
    sheet['G1'] = 'RATE_SCPC_AVE_FIXED_FEE' 
    sheet['H1'] = 'RATE_SCPC_AVE_VARIABLE_FEE'
    sheet['I1'] = 'BCQ_KSPC' 
    sheet['J1'] = 'RATE_KSPC_B1_FIXED_FEE' 
    sheet['K1'] = 'RATE_KSPC_B1_VARIABLE_FEE'
    sheet['L1'] = 'RATE_KSPC_B2_FIXED_FEE' 
    sheet['M1'] = 'RATE_KSPC_B2_VARIABLE_FEE'
    sheet['N1'] = 'RATE_KSPC_AVE_FIXED_FEE' 
    sheet['O1'] = 'RATE_KSPC_AVE_VARIABLE_FEE'
    sheet['P1'] = 'BCQ_EDC' 
    sheet['Q1'] = 'RATE_EDC_FIXED_FEE' 
    sheet['R1'] = 'RATE_EDC_VARIABLE_FEE'
    sheet['S1'] = 'TOTAL_SS_LOAD' 
    sheet['T1'] = 'CONTESTABLE_ENERGY'
    sheet['U1'] = 'TOTAL_BCQ' 
    sheet['V1'] = 'WESM_RATE'
    sheet['W1'] = 'CURRENT_RATE'

    workbook.save('current_rate.xlsx')
    workbook.close()
    print("Excel file 'current_rate.xlsx' created successfully.")

    # File path for destination Excel file
    destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\current_rate.xlsx'

    # Load source workbook
    source_wb = load_workbook(folder_path_3, data_only=True)  # Use data_only=True to get values instead of formulas
    # Sheet name of the source workbook
    source_sheet = source_wb['6. DAP Report']

    # Load destination workbook
    dest_wb = load_workbook(destination_file_path)

    # Sheet name of the destination workbook
    dest_sheet = dest_wb['Sheet']

    # Define source and destination ranges
    source_ranges = [('C11', 'C34'), ('G11', 'G34'), ('H11', 'H34'), ('I11', 'I34')]
    dest_ranges = [('A2', 'A25'), ('B2', 'B25'), ('I2', 'I25'), ('P2', 'P25')]

    # Copy data from source to destination based on specified ranges
    for i in range(len(source_ranges)):
        copy_values(source_sheet, dest_sheet, source_ranges[i], dest_ranges[i])

    # Save the destination workbook
    dest_wb.save(destination_file_path)

    # Close workbooks
    source_wb.close()
    dest_wb.close()
    print("Data transfer completed.")

    # 1st Hour
    value_B2 = dest_sheet['B2'].value
    value_I2 = dest_sheet['I2'].value
    value_P2 = dest_sheet['P2'].value

    result2 = value_B2 + value_I2 + value_P2
    dest_sheet['U2'] = result2

    # 2nd Hour
    value_B3 = dest_sheet['B3'].value
    value_I3 = dest_sheet['I3'].value
    value_P3 = dest_sheet['P3'].value

    result3 = value_B3 + value_I3 + value_P3
    dest_sheet['U3'] = result3

    # 3rd Hour
    value_B4 = dest_sheet['B4'].value
    value_I4 = dest_sheet['I4'].value
    value_P4 = dest_sheet['P4'].value

    result4 = value_B4 + value_I4 + value_P4
    dest_sheet['U4'] = result4

    # 4th Hour
    value_B5 = dest_sheet['B5'].value
    value_I5 = dest_sheet['I5'].value
    value_P5 = dest_sheet['P5'].value

    result5 = value_B5 + value_I5 + value_P5
    dest_sheet['U5'] = result5

    # 5th Hour
    value_B6 = dest_sheet['B6'].value
    value_I6 = dest_sheet['I6'].value
    value_P6 = dest_sheet['P6'].value

    result6 = value_B6 + value_I6 + value_P6
    dest_sheet['U6'] = result6

    # 6th Hour
    value_B7 = dest_sheet['B7'].value
    value_I7 = dest_sheet['I7'].value
    value_P7 = dest_sheet['P7'].value

    result7 = value_B7 + value_I7 + value_P7
    dest_sheet['U7'] = result7

    # 7th Hour
    value_B8 = dest_sheet['B8'].value
    value_I8 = dest_sheet['I8'].value
    value_P8 = dest_sheet['P8'].value

    result8 = value_B8 + value_I8 + value_P8
    dest_sheet['U8'] = result8

    # 8th Hour
    value_B9 = dest_sheet['B9'].value
    value_I9 = dest_sheet['I9'].value
    value_P9 = dest_sheet['P9'].value

    result9 = value_B9 + value_I9 + value_P9
    dest_sheet['U9'] = result9

    # 9th Hour
    value_B10 = dest_sheet['B10'].value
    value_I10 = dest_sheet['I10'].value
    value_P10 = dest_sheet['P10'].value

    result10 = value_B10 + value_I10 + value_P10
    dest_sheet['U10'] = result10

    # 10th Hour
    value_B11 = dest_sheet['B11'].value
    value_I11 = dest_sheet['I11'].value
    value_P11 = dest_sheet['P11'].value

    result11 = value_B11 + value_I11 + value_P11
    dest_sheet['U11'] = result11

    # 11th Hour
    value_B12 = dest_sheet['B12'].value
    value_I12 = dest_sheet['I12'].value
    value_P12 = dest_sheet['P12'].value

    result12 = value_B12 + value_I12 + value_P12
    dest_sheet['U12'] = result12

    # 12th Hour
    value_B13 = dest_sheet['B13'].value
    value_I13 = dest_sheet['I13'].value
    value_P13 = dest_sheet['P13'].value

    result13 = value_B13 + value_I13 + value_P13
    dest_sheet['U13'] = result13

    # 13th Hour
    value_B14 = dest_sheet['B14'].value
    value_I14 = dest_sheet['I14'].value
    value_P14 = dest_sheet['P14'].value

    result14 = value_B14 + value_I14 + value_P14
    dest_sheet['U14'] = result14

    # 14th Hour
    value_B15 = dest_sheet['B15'].value
    value_I15 = dest_sheet['I15'].value
    value_P15 = dest_sheet['P15'].value

    result15 = value_B15 + value_I15 + value_P15
    dest_sheet['U15'] = result15

    # 15th Hour
    value_B16 = dest_sheet['B16'].value
    value_I16 = dest_sheet['I16'].value
    value_P16 = dest_sheet['P16'].value

    result16 = value_B16 + value_I16 + value_P16
    dest_sheet['U16'] = result16

    # 16th Hour
    value_B17 = dest_sheet['B17'].value
    value_I17 = dest_sheet['I17'].value
    value_P17 = dest_sheet['P17'].value

    result17 = value_B17 + value_I17 + value_P17 
    dest_sheet['U17'] = result17

    # 17th Hour
    value_B18 = dest_sheet['B18'].value
    value_I18 = dest_sheet['I18'].value
    value_P18 = dest_sheet['P18'].value

    result18 = value_B18 + value_I18 + value_P18
    dest_sheet['U18'] = result18

    # 18th Hour
    value_B19 = dest_sheet['B19'].value
    value_I19 = dest_sheet['I19'].value
    value_P19 = dest_sheet['P19'].value

    result19 = value_B19 + value_I19 + value_P19
    dest_sheet['U19'] = result19

    # 19th Hour
    value_B20 = dest_sheet['B20'].value
    value_I20 = dest_sheet['I20'].value
    value_P20 = dest_sheet['P20'].value

    result20 = value_B20 + value_I20 + value_P20
    dest_sheet['U20'] = result20

    # 20th Hour
    value_B21 = dest_sheet['B21'].value
    value_I21 = dest_sheet['I21'].value
    value_P21 = dest_sheet['P21'].value

    result21 = value_B21 + value_I21 + value_P21
    dest_sheet['U21'] = result21

    # 21th Hour
    value_B22 = dest_sheet['B22'].value
    value_I22 = dest_sheet['I22'].value
    value_P22 = dest_sheet['P22'].value

    result22 = value_B22 + value_I22 + value_P22
    dest_sheet['U22'] = result22

    # 22th Hour
    value_B23 = dest_sheet['B23'].value
    value_I23 = dest_sheet['I23'].value
    value_P23 = dest_sheet['P23'].value

    result23 = value_B23 + value_I23 + value_P23
    dest_sheet['U23'] = result23

    # 23rd Hour
    value_B24 = dest_sheet['B24'].value
    value_I24 = dest_sheet['I24'].value
    value_P24 = dest_sheet['P24'].value

    result24 = value_B24 + value_I24 + value_P24
    dest_sheet['U24'] = result24

    # 24th Hour
    value_B25 = dest_sheet['B25'].value
    value_I25 = dest_sheet['I25'].value
    value_P25 = dest_sheet['P25'].value

    result25 = value_B25 + value_I25 + value_P25
    dest_sheet['U25'] = result25

    # Save the workbook back to file
    dest_wb.save(destination_file_path)

    # Example: Supplier Rates for the Month_August 2024
    srftm = "Supplier Rates for the Month_"
    date_currently = datetime.now()
    month_var = date_currently.strftime("%B") # Example: "August"
    extension = ".xlsx"
    supplier_rates_folder = f"{srftm}{month_var} {current_year}{extension}"
    supplier_rates_path = os.path.join(folder_path_2, supplier_rates_folder)

    # File path for destination Excel file
    destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\current_rate.xlsx'

    # Load source workbook
    source_wb_sup = load_workbook(supplier_rates_path, data_only=True)  # Use data_only=True to get values instead of formulas
    # Sheet name of the source workbook
    source_sheet_sup = source_wb_sup['RATES']

    # Load destination workbook
    dest_wb = load_workbook(destination_file_path)

    # Sheet name of the destination workbook
    dest_sheet = dest_wb['Sheet']

    scpc_b1_fixed_fee = source_sheet_sup['C17'].value
    scpc_b1_variable_fee = source_sheet_sup['C18'].value
    scpc_b2_fixed_fee = source_sheet_sup['D17'].value
    scpc_b2_variable_fee = source_sheet_sup['D18'].value
    kspc_b1_fixed_fee = source_sheet_sup['E17'].value
    kspc_b1_variable_fee = source_sheet_sup['E18'].value
    kspc_b2_fixed_fee = source_sheet_sup['F17'].value
    kspc_b2_variable_fee = source_sheet_sup['F18'].value
    edc_fixed_fee = source_sheet_sup['G17'].value
    edc_variable_fee = source_sheet_sup['G18'].value
    
    dest_sheet['C2'].value = scpc_b1_fixed_fee
    dest_sheet['C3'].value = scpc_b1_fixed_fee
    dest_sheet['C4'].value = scpc_b1_fixed_fee
    dest_sheet['C5'].value = scpc_b1_fixed_fee
    dest_sheet['C6'].value = scpc_b1_fixed_fee
    dest_sheet['C7'].value = scpc_b1_fixed_fee
    dest_sheet['C8'].value = scpc_b1_fixed_fee
    dest_sheet['C9'].value = scpc_b1_fixed_fee
    dest_sheet['C10'].value = scpc_b1_fixed_fee
    dest_sheet['C11'].value = scpc_b1_fixed_fee
    dest_sheet['C12'].value = scpc_b1_fixed_fee
    dest_sheet['C13'].value = scpc_b1_fixed_fee
    dest_sheet['C14'].value = scpc_b1_fixed_fee
    dest_sheet['C15'].value = scpc_b1_fixed_fee
    dest_sheet['C16'].value = scpc_b1_fixed_fee
    dest_sheet['C17'].value = scpc_b1_fixed_fee
    dest_sheet['C18'].value = scpc_b1_fixed_fee
    dest_sheet['C19'].value = scpc_b1_fixed_fee
    dest_sheet['C20'].value = scpc_b1_fixed_fee
    dest_sheet['C21'].value = scpc_b1_fixed_fee
    dest_sheet['C22'].value = scpc_b1_fixed_fee
    dest_sheet['C23'].value = scpc_b1_fixed_fee
    dest_sheet['C24'].value = scpc_b1_fixed_fee
    dest_sheet['C25'].value = scpc_b1_fixed_fee

    dest_sheet['D2'].value = scpc_b1_variable_fee
    dest_sheet['D3'].value = scpc_b1_variable_fee
    dest_sheet['D4'].value = scpc_b1_variable_fee
    dest_sheet['D5'].value = scpc_b1_variable_fee
    dest_sheet['D6'].value = scpc_b1_variable_fee
    dest_sheet['D7'].value = scpc_b1_variable_fee
    dest_sheet['D8'].value = scpc_b1_variable_fee
    dest_sheet['D9'].value = scpc_b1_variable_fee
    dest_sheet['D10'].value = scpc_b1_variable_fee
    dest_sheet['D11'].value = scpc_b1_variable_fee
    dest_sheet['D12'].value = scpc_b1_variable_fee
    dest_sheet['D13'].value = scpc_b1_variable_fee
    dest_sheet['D14'].value = scpc_b1_variable_fee
    dest_sheet['D15'].value = scpc_b1_variable_fee
    dest_sheet['D16'].value = scpc_b1_variable_fee
    dest_sheet['D17'].value = scpc_b1_variable_fee
    dest_sheet['D18'].value = scpc_b1_variable_fee
    dest_sheet['D19'].value = scpc_b1_variable_fee
    dest_sheet['D20'].value = scpc_b1_variable_fee
    dest_sheet['D21'].value = scpc_b1_variable_fee
    dest_sheet['D22'].value = scpc_b1_variable_fee
    dest_sheet['D23'].value = scpc_b1_variable_fee
    dest_sheet['D24'].value = scpc_b1_variable_fee
    dest_sheet['D25'].value = scpc_b1_variable_fee

    dest_sheet['E2'].value = scpc_b2_fixed_fee
    dest_sheet['E3'].value = scpc_b2_fixed_fee
    dest_sheet['E4'].value = scpc_b2_fixed_fee
    dest_sheet['E5'].value = scpc_b2_fixed_fee
    dest_sheet['E6'].value = scpc_b2_fixed_fee
    dest_sheet['E7'].value = scpc_b2_fixed_fee
    dest_sheet['E8'].value = scpc_b2_fixed_fee
    dest_sheet['E9'].value = scpc_b2_fixed_fee
    dest_sheet['E10'].value = scpc_b2_fixed_fee
    dest_sheet['E11'].value = scpc_b2_fixed_fee
    dest_sheet['E12'].value = scpc_b2_fixed_fee
    dest_sheet['E13'].value = scpc_b2_fixed_fee
    dest_sheet['E14'].value = scpc_b2_fixed_fee
    dest_sheet['E15'].value = scpc_b2_fixed_fee
    dest_sheet['E16'].value = scpc_b2_fixed_fee
    dest_sheet['E17'].value = scpc_b2_fixed_fee
    dest_sheet['E18'].value = scpc_b2_fixed_fee
    dest_sheet['E19'].value = scpc_b2_fixed_fee
    dest_sheet['E20'].value = scpc_b2_fixed_fee
    dest_sheet['E21'].value = scpc_b2_fixed_fee
    dest_sheet['E22'].value = scpc_b2_fixed_fee
    dest_sheet['E23'].value = scpc_b2_fixed_fee
    dest_sheet['E24'].value = scpc_b2_fixed_fee
    dest_sheet['E25'].value = scpc_b2_fixed_fee

    dest_sheet['F2'].value = scpc_b2_variable_fee
    dest_sheet['F3'].value = scpc_b2_variable_fee
    dest_sheet['F4'].value = scpc_b2_variable_fee
    dest_sheet['F5'].value = scpc_b2_variable_fee
    dest_sheet['F6'].value = scpc_b2_variable_fee
    dest_sheet['F7'].value = scpc_b2_variable_fee
    dest_sheet['F8'].value = scpc_b2_variable_fee
    dest_sheet['F9'].value = scpc_b2_variable_fee
    dest_sheet['F10'].value = scpc_b2_variable_fee
    dest_sheet['F11'].value = scpc_b2_variable_fee
    dest_sheet['F12'].value = scpc_b2_variable_fee
    dest_sheet['F13'].value = scpc_b2_variable_fee
    dest_sheet['F14'].value = scpc_b2_variable_fee
    dest_sheet['F15'].value = scpc_b2_variable_fee
    dest_sheet['F16'].value = scpc_b2_variable_fee
    dest_sheet['F17'].value = scpc_b2_variable_fee
    dest_sheet['F18'].value = scpc_b2_variable_fee
    dest_sheet['F19'].value = scpc_b2_variable_fee
    dest_sheet['F20'].value = scpc_b2_variable_fee
    dest_sheet['F21'].value = scpc_b2_variable_fee
    dest_sheet['F22'].value = scpc_b2_variable_fee
    dest_sheet['F23'].value = scpc_b2_variable_fee
    dest_sheet['F24'].value = scpc_b2_variable_fee
    dest_sheet['F25'].value = scpc_b2_variable_fee

    dest_sheet['G2'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G3'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G4'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G5'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G6'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G7'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G8'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G9'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G10'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G11'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G12'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G13'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G14'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G15'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G16'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G17'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G18'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G19'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G20'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G21'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G22'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G23'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G24'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2
    dest_sheet['G25'].value = (scpc_b1_fixed_fee + scpc_b2_fixed_fee)/2

    dest_sheet['H2'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H3'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H4'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H5'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H6'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H7'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H8'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H9'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H10'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H11'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H12'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H13'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H14'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H15'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H16'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H17'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H18'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H19'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H20'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H21'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H22'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H23'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H24'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2
    dest_sheet['H25'].value = (scpc_b1_variable_fee + scpc_b2_variable_fee)/2

    dest_sheet['J2'].value = kspc_b1_fixed_fee
    dest_sheet['J3'].value = kspc_b1_fixed_fee
    dest_sheet['J4'].value = kspc_b1_fixed_fee
    dest_sheet['J5'].value = kspc_b1_fixed_fee
    dest_sheet['J6'].value = kspc_b1_fixed_fee
    dest_sheet['J7'].value = kspc_b1_fixed_fee
    dest_sheet['J8'].value = kspc_b1_fixed_fee
    dest_sheet['J9'].value = kspc_b1_fixed_fee
    dest_sheet['J10'].value = kspc_b1_fixed_fee
    dest_sheet['J11'].value = kspc_b1_fixed_fee
    dest_sheet['J12'].value = kspc_b1_fixed_fee
    dest_sheet['J13'].value = kspc_b1_fixed_fee
    dest_sheet['J14'].value = kspc_b1_fixed_fee
    dest_sheet['J15'].value = kspc_b1_fixed_fee
    dest_sheet['J16'].value = kspc_b1_fixed_fee
    dest_sheet['J17'].value = kspc_b1_fixed_fee
    dest_sheet['J18'].value = kspc_b1_fixed_fee
    dest_sheet['J19'].value = kspc_b1_fixed_fee
    dest_sheet['J20'].value = kspc_b1_fixed_fee
    dest_sheet['J21'].value = kspc_b1_fixed_fee
    dest_sheet['J22'].value = kspc_b1_fixed_fee
    dest_sheet['J23'].value = kspc_b1_fixed_fee
    dest_sheet['J24'].value = kspc_b1_fixed_fee
    dest_sheet['J25'].value = kspc_b1_fixed_fee

    dest_sheet['K2'].value = kspc_b1_variable_fee
    dest_sheet['K3'].value = kspc_b1_variable_fee
    dest_sheet['K4'].value = kspc_b1_variable_fee
    dest_sheet['K5'].value = kspc_b1_variable_fee
    dest_sheet['K6'].value = kspc_b1_variable_fee
    dest_sheet['K7'].value = kspc_b1_variable_fee
    dest_sheet['K8'].value = kspc_b1_variable_fee
    dest_sheet['K9'].value = kspc_b1_variable_fee
    dest_sheet['K10'].value = kspc_b1_variable_fee
    dest_sheet['K11'].value = kspc_b1_variable_fee
    dest_sheet['K12'].value = kspc_b1_variable_fee
    dest_sheet['K13'].value = kspc_b1_variable_fee
    dest_sheet['K14'].value = kspc_b1_variable_fee
    dest_sheet['K15'].value = kspc_b1_variable_fee
    dest_sheet['K16'].value = kspc_b1_variable_fee
    dest_sheet['K17'].value = kspc_b1_variable_fee
    dest_sheet['K18'].value = kspc_b1_variable_fee
    dest_sheet['K19'].value = kspc_b1_variable_fee
    dest_sheet['K20'].value = kspc_b1_variable_fee
    dest_sheet['K21'].value = kspc_b1_variable_fee
    dest_sheet['K22'].value = kspc_b1_variable_fee
    dest_sheet['K23'].value = kspc_b1_variable_fee
    dest_sheet['K24'].value = kspc_b1_variable_fee
    dest_sheet['K25'].value = kspc_b1_variable_fee

    dest_sheet['L2'].value = kspc_b2_fixed_fee
    dest_sheet['L3'].value = kspc_b2_fixed_fee
    dest_sheet['L4'].value = kspc_b2_fixed_fee
    dest_sheet['L5'].value = kspc_b2_fixed_fee
    dest_sheet['L6'].value = kspc_b2_fixed_fee
    dest_sheet['L7'].value = kspc_b2_fixed_fee
    dest_sheet['L8'].value = kspc_b2_fixed_fee
    dest_sheet['L9'].value = kspc_b2_fixed_fee
    dest_sheet['L10'].value = kspc_b2_fixed_fee
    dest_sheet['L11'].value = kspc_b2_fixed_fee
    dest_sheet['L12'].value = kspc_b2_fixed_fee
    dest_sheet['L13'].value = kspc_b2_fixed_fee
    dest_sheet['L14'].value = kspc_b2_fixed_fee
    dest_sheet['L15'].value = kspc_b2_fixed_fee
    dest_sheet['L16'].value = kspc_b2_fixed_fee
    dest_sheet['L17'].value = kspc_b2_fixed_fee
    dest_sheet['L18'].value = kspc_b2_fixed_fee
    dest_sheet['L19'].value = kspc_b2_fixed_fee
    dest_sheet['L20'].value = kspc_b2_fixed_fee
    dest_sheet['L21'].value = kspc_b2_fixed_fee
    dest_sheet['L22'].value = kspc_b2_fixed_fee
    dest_sheet['L23'].value = kspc_b2_fixed_fee
    dest_sheet['L24'].value = kspc_b2_fixed_fee
    dest_sheet['L25'].value = kspc_b2_fixed_fee

    dest_sheet['M2'].value = kspc_b2_variable_fee
    dest_sheet['M3'].value = kspc_b2_variable_fee
    dest_sheet['M4'].value = kspc_b2_variable_fee
    dest_sheet['M5'].value = kspc_b2_variable_fee
    dest_sheet['M6'].value = kspc_b2_variable_fee
    dest_sheet['M7'].value = kspc_b2_variable_fee
    dest_sheet['M8'].value = kspc_b2_variable_fee
    dest_sheet['M9'].value = kspc_b2_variable_fee
    dest_sheet['M10'].value = kspc_b2_variable_fee
    dest_sheet['M11'].value = kspc_b2_variable_fee
    dest_sheet['M12'].value = kspc_b2_variable_fee
    dest_sheet['M13'].value = kspc_b2_variable_fee
    dest_sheet['M14'].value = kspc_b2_variable_fee
    dest_sheet['M15'].value = kspc_b2_variable_fee
    dest_sheet['M16'].value = kspc_b2_variable_fee
    dest_sheet['M17'].value = kspc_b2_variable_fee
    dest_sheet['M18'].value = kspc_b2_variable_fee
    dest_sheet['M19'].value = kspc_b2_variable_fee
    dest_sheet['M20'].value = kspc_b2_variable_fee
    dest_sheet['M21'].value = kspc_b2_variable_fee
    dest_sheet['M22'].value = kspc_b2_variable_fee
    dest_sheet['M23'].value = kspc_b2_variable_fee
    dest_sheet['M24'].value = kspc_b2_variable_fee
    dest_sheet['M25'].value = kspc_b2_variable_fee

    dest_sheet['N2'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N3'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N4'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N5'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N6'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N7'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N8'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N9'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N10'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N11'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N12'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N13'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N14'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N15'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N16'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N17'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N18'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N19'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N20'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N21'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N22'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N23'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N24'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2
    dest_sheet['N25'].value = (kspc_b1_fixed_fee + kspc_b2_fixed_fee)/2

    dest_sheet['O2'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O3'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O4'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O5'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O6'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O7'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O8'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O9'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O10'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O11'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O12'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O13'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O14'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O15'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O16'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O17'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O18'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O19'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O20'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O21'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O22'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O23'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O24'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2
    dest_sheet['O25'].value = (kspc_b1_variable_fee + kspc_b2_variable_fee)/2

    dest_sheet['Q2'].value = edc_fixed_fee
    dest_sheet['Q3'].value = edc_fixed_fee
    dest_sheet['Q4'].value = edc_fixed_fee
    dest_sheet['Q5'].value = edc_fixed_fee
    dest_sheet['Q6'].value = edc_fixed_fee
    dest_sheet['Q7'].value = edc_fixed_fee
    dest_sheet['Q8'].value = edc_fixed_fee
    dest_sheet['Q9'].value = edc_fixed_fee
    dest_sheet['Q10'].value = edc_fixed_fee
    dest_sheet['Q11'].value = edc_fixed_fee
    dest_sheet['Q12'].value = edc_fixed_fee
    dest_sheet['Q13'].value = edc_fixed_fee
    dest_sheet['Q14'].value = edc_fixed_fee
    dest_sheet['Q15'].value = edc_fixed_fee
    dest_sheet['Q16'].value = edc_fixed_fee
    dest_sheet['Q17'].value = edc_fixed_fee
    dest_sheet['Q18'].value = edc_fixed_fee
    dest_sheet['Q19'].value = edc_fixed_fee
    dest_sheet['Q20'].value = edc_fixed_fee
    dest_sheet['Q21'].value = edc_fixed_fee
    dest_sheet['Q22'].value = edc_fixed_fee
    dest_sheet['Q23'].value = edc_fixed_fee
    dest_sheet['Q24'].value = edc_fixed_fee
    dest_sheet['Q25'].value = edc_fixed_fee

    dest_sheet['R2'].value = edc_variable_fee
    dest_sheet['R3'].value = edc_variable_fee
    dest_sheet['R4'].value = edc_variable_fee
    dest_sheet['R5'].value = edc_variable_fee
    dest_sheet['R6'].value = edc_variable_fee
    dest_sheet['R7'].value = edc_variable_fee
    dest_sheet['R8'].value = edc_variable_fee
    dest_sheet['R9'].value = edc_variable_fee
    dest_sheet['R10'].value = edc_variable_fee
    dest_sheet['R11'].value = edc_variable_fee
    dest_sheet['R12'].value = edc_variable_fee
    dest_sheet['R13'].value = edc_variable_fee
    dest_sheet['R14'].value = edc_variable_fee
    dest_sheet['R15'].value = edc_variable_fee
    dest_sheet['R16'].value = edc_variable_fee
    dest_sheet['R17'].value = edc_variable_fee
    dest_sheet['R18'].value = edc_variable_fee
    dest_sheet['R19'].value = edc_variable_fee
    dest_sheet['R20'].value = edc_variable_fee
    dest_sheet['R21'].value = edc_variable_fee
    dest_sheet['R22'].value = edc_variable_fee
    dest_sheet['R23'].value = edc_variable_fee
    dest_sheet['R24'].value = edc_variable_fee
    dest_sheet['R25'].value = edc_variable_fee

    dest_wb.save(destination_file_path)

# TOTAL SS LOAD 
def total_ss_load():

    total_ss_load_path = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load.xlsx"
    destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\current_rate.xlsx'
    dest_wb = load_workbook(destination_file_path)
    dest_sheet = dest_wb['Sheet']

    # Load source workbook
    source_wb_2 = load_workbook(total_ss_load_path)  
    # Sheet name of the source workbook
    source_sheet_2 = source_wb_2['Sheet1']

    # Define source and destination ranges
    source_ranges_2 = [('K2', 'K25')]
    dest_ranges_2 = [('S2', 'S25')]

    # Copy data from source to destination based on specified ranges
    for i in range(len(source_ranges_2)):
        copy_values(source_sheet_2, dest_sheet, source_ranges_2[i], dest_ranges_2[i])

    # Save the destination workbook
    dest_wb.save(destination_file_path)

# CONTESTABLE ENERGY
def contestable_energy():
    CE_filepath = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\contestable_energy.xlsx'
    destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\current_rate.xlsx'
    dest_wb = load_workbook(destination_file_path)
    dest_sheet = dest_wb['Sheet']

    # Load source workbook
    source_wb_CE = load_workbook(CE_filepath)
    # Sheet name of the source workbook
    source_sheet_CE = source_wb_CE['Sheet']

    # Define source and destination ranges
    source_ranges_CE = [('B2', 'B25')]
    dest_ranges_CE = [('T2', 'T25')]

    # Copy data from source to destination based on specified ranges
    for i in range(len(source_ranges_CE)):
        copy_values(source_sheet_CE, dest_sheet, source_ranges_CE[i], dest_ranges_CE[i])

    # Save the destination workbook
    dest_wb.save(destination_file_path)

def find_total():
    directory = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\MORE Trading Node'
    now = datetime.now()
    # Example folder name: 20240711
    date_today = now.strftime("%Y%m%d")
    contents = os.listdir(directory)
    for item in contents:
        if item == date_today:
            folder_path = os.path.join(directory, item) 
            if os.path.isdir(folder_path):
                files_in_folder = os.listdir(folder_path) # MPI_LMP_DIPC_202407110000, etc.
                current_hour = now.strftime("%H")

                if current_hour == '01':

                    try: 
                        first_file_path = os.path.join(folder_path, files_in_folder[1])
                        second_file_path = os.path.join(folder_path, files_in_folder[2])
                        third_file_path = os.path.join(folder_path, files_in_folder[3])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[4])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[5])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[6])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[7])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[8])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[9])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[10])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[11])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[12])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                    except:
                        final_total = 0
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '02':

                    try:

                        first_file_path = os.path.join(folder_path, files_in_folder[13])
                        second_file_path = os.path.join(folder_path, files_in_folder[14])
                        third_file_path = os.path.join(folder_path, files_in_folder[15])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[16])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[17])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[18])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[19])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[20])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[21])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[22])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[23])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[24])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                        
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                            
                        first_file_path = os.path.join(folder_path, files_in_folder[1])
                        second_file_path = os.path.join(folder_path, files_in_folder[2])
                        third_file_path = os.path.join(folder_path, files_in_folder[3])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[4])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[5])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[6])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[7])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[8])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[9])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[10])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[11])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[12])
                        
                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '03':
                    try: 
                        first_file_path = os.path.join(folder_path, files_in_folder[25])
                        second_file_path = os.path.join(folder_path, files_in_folder[26])
                        third_file_path = os.path.join(folder_path, files_in_folder[27])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[28])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[29])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[30])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[31])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[32])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[33])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[34])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[35])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[36])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                        
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[13])
                        second_file_path = os.path.join(folder_path, files_in_folder[14])
                        third_file_path = os.path.join(folder_path, files_in_folder[15])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[16])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[17])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[18])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[19])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[20])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[21])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[22])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[23])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[24])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_1 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_1 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_1 + file_08STBAR_T1L1_1)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                        
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '04':
                    try: 
                        first_file_path = os.path.join(folder_path, files_in_folder[37])
                        second_file_path = os.path.join(folder_path, files_in_folder[38])
                        third_file_path = os.path.join(folder_path, files_in_folder[39])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[40])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[41])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[42])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[43])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[44])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[45])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[46])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[47])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[48])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:

                        first_file_path = os.path.join(folder_path, files_in_folder[25])
                        second_file_path = os.path.join(folder_path, files_in_folder[26])
                        third_file_path = os.path.join(folder_path, files_in_folder[27])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[28])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[29])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[30])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[31])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[32])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[33])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[34])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[35])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[36])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                        
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '05':

                    try: 
                        first_file_path = os.path.join(folder_path, files_in_folder[49])
                        second_file_path = os.path.join(folder_path, files_in_folder[50])
                        third_file_path = os.path.join(folder_path, files_in_folder[51])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[52])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[53])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[54])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[55])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[56])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[57])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[58])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[59])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[60])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                    except:

                        first_file_path = os.path.join(folder_path, files_in_folder[37])
                        second_file_path = os.path.join(folder_path, files_in_folder[38])
                        third_file_path = os.path.join(folder_path, files_in_folder[39])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[40])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[41])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[42])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[43])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[44])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[45])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[46])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[47])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[48])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '06':

                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[61])
                        second_file_path = os.path.join(folder_path, files_in_folder[62])
                        third_file_path = os.path.join(folder_path, files_in_folder[63])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[64])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[65])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[66])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[67])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[68])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[69])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[70])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[71])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[72])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[49])
                        second_file_path = os.path.join(folder_path, files_in_folder[50])
                        third_file_path = os.path.join(folder_path, files_in_folder[51])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[52])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[53])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[54])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[55])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[56])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[57])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[58])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[59])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[60])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '07':

                    try: 

                        first_file_path = os.path.join(folder_path, files_in_folder[73])
                        second_file_path = os.path.join(folder_path, files_in_folder[74])
                        third_file_path = os.path.join(folder_path, files_in_folder[75])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[76])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[77])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[78])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[79])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[80])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[81])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[82])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[83])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[84])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[61])
                        second_file_path = os.path.join(folder_path, files_in_folder[62])
                        third_file_path = os.path.join(folder_path, files_in_folder[63])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[64])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[65])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[66])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[67])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[68])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[69])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[70])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[71])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[72])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '08':

                    try: 

                        first_file_path = os.path.join(folder_path, files_in_folder[85])
                        second_file_path = os.path.join(folder_path, files_in_folder[86])
                        third_file_path = os.path.join(folder_path, files_in_folder[87])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[88])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[89])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[90])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[91])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[92])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[93])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[94])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[95])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[96])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_1 + file_08STBAR_T1L1_1)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:

                        first_file_path = os.path.join(folder_path, files_in_folder[73])
                        second_file_path = os.path.join(folder_path, files_in_folder[74])
                        third_file_path = os.path.join(folder_path, files_in_folder[75])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[76])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[77])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[78])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[79])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[80])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[81])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[82])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[83])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[84])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '09':

                    try: 
                        first_file_path = os.path.join(folder_path, files_in_folder[97])
                        second_file_path = os.path.join(folder_path, files_in_folder[98])
                        third_file_path = os.path.join(folder_path, files_in_folder[99])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[100])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[101])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[102])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[103])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[104])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[105])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[106])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[107])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[108])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:

                        first_file_path = os.path.join(folder_path, files_in_folder[85])
                        second_file_path = os.path.join(folder_path, files_in_folder[86])
                        third_file_path = os.path.join(folder_path, files_in_folder[87])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[88])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[89])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[90])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[91])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[92])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[93])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[94])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[95])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[96])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '10':

                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[109])
                        second_file_path = os.path.join(folder_path, files_in_folder[110])
                        third_file_path = os.path.join(folder_path, files_in_folder[111])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[112])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[113])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[114])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[115])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[116])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[117])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[118])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[119])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[120])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                    
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_file_path = os.path.join(folder_path, files_in_folder[114])
                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[97])
                        second_file_path = os.path.join(folder_path, files_in_folder[98])
                        third_file_path = os.path.join(folder_path, files_in_folder[99])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[100])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[101])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[102])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[103])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[104])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[105])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[106])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[107])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[108])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                elif current_hour == '11':
                    try: 
                        first_file_path = os.path.join(folder_path, files_in_folder[121])
                        second_file_path = os.path.join(folder_path, files_in_folder[122])
                        third_file_path = os.path.join(folder_path, files_in_folder[123])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[124])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[125])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[126])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[127])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[128])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[129])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[130])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[131])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[132])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3
                        
                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[109])
                        second_file_path = os.path.join(folder_path, files_in_folder[110])
                        third_file_path = os.path.join(folder_path, files_in_folder[111])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[112])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[113])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[114])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[115])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[116])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[117])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[118])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[119])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[120])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                    
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3
                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_file_path = os.path.join(folder_path, files_in_folder[114])
                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '12':
                    try: 
                        first_file_path = os.path.join(folder_path, files_in_folder[133])
                        second_file_path = os.path.join(folder_path, files_in_folder[134])
                        third_file_path = os.path.join(folder_path, files_in_folder[135])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[136])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[137])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[138])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[139])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[140])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[141])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[142])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[143])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[144])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[121])
                        second_file_path = os.path.join(folder_path, files_in_folder[122])
                        third_file_path = os.path.join(folder_path, files_in_folder[123])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[124])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[125])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[126])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[127])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[128])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[129])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[130])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[131])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[132])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '13':

                    try: 

                        first_file_path = os.path.join(folder_path, files_in_folder[145])
                        second_file_path = os.path.join(folder_path, files_in_folder[146])
                        third_file_path = os.path.join(folder_path, files_in_folder[147])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[148])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[149])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[150])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[151])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[152])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[153])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[154])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[155])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[156])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                    
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[133])
                        second_file_path = os.path.join(folder_path, files_in_folder[134])
                        third_file_path = os.path.join(folder_path, files_in_folder[135])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[136])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[137])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[138])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[139])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[140])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[141])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[142])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[143])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[144])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '14':
                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[157])
                        second_file_path = os.path.join(folder_path, files_in_folder[158])
                        third_file_path = os.path.join(folder_path, files_in_folder[159])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[160])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[161])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[162])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[163])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[164])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[165])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[166])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[167])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[168])
                    
                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_file_path = os.path.join(folder_path, files_in_folder[168])
                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[145])
                        second_file_path = os.path.join(folder_path, files_in_folder[146])
                        third_file_path = os.path.join(folder_path, files_in_folder[147])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[148])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[149])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[150])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[151])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[152])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[153])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[154])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[155])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[156])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                    
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '15':

                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[169])
                        second_file_path = os.path.join(folder_path, files_in_folder[170])
                        third_file_path = os.path.join(folder_path, files_in_folder[171])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[172])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[173])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[174])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[175])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[176])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[177])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[178])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[179])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[180])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                    
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:

                        first_file_path = os.path.join(folder_path, files_in_folder[157])
                        second_file_path = os.path.join(folder_path, files_in_folder[158])
                        third_file_path = os.path.join(folder_path, files_in_folder[159])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[160])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[161])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[162])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[163])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[164])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[165])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[166])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[167])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[168])
                    
                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_file_path = os.path.join(folder_path, files_in_folder[168])
                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
            
                elif current_hour == '16':
                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[181])
                        second_file_path = os.path.join(folder_path, files_in_folder[182])
                        third_file_path = os.path.join(folder_path, files_in_folder[183])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[184])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[185])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[186])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[187])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[188])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[189])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[190])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[191])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[192])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[169])
                        second_file_path = os.path.join(folder_path, files_in_folder[170])
                        third_file_path = os.path.join(folder_path, files_in_folder[171])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[172])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[173])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[174])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[175])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[176])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[177])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[178])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[179])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[180])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                    
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '17':
                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[193])
                        second_file_path = os.path.join(folder_path, files_in_folder[194])
                        third_file_path = os.path.join(folder_path, files_in_folder[195])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[196])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[197])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[198])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[199])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[200])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[201])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[202])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[203])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[204])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[181])
                        second_file_path = os.path.join(folder_path, files_in_folder[182])
                        third_file_path = os.path.join(folder_path, files_in_folder[183])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[184])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[185])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[186])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[187])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[188])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[189])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[190])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[191])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[192])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                
                elif current_hour == '18':
                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[205])
                        second_file_path = os.path.join(folder_path, files_in_folder[206])
                        third_file_path = os.path.join(folder_path, files_in_folder[207])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[208])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[209])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[210])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[211])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[212])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[213])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[214])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[215])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[216])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                    
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                    
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[193])
                        second_file_path = os.path.join(folder_path, files_in_folder[194])
                        third_file_path = os.path.join(folder_path, files_in_folder[195])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[196])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[197])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[198])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[199])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[200])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[201])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[202])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[203])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[204])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '19':
                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[217])
                        second_file_path = os.path.join(folder_path, files_in_folder[218])
                        third_file_path = os.path.join(folder_path, files_in_folder[219])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[220])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[221])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[222])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[223])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[224])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[225])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[226])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[227])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[228])
                    
                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:

                        first_file_path = os.path.join(folder_path, files_in_folder[205])
                        second_file_path = os.path.join(folder_path, files_in_folder[206])
                        third_file_path = os.path.join(folder_path, files_in_folder[207])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[208])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[209])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[210])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[211])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[212])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[213])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[214])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[215])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[216])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                    
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                    
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '20':

                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[229])
                        second_file_path = os.path.join(folder_path, files_in_folder[230])
                        third_file_path = os.path.join(folder_path, files_in_folder[231])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[232])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[233])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[234])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[235])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[236])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[237])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[238])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[239])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[240])
                    
                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[217])
                        second_file_path = os.path.join(folder_path, files_in_folder[218])
                        third_file_path = os.path.join(folder_path, files_in_folder[219])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[220])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[221])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[222])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[223])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[224])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[225])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[226])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[227])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[228])
                    
                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '21': 
                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[241])
                        second_file_path = os.path.join(folder_path, files_in_folder[242])
                        third_file_path = os.path.join(folder_path, files_in_folder[243])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[244])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[245])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[246])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[247])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[248])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[249])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[250])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[251])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[252])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[229])
                        second_file_path = os.path.join(folder_path, files_in_folder[230])
                        third_file_path = os.path.join(folder_path, files_in_folder[231])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[232])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[233])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[234])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[235])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[236])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[237])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[238])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[239])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[240])
                    
                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                elif current_hour == '22':
                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[253])
                        second_file_path = os.path.join(folder_path, files_in_folder[254])
                        third_file_path = os.path.join(folder_path, files_in_folder[255])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[256])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[257])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[258])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[259])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[260])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[261])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[262])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[263])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[264])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:

                        first_file_path = os.path.join(folder_path, files_in_folder[241])
                        second_file_path = os.path.join(folder_path, files_in_folder[242])
                        third_file_path = os.path.join(folder_path, files_in_folder[243])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[244])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[245])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[246])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[247])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[248])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[249])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[250])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[251])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[252])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '23':
                    try:
                        first_file_path = os.path.join(folder_path, files_in_folder[265])
                        second_file_path = os.path.join(folder_path, files_in_folder[266])
                        third_file_path = os.path.join(folder_path, files_in_folder[267])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[268])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[269])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[270])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[271])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[272])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[273])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[274])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[275])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[276])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total
                    
                    except:
                        
                        first_file_path = os.path.join(folder_path, files_in_folder[253])
                        second_file_path = os.path.join(folder_path, files_in_folder[254])
                        third_file_path = os.path.join(folder_path, files_in_folder[255])
                        fourth_file_path = os.path.join(folder_path, files_in_folder[256])
                        fifth_file_path = os.path.join(folder_path, files_in_folder[257])
                        sixth_file_path = os.path.join(folder_path, files_in_folder[258])
                        seventh_file_path = os.path.join(folder_path, files_in_folder[259])
                        eighth_file_path = os.path.join(folder_path, files_in_folder[260])
                        ninth_file_path = os.path.join(folder_path, files_in_folder[261])
                        tenth_file_path = os.path.join(folder_path, files_in_folder[262])
                        eleventh_file_path = os.path.join(folder_path, files_in_folder[263])
                        twelvth_file_path = os.path.join(folder_path, files_in_folder[264])

                        df = pd.read_csv(first_file_path)

                        search_value_1 = '08PEDC_T1L1'
                        matching_row_1 = df[df['RESOURCE_NAME'] == search_value_1]
                        file_08PEDC_T1L1_1 = matching_row_1['DIPC_PRICE'].values[0]

                        search_value_2 = '08PEDC_T1L2'
                        matching_row_2 = df[df['RESOURCE_NAME'] == search_value_2]
                        file_08PEDC_T1L2_2 = matching_row_2['DIPC_PRICE'].values[0]

                        search_value_3 = '08STBAR_T1L1'
                        matching_row_3 = df[df['RESOURCE_NAME'] == search_value_3]
                        file_08STBAR_T1L1_3 = matching_row_3['DIPC_PRICE'].values[0]

                        average_1 = (file_08PEDC_T1L1_1 + file_08PEDC_T1L2_2 + file_08STBAR_T1L1_3)/3

                        second_df = pd.read_csv(second_file_path)

                        search_value_4 = '08PEDC_T1L1'
                        matching_row_4 = second_df[second_df['RESOURCE_NAME'] == search_value_4]
                        file_08PEDC_T1L1_4 = matching_row_4['DIPC_PRICE'].values[0]
                        
                        search_value_5 = '08PEDC_T1L2'
                        matching_row_5 = second_df[second_df['RESOURCE_NAME'] == search_value_5]
                        file_08PEDC_T1L2_5 = matching_row_5['DIPC_PRICE'].values[0]

                        search_value_6 = '08STBAR_T1L1'
                        matching_row_6 = second_df[second_df['RESOURCE_NAME'] == search_value_6]
                        file_08STBAR_T1L1_6 = matching_row_6['DIPC_PRICE'].values[0]

                        average_2 = (file_08PEDC_T1L1_4 + file_08PEDC_T1L2_5 + file_08STBAR_T1L1_6)/3

                        third_df = pd.read_csv(third_file_path)

                        search_value_7 = '08PEDC_T1L1'
                        matching_row_7 = third_df[third_df['RESOURCE_NAME'] == search_value_7]
                        file_08PEDC_T1L1_7 = matching_row_7['DIPC_PRICE'].values[0]

                        search_value_8 = '08PEDC_T1L2'
                        matching_row_8 = third_df[third_df['RESOURCE_NAME'] == search_value_8]
                        file_08PEDC_T1L2_8 = matching_row_8['DIPC_PRICE'].values[0]

                        search_value_9 = '08STBAR_T1L1'
                        matching_row_9 = third_df[third_df['RESOURCE_NAME'] == search_value_9]
                        file_08STBAR_T1L1_9 = matching_row_9['DIPC_PRICE'].values[0]

                        average_3 = (file_08PEDC_T1L1_7 + file_08PEDC_T1L2_8 + file_08STBAR_T1L1_9)/3

                        fourth_df = pd.read_csv(fourth_file_path)

                        search_value_10 = '08PEDC_T1L1'
                        matching_row_10 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_10]
                        file_08PEDC_T1L1_10 = matching_row_10['DIPC_PRICE'].values[0]

                        search_value_11 = '08PEDC_T1L2'
                        matching_row_11 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_11]
                        file_08PEDC_T1L2_11 = matching_row_11['DIPC_PRICE'].values[0]

                        search_value_12 = '08STBAR_T1L1'
                        matching_row_12 = fourth_df[fourth_df['RESOURCE_NAME'] == search_value_12]
                        file_08STBAR_T1L1_12 = matching_row_12['DIPC_PRICE'].values[0]

                        average_4 = (file_08PEDC_T1L1_10 + file_08PEDC_T1L2_11 + file_08STBAR_T1L1_12)/3

                        fifth_df = pd.read_csv(fifth_file_path)

                        search_value_13 = '08PEDC_T1L1'
                        matching_row_13 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_13]
                        file_08PEDC_T1L1_13 = matching_row_13['DIPC_PRICE'].values[0]

                        search_value_14 = '08PEDC_T1L2'
                        matching_row_14 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_14]
                        file_08PEDC_T1L2_14 = matching_row_14['DIPC_PRICE'].values[0]

                        search_value_15 = '08STBAR_T1L1'
                        matching_row_15 = fifth_df[fifth_df['RESOURCE_NAME'] == search_value_15]
                        file_08STBAR_T1L1_15 = matching_row_15['DIPC_PRICE'].values[0]

                        average_5 = (file_08PEDC_T1L1_13 + file_08PEDC_T1L2_14 + file_08STBAR_T1L1_15)/3

                        sixth_df = pd.read_csv(sixth_file_path)

                        search_value_16 = '08PEDC_T1L1'
                        matching_row_16 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_16]
                        file_08PEDC_T1L1_16 = matching_row_16['DIPC_PRICE'].values[0]

                        search_value_17 = '08PEDC_T1L2'
                        matching_row_17 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_17]
                        file_08PEDC_T1L2_17 = matching_row_17['DIPC_PRICE'].values[0]

                        search_value_18 = '08STBAR_T1L1'
                        matching_row_18 = sixth_df[sixth_df['RESOURCE_NAME'] == search_value_18]
                        file_08STBAR_T1L1_18 = matching_row_18['DIPC_PRICE'].values[0]

                        average_6 = (file_08PEDC_T1L1_16 + file_08PEDC_T1L2_17 + file_08STBAR_T1L1_18)/3

                        seventh_df = pd.read_csv(seventh_file_path)

                        search_value_19 = '08PEDC_T1L1'
                        matching_row_19 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_19]
                        file_08PEDC_T1L1_19 = matching_row_19['DIPC_PRICE'].values[0]

                        search_value_20 = '08PEDC_T1L2'
                        matching_row_20 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_20]
                        file_08PEDC_T1L2_20 = matching_row_20['DIPC_PRICE'].values[0]

                        search_value_21 = '08STBAR_T1L1'
                        matching_row_21 = seventh_df[seventh_df['RESOURCE_NAME'] == search_value_21]
                        file_08STBAR_T1L1_21 = matching_row_21['DIPC_PRICE'].values[0]

                        average_7 = (file_08PEDC_T1L1_19 + file_08PEDC_T1L2_20 + file_08STBAR_T1L1_21)/3

                        eighth_df = pd.read_csv(eighth_file_path)

                        search_value_22 = '08PEDC_T1L1'
                        matching_row_22 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_22]
                        file_08PEDC_T1L1_22 = matching_row_22['DIPC_PRICE'].values[0]

                        search_value_23 = '08PEDC_T1L2'
                        matching_row_23 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_23]
                        file_08PEDC_T1L2_23 = matching_row_23['DIPC_PRICE'].values[0]

                        search_value_24 = '08STBAR_T1L1'
                        matching_row_24 = eighth_df[eighth_df['RESOURCE_NAME'] == search_value_24]
                        file_08STBAR_T1L1_24 = matching_row_24['DIPC_PRICE'].values[0]

                        average_8 = (file_08PEDC_T1L1_22 + file_08PEDC_T1L2_23 + file_08STBAR_T1L1_24)/3
                            
                        ninth_df = pd.read_csv(ninth_file_path)

                        search_value_25 = '08PEDC_T1L1'
                        matching_row_25 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_25]
                        file_08PEDC_T1L1_25 = matching_row_25['DIPC_PRICE'].values[0]

                        search_value_26 = '08PEDC_T1L2'
                        matching_row_26 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_26]
                        file_08PEDC_T1L2_26 = matching_row_26['DIPC_PRICE'].values[0]

                        search_value_27 = '08STBAR_T1L1'
                        matching_row_27 = ninth_df[ninth_df['RESOURCE_NAME'] == search_value_27]
                        file_08STBAR_T1L1_27 = matching_row_27['DIPC_PRICE'].values[0]

                        average_9 = (file_08PEDC_T1L1_25 + file_08PEDC_T1L2_26 + file_08STBAR_T1L1_27)/3

                        tenth_df = pd.read_csv(tenth_file_path)

                        search_value_28 = '08PEDC_T1L1'
                        matching_row_28 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_28]
                        file_08PEDC_T1L1_28 = matching_row_28['DIPC_PRICE'].values[0]

                        search_value_29 = '08PEDC_T1L2'
                        matching_row_29 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_29]
                        file_08PEDC_T1L2_29 = matching_row_29['DIPC_PRICE'].values[0]

                        search_value_30 = '08STBAR_T1L1'
                        matching_row_30 = tenth_df[tenth_df['RESOURCE_NAME'] == search_value_30]
                        file_08STBAR_T1L1_30 = matching_row_30['DIPC_PRICE'].values[0]

                        average_10 = (file_08PEDC_T1L1_28 + file_08PEDC_T1L2_29 + file_08STBAR_T1L1_30)/3

                        eleventh_df = pd.read_csv(eleventh_file_path)

                        search_value_31 = '08PEDC_T1L1'
                        matching_row_31 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_31]
                        file_08PEDC_T1L1_31 = matching_row_31['DIPC_PRICE'].values[0]

                        search_value_32 = '08PEDC_T1L2'
                        matching_row_32 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_32]
                        file_08PEDC_T1L2_32 = matching_row_32['DIPC_PRICE'].values[0]

                        search_value_33 = '08STBAR_T1L1'
                        matching_row_33 = eleventh_df[eleventh_df['RESOURCE_NAME'] == search_value_33]
                        file_08STBAR_T1L1_33 = matching_row_33['DIPC_PRICE'].values[0]

                        average_11 = (file_08PEDC_T1L1_31 + file_08PEDC_T1L2_32 + file_08STBAR_T1L1_33)/3

                        twelvth_df = pd.read_csv(twelvth_file_path)

                        search_value_34 = '08PEDC_T1L1'
                        matching_row_34 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_34]
                        file_08PEDC_T1L1_34 = matching_row_34['DIPC_PRICE'].values[0]
                        
                        search_value_35 = '08PEDC_T1L2'
                        matching_row_35 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_35]
                        file_08PEDC_T1L2_35 = matching_row_35['DIPC_PRICE'].values[0]

                        search_value_36 = '08STBAR_T1L1'
                        matching_row_36 = twelvth_df[twelvth_df['RESOURCE_NAME'] == search_value_36]
                        file_08STBAR_T1L1_36 = matching_row_36['DIPC_PRICE'].values[0]

                        average_12 = (file_08PEDC_T1L1_34 + file_08PEDC_T1L2_35 + file_08STBAR_T1L1_36)/3

                        final_total = (average_1 + average_2 + average_3 + average_4 + average_5 + average_6 + average_7 + average_8 + average_9 + average_10 + average_11 + average_12)/12
                        float_final_total = float(final_total)
                        return float_final_total

                elif current_hour == '00':
                        final_total = 0
                        float_final_total = float(final_total)
                        return float_final_total
                    
# STORE WESM RATE IN THE EXCEL FILE
def find_total_2():
    destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\current_rate.xlsx'
    dest_wb = load_workbook(destination_file_path)
    dest_sheet = dest_wb['Sheet']
    
    c = datetime.now()
    hour_now = c.strftime('%H')
    if hour_now == '01':
        wesm_rate_var = find_total()
        # cell N2
        dest_sheet['V2'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '02':
        wesm_rate_var = find_total()
        # cell N3
        dest_sheet['V3'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '03':
        wesm_rate_var = find_total()
        # cell N4
        dest_sheet['V4'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '04':
        wesm_rate_var = find_total()
        # cell N5
        dest_sheet['V5'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '05':
        wesm_rate_var = find_total()
        # cell N6
        dest_sheet['V6'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '06':
        wesm_rate_var = find_total()
        # cell N7
        dest_sheet['V7'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '07':
        wesm_rate_var = find_total()
        # cell N8
        dest_sheet['V8'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '08':
        wesm_rate_var = find_total()
        # cell N9
        dest_sheet['V9'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '09':
        wesm_rate_var = find_total()
        # cell N10
        dest_sheet['V10'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '10':
        wesm_rate_var = find_total()
        # cell N11
        dest_sheet['V11'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '11':
        wesm_rate_var = find_total()
        # cell N12
        dest_sheet['V12'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '12':
        wesm_rate_var = find_total()
        # cell N13
        dest_sheet['V13'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '13':
        wesm_rate_var = find_total()
        # cell N14
        dest_sheet['V14'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '14':
        wesm_rate_var = find_total()
        # cell N15
        dest_sheet['V15'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '15':
        wesm_rate_var = find_total()
        # cell N16
        dest_sheet['V16'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '16':
        wesm_rate_var = find_total()
        # cell N17
        dest_sheet['V17'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '17':
        wesm_rate_var = find_total()
        # cell N18
        dest_sheet['V18'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '18':
        wesm_rate_var = find_total()
        # cell N19
        dest_sheet['V19'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '19':
        wesm_rate_var = find_total()
        # cell N20
        dest_sheet['V20'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '20':
        wesm_rate_var = find_total()
        # cell N21
        dest_sheet['V21'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '21':
        wesm_rate_var = find_total()
        # cell N22
        dest_sheet['V22'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '22':
        wesm_rate_var = find_total()
        # cell N23
        dest_sheet['V23'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '23':
        wesm_rate_var = find_total()
        # cell N24
        dest_sheet['V24'] = wesm_rate_var
        dest_wb.save(destination_file_path)
    elif hour_now == '00':
        wesm_rate_var = 0
        # cell N25
        dest_sheet['V25'] = wesm_rate_var
        dest_wb.save(destination_file_path)

# Calculating the Current Rate
# Current Rate Formula = [((25000 x Fixed Fee)+(BCQSCPCxVariable Fee)) + ((20000xFixed Fee)+(BCQKSPCxVariable Fee)) + ((20000xFixed Fee)+(BCQEDCxVariable Fee)) + (TotalSSLoad-Contestable Energy-SummationBCQ) x DIPC HOURLY]/Total SS Load - Contestable Energy 

# CALCULATE CURRENT RATE AND STORE IN THE EXCEL FILE
def find_total_3():
    destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\current_rate.xlsx'
    dest_wb = load_workbook(destination_file_path)
    dest_sheet = dest_wb['Sheet']
    c = datetime.now()
    hour_now = c.strftime('%H')

    if hour_now == '01':
        # HOUR 1
        scpc_value_1 = 25000
        scpc_fixed_cost_1 = dest_sheet['G2'].value
        bcq_scpc_1 = dest_sheet['B2'].value
        scpc_variable_cost_1 = dest_sheet['H2'].value

        kspc_value_1 = 20000
        kspc_fixed_cost_1 = dest_sheet['N2'].value
        bcq_kspc_1 = dest_sheet['I2'].value
        kspc_variable_cost_1 = dest_sheet['O2'].value

        edc_value_1 = 20000
        edc_fixed_cost_1 = dest_sheet['Q2'].value
        bcq_edc_1 = dest_sheet['P2'].value
        edc_variable_cost_1 = dest_sheet['R2'].value

        total_ssload_var_1 = dest_sheet['S2'].value
        contestable_energy_var_1 = dest_sheet['T2'].value
        total_bcq_var_1 = dest_sheet['U2'].value
        wesm_rate_var_1 = (dest_sheet['V2'].value)/1000

        current_rate_formula_1 = (((scpc_value_1 * scpc_fixed_cost_1) + (bcq_scpc_1 * scpc_variable_cost_1)) + ((kspc_value_1 * kspc_fixed_cost_1) + (bcq_kspc_1 * kspc_variable_cost_1)) + ((edc_value_1 * edc_fixed_cost_1) + (bcq_edc_1 * edc_variable_cost_1)) + (total_ssload_var_1 - contestable_energy_var_1 - total_bcq_var_1) * wesm_rate_var_1)/(total_ssload_var_1 - contestable_energy_var_1)
        
        dest_sheet['W2'] = current_rate_formula_1
        dest_wb.save(destination_file_path)

    elif hour_now == '02':
        # HOUR 2
        scpc_value_2 = 25000
        scpc_fixed_cost_2 = dest_sheet['G3'].value
        bcq_scpc_2 = dest_sheet['B3'].value
        scpc_variable_cost_2 = dest_sheet['H3'].value

        kspc_value_2 = 20000
        kspc_fixed_cost_2 = dest_sheet['N3'].value
        bcq_kspc_2 = dest_sheet['I3'].value
        kspc_variable_cost_2 = dest_sheet['O3'].value

        edc_value_2 = 20000
        edc_fixed_cost_2 = dest_sheet['Q3'].value
        bcq_edc_2 = dest_sheet['P3'].value
        edc_variable_cost_2 = dest_sheet['R3'].value

        total_ssload_var_2 = dest_sheet['S3'].value
        contestable_energy_var_2 = dest_sheet['T3'].value
        total_bcq_var_2 = dest_sheet['U3'].value
        wesm_rate_var_2 = (dest_sheet['V3'].value)/1000

        current_rate_formula_2 = (((scpc_value_2 * scpc_fixed_cost_2) + (bcq_scpc_2 * scpc_variable_cost_2)) + ((kspc_value_2 * kspc_fixed_cost_2) + (bcq_kspc_2 * kspc_variable_cost_2)) + ((edc_value_2 * edc_fixed_cost_2) + (bcq_edc_2 * edc_variable_cost_2)) + (total_ssload_var_2 - contestable_energy_var_2 - total_bcq_var_2) * wesm_rate_var_2)/(total_ssload_var_2 - contestable_energy_var_2)
        
        dest_sheet['W3'] = current_rate_formula_2
        dest_wb.save(destination_file_path)

    elif hour_now == '03':
        # HOUR 3
        scpc_value_3 = 25000
        scpc_fixed_cost_3 = dest_sheet['G4'].value
        bcq_scpc_3 = dest_sheet['B4'].value
        scpc_variable_cost_3 = dest_sheet['H4'].value

        kspc_value_3 = 20000
        kspc_fixed_cost_3 = dest_sheet['N4'].value
        bcq_kspc_3 = dest_sheet['I4'].value
        kspc_variable_cost_3 = dest_sheet['O4'].value

        edc_value_3 = 20000
        edc_fixed_cost_3 = dest_sheet['Q4'].value
        bcq_edc_3 = dest_sheet['P4'].value
        edc_variable_cost_3 = dest_sheet['R4'].value

        total_ssload_var_3 = dest_sheet['S4'].value
        contestable_energy_var_3 = dest_sheet['T4'].value
        total_bcq_var_3 = dest_sheet['U4'].value
        wesm_rate_var_3 = (dest_sheet['V4'].value)/1000

        current_rate_formula_3 = (((scpc_value_3 * scpc_fixed_cost_3) + (bcq_scpc_3 * scpc_variable_cost_3)) + ((kspc_value_3 * kspc_fixed_cost_3) + (bcq_kspc_3 * kspc_variable_cost_3)) + ((edc_value_3 * edc_fixed_cost_3) + (bcq_edc_3 * edc_variable_cost_3)) + (total_ssload_var_3 - contestable_energy_var_3 - total_bcq_var_3) * wesm_rate_var_3)/(total_ssload_var_3 - contestable_energy_var_3)
        
        dest_sheet['W4'] = current_rate_formula_3
        dest_wb.save(destination_file_path)
        
    elif hour_now == '04':
        # HOUR 4
        scpc_value_4 = 25000
        scpc_fixed_cost_4 = dest_sheet['G5'].value
        bcq_scpc_4 = dest_sheet['B5'].value
        scpc_variable_cost_4 = dest_sheet['H5'].value

        kspc_value_4 = 20000
        kspc_fixed_cost_4 = dest_sheet['N5'].value
        bcq_kspc_4 = dest_sheet['I5'].value
        kspc_variable_cost_4 = dest_sheet['O5'].value

        edc_value_4 = 20000
        edc_fixed_cost_4 = dest_sheet['Q5'].value
        bcq_edc_4 = dest_sheet['P5'].value
        edc_variable_cost_4 = dest_sheet['R5'].value

        total_ssload_var_4 = dest_sheet['S5'].value
        contestable_energy_var_4 = dest_sheet['T5'].value
        total_bcq_var_4 = dest_sheet['U5'].value
        wesm_rate_var_4 = (dest_sheet['V5'].value)/1000

        current_rate_formula_4 = (((scpc_value_4 * scpc_fixed_cost_4) + (bcq_scpc_4 * scpc_variable_cost_4)) + ((kspc_value_4 * kspc_fixed_cost_4) + (bcq_kspc_4 * kspc_variable_cost_4)) + ((edc_value_4 * edc_fixed_cost_4) + (bcq_edc_4 * edc_variable_cost_4)) + (total_ssload_var_4 - contestable_energy_var_4 - total_bcq_var_4) * wesm_rate_var_4)/(total_ssload_var_4 - contestable_energy_var_4)
        
        dest_sheet['W5'] = current_rate_formula_4
        dest_wb.save(destination_file_path)
        
    elif hour_now == '05':
        # HOUR 5
        scpc_value_5 = 25000
        scpc_fixed_cost_5 = dest_sheet['G6'].value
        bcq_scpc_5 = dest_sheet['B6'].value
        scpc_variable_cost_5 = dest_sheet['H6'].value

        kspc_value_5 = 20000
        kspc_fixed_cost_5 = dest_sheet['N6'].value
        bcq_kspc_5 = dest_sheet['I6'].value
        kspc_variable_cost_5 = dest_sheet['O6'].value

        edc_value_5 = 20000
        edc_fixed_cost_5 = dest_sheet['Q6'].value
        bcq_edc_5 = dest_sheet['P6'].value
        edc_variable_cost_5 = dest_sheet['R6'].value

        total_ssload_var_5 = dest_sheet['S6'].value
        contestable_energy_var_5 = dest_sheet['T6'].value
        total_bcq_var_5 = dest_sheet['U6'].value
        wesm_rate_var_5 = (dest_sheet['V6'].value)/1000

        current_rate_formula_5 = (((scpc_value_5 * scpc_fixed_cost_5) + (bcq_scpc_5 * scpc_variable_cost_5)) + ((kspc_value_5 * kspc_fixed_cost_5) + (bcq_kspc_5 * kspc_variable_cost_5)) + ((edc_value_5 * edc_fixed_cost_5) + (bcq_edc_5 * edc_variable_cost_5)) + (total_ssload_var_5 - contestable_energy_var_5 - total_bcq_var_5) * wesm_rate_var_5)/(total_ssload_var_5 - contestable_energy_var_5)
        
        dest_sheet['W6'] = current_rate_formula_5
        dest_wb.save(destination_file_path)

    elif hour_now == '06':
        # HOUR 6
        scpc_value_6 = 25000
        scpc_fixed_cost_6 = dest_sheet['G7'].value
        bcq_scpc_6 = dest_sheet['B7'].value
        scpc_variable_cost_6 = dest_sheet['H7'].value

        kspc_value_6 = 20000
        kspc_fixed_cost_6 = dest_sheet['N7'].value
        bcq_kspc_6 = dest_sheet['I7'].value
        kspc_variable_cost_6 = dest_sheet['O7'].value

        edc_value_6 = 20000
        edc_fixed_cost_6 = dest_sheet['Q7'].value
        bcq_edc_6 = dest_sheet['P7'].value
        edc_variable_cost_6 = dest_sheet['R7'].value

        total_ssload_var_6 = dest_sheet['S7'].value
        contestable_energy_var_6 = dest_sheet['T7'].value
        total_bcq_var_6 = dest_sheet['U7'].value
        wesm_rate_var_6 = (dest_sheet['V7'].value)/1000

        current_rate_formula_6 = (((scpc_value_6 * scpc_fixed_cost_6) + (bcq_scpc_6 * scpc_variable_cost_6)) + ((kspc_value_6 * kspc_fixed_cost_6) + (bcq_kspc_6 * kspc_variable_cost_6)) + ((edc_value_6 * edc_fixed_cost_6) + (bcq_edc_6 * edc_variable_cost_6)) + (total_ssload_var_6 - contestable_energy_var_6 - total_bcq_var_6) * wesm_rate_var_6)/(total_ssload_var_6 - contestable_energy_var_6)
        
        dest_sheet['W7'] = current_rate_formula_6
        dest_wb.save(destination_file_path)

    elif hour_now == '07':
        # HOUR 7
        scpc_value_7 = 25000
        scpc_fixed_cost_7 = dest_sheet['G8'].value
        bcq_scpc_7 = dest_sheet['B8'].value
        scpc_variable_cost_7 = dest_sheet['H8'].value

        kspc_value_7 = 20000
        kspc_fixed_cost_7 = dest_sheet['N8'].value
        bcq_kspc_7 = dest_sheet['I8'].value
        kspc_variable_cost_7 = dest_sheet['O8'].value

        edc_value_7 = 20000
        edc_fixed_cost_7 = dest_sheet['Q8'].value
        bcq_edc_7 = dest_sheet['P8'].value
        edc_variable_cost_7 = dest_sheet['R8'].value

        total_ssload_var_7 = dest_sheet['S8'].value
        contestable_energy_var_7 = dest_sheet['T8'].value
        total_bcq_var_7 = dest_sheet['U8'].value
        wesm_rate_var_7 = (dest_sheet['V8'].value)/1000

        current_rate_formula_7 = (((scpc_value_7 * scpc_fixed_cost_7) + (bcq_scpc_7 * scpc_variable_cost_7)) + ((kspc_value_7 * kspc_fixed_cost_7) + (bcq_kspc_7 * kspc_variable_cost_7)) + ((edc_value_7 * edc_fixed_cost_7) + (bcq_edc_7 * edc_variable_cost_7)) + (total_ssload_var_7 - contestable_energy_var_7 - total_bcq_var_7) * wesm_rate_var_7)/(total_ssload_var_7 - contestable_energy_var_7)
        
        dest_sheet['W8'] = current_rate_formula_7
        dest_wb.save(destination_file_path)

    elif hour_now == '08':
        # HOUR 8
        scpc_value_8 = 25000
        scpc_fixed_cost_8 = dest_sheet['G9'].value
        bcq_scpc_8 = dest_sheet['B9'].value
        scpc_variable_cost_8 = dest_sheet['H9'].value

        kspc_value_8 = 20000
        kspc_fixed_cost_8 = dest_sheet['N9'].value
        bcq_kspc_8 = dest_sheet['I9'].value
        kspc_variable_cost_8 = dest_sheet['O9'].value

        edc_value_8 = 20000
        edc_fixed_cost_8 = dest_sheet['Q9'].value
        bcq_edc_8 = dest_sheet['P9'].value
        edc_variable_cost_8 = dest_sheet['R9'].value

        total_ssload_var_8 = dest_sheet['S9'].value
        contestable_energy_var_8 = dest_sheet['T9'].value
        total_bcq_var_8 = dest_sheet['U9'].value
        wesm_rate_var_8 = (dest_sheet['V9'].value)/1000

        current_rate_formula_8 = (((scpc_value_8 * scpc_fixed_cost_8) + (bcq_scpc_8 * scpc_variable_cost_8)) + ((kspc_value_8 * kspc_fixed_cost_8) + (bcq_kspc_8 * kspc_variable_cost_8)) + ((edc_value_8 * edc_fixed_cost_8) + (bcq_edc_8 * edc_variable_cost_8)) + (total_ssload_var_8 - contestable_energy_var_8 - total_bcq_var_8) * wesm_rate_var_8)/(total_ssload_var_8 - contestable_energy_var_8)
        
        dest_sheet['W9'] = current_rate_formula_8
        dest_wb.save(destination_file_path)

    elif hour_now == '09':
        # HOUR 9
        scpc_value_9 = 25000
        scpc_fixed_cost_9 = dest_sheet['G10'].value
        bcq_scpc_9 = dest_sheet['B10'].value
        scpc_variable_cost_9 = dest_sheet['H10'].value

        kspc_value_9 = 20000
        kspc_fixed_cost_9 = dest_sheet['N10'].value
        bcq_kspc_9 = dest_sheet['I10'].value
        kspc_variable_cost_9 = dest_sheet['O10'].value

        edc_value_9 = 20000
        edc_fixed_cost_9 = dest_sheet['Q10'].value
        bcq_edc_9 = dest_sheet['P10'].value
        edc_variable_cost_9 = dest_sheet['R10'].value

        total_ssload_var_9 = dest_sheet['S10'].value
        contestable_energy_var_9 = dest_sheet['T10'].value
        total_bcq_var_9 = dest_sheet['U10'].value
        wesm_rate_var_9 = (dest_sheet['V10'].value)/1000

        current_rate_formula_9 = (((scpc_value_9 * scpc_fixed_cost_9) + (bcq_scpc_9 * scpc_variable_cost_9)) + ((kspc_value_9 * kspc_fixed_cost_9) + (bcq_kspc_9 * kspc_variable_cost_9)) + ((edc_value_9 * edc_fixed_cost_9) + (bcq_edc_9 * edc_variable_cost_9)) + (total_ssload_var_9 - contestable_energy_var_9 - total_bcq_var_9) * wesm_rate_var_9)/(total_ssload_var_9 - contestable_energy_var_9)
        
        dest_sheet['W10'] = current_rate_formula_9
        dest_wb.save(destination_file_path)

    elif hour_now == '10':
        scpc_value_10 = 25000
        scpc_fixed_cost_10 = dest_sheet['G11'].value
        bcq_scpc_10 = dest_sheet['B11'].value
        scpc_variable_cost_10 = dest_sheet['H11'].value

        kspc_value_10 = 20000
        kspc_fixed_cost_10 = dest_sheet['N11'].value
        bcq_kspc_10 = dest_sheet['I11'].value
        kspc_variable_cost_10 = dest_sheet['O11'].value

        edc_value_10 = 20000
        edc_fixed_cost_10 = dest_sheet['Q11'].value
        bcq_edc_10 = dest_sheet['P11'].value
        edc_variable_cost_10 = dest_sheet['R11'].value

        total_ssload_var_10 = dest_sheet['S11'].value
        contestable_energy_var_10 = dest_sheet['T11'].value
        total_bcq_var_10 = dest_sheet['U11'].value
        wesm_rate_var_10 = (dest_sheet['V11'].value)/1000

        current_rate_formula_10 = (((scpc_value_10 * scpc_fixed_cost_10) + (bcq_scpc_10 * scpc_variable_cost_10)) + ((kspc_value_10 * kspc_fixed_cost_10) + (bcq_kspc_10 * kspc_variable_cost_10)) + ((edc_value_10 * edc_fixed_cost_10) + (bcq_edc_10 * edc_variable_cost_10)) + (total_ssload_var_10 - contestable_energy_var_10 - total_bcq_var_10) * wesm_rate_var_10)/(total_ssload_var_10 - contestable_energy_var_10)
        
        dest_sheet['W11'] = current_rate_formula_10
        dest_wb.save(destination_file_path)

    elif hour_now == '11':
        # HOUR 11
        scpc_value_11 = 25000
        scpc_fixed_cost_11 = dest_sheet['G12'].value
        bcq_scpc_11 = dest_sheet['B12'].value
        scpc_variable_cost_11 = dest_sheet['H12'].value

        kspc_value_11 = 20000
        kspc_fixed_cost_11 = dest_sheet['N12'].value
        bcq_kspc_11 = dest_sheet['I12'].value
        kspc_variable_cost_11 = dest_sheet['O12'].value

        edc_value_11 = 20000
        edc_fixed_cost_11 = dest_sheet['Q12'].value
        bcq_edc_11 = dest_sheet['P12'].value
        edc_variable_cost_11 = dest_sheet['R12'].value

        total_ssload_var_11 = dest_sheet['S12'].value
        contestable_energy_var_11 = dest_sheet['T12'].value
        total_bcq_var_11 = dest_sheet['U12'].value
        wesm_rate_var_11 = (dest_sheet['V12'].value)/1000

        current_rate_formula_11 = (((scpc_value_11 * scpc_fixed_cost_11) + (bcq_scpc_11 * scpc_variable_cost_11)) + ((kspc_value_11 * kspc_fixed_cost_11) + (bcq_kspc_11 * kspc_variable_cost_11)) + ((edc_value_11 * edc_fixed_cost_11) + (bcq_edc_11 * edc_variable_cost_11)) + (total_ssload_var_11 - contestable_energy_var_11 - total_bcq_var_11) * wesm_rate_var_11)/(total_ssload_var_11 - contestable_energy_var_11)
        
        dest_sheet['W12'] = current_rate_formula_11
        dest_wb.save(destination_file_path)

    elif hour_now == '12':
        # HOUR 12
        scpc_value_12 = 25000
        scpc_fixed_cost_12 = dest_sheet['G13'].value
        bcq_scpc_12 = dest_sheet['B13'].value
        scpc_variable_cost_12 = dest_sheet['H13'].value

        kspc_value_12 = 20000
        kspc_fixed_cost_12 = dest_sheet['N13'].value
        bcq_kspc_12 = dest_sheet['I13'].value
        kspc_variable_cost_12 = dest_sheet['O13'].value

        edc_value_12 = 20000
        edc_fixed_cost_12 = dest_sheet['Q13'].value
        bcq_edc_12 = dest_sheet['P13'].value
        edc_variable_cost_12 = dest_sheet['R13'].value

        total_ssload_var_12 = dest_sheet['S13'].value
        contestable_energy_var_12 = dest_sheet['T13'].value
        total_bcq_var_12 = dest_sheet['U13'].value
        wesm_rate_var_12 = (dest_sheet['V13'].value)/1000

        current_rate_formula_12 = (((scpc_value_12 * scpc_fixed_cost_12) + (bcq_scpc_12 * scpc_variable_cost_12)) + ((kspc_value_12 * kspc_fixed_cost_12) + (bcq_kspc_12 * kspc_variable_cost_12)) + ((edc_value_12 * edc_fixed_cost_12) + (bcq_edc_12 * edc_variable_cost_12)) + (total_ssload_var_12 - contestable_energy_var_12 - total_bcq_var_12) * wesm_rate_var_12)/(total_ssload_var_12 - contestable_energy_var_12)
        
        dest_sheet['W13'] = current_rate_formula_12
        dest_wb.save(destination_file_path)

    elif hour_now == '13':
        # HOUR 13
        scpc_value_13 = 25000
        scpc_fixed_cost_13 = dest_sheet['G14'].value
        bcq_scpc_13 = dest_sheet['B14'].value
        scpc_variable_cost_13 = dest_sheet['H14'].value

        kspc_value_13 = 20000
        kspc_fixed_cost_13 = dest_sheet['N14'].value
        bcq_kspc_13 = dest_sheet['I14'].value
        kspc_variable_cost_13 = dest_sheet['O14'].value

        edc_value_13 = 20000
        edc_fixed_cost_13 = dest_sheet['Q14'].value
        bcq_edc_13 = dest_sheet['P14'].value
        edc_variable_cost_13 = dest_sheet['R14'].value

        total_ssload_var_13 = dest_sheet['S14'].value
        contestable_energy_var_13 = dest_sheet['T14'].value
        total_bcq_var_13 = dest_sheet['U14'].value
        wesm_rate_var_13 = (dest_sheet['V14'].value)/1000

        current_rate_formula_13 = (((scpc_value_13 * scpc_fixed_cost_13) + (bcq_scpc_13 * scpc_variable_cost_13)) + ((kspc_value_13 * kspc_fixed_cost_13) + (bcq_kspc_13 * kspc_variable_cost_13)) + ((edc_value_13 * edc_fixed_cost_13) + (bcq_edc_13 * edc_variable_cost_13)) + (total_ssload_var_13 - contestable_energy_var_13 - total_bcq_var_13) * wesm_rate_var_13)/(total_ssload_var_13 - contestable_energy_var_13)
        
        dest_sheet['W14'] = current_rate_formula_13
        dest_wb.save(destination_file_path)

    elif hour_now == '14':
        # HOUR 14
        scpc_value_14 = 25000
        scpc_fixed_cost_14 = dest_sheet['G15'].value
        bcq_scpc_14 = dest_sheet['B15'].value
        scpc_variable_cost_14 = dest_sheet['H15'].value

        kspc_value_14 = 20000
        kspc_fixed_cost_14 = dest_sheet['N15'].value
        bcq_kspc_14 = dest_sheet['I15'].value
        kspc_variable_cost_14 = dest_sheet['O15'].value

        edc_value_14 = 20000
        edc_fixed_cost_14 = dest_sheet['Q15'].value
        bcq_edc_14 = dest_sheet['P15'].value
        edc_variable_cost_14 = dest_sheet['R15'].value

        total_ssload_var_14 = dest_sheet['S15'].value
        contestable_energy_var_14 = dest_sheet['T15'].value
        total_bcq_var_14 = dest_sheet['U15'].value
        wesm_rate_var_14 = (dest_sheet['V15'].value)/1000

        current_rate_formula_14 = (((scpc_value_14 * scpc_fixed_cost_14) + (bcq_scpc_14 * scpc_variable_cost_14)) + ((kspc_value_14 * kspc_fixed_cost_14) + (bcq_kspc_14 * kspc_variable_cost_14)) + ((edc_value_14 * edc_fixed_cost_14) + (bcq_edc_14 * edc_variable_cost_14)) + (total_ssload_var_14 - contestable_energy_var_14 - total_bcq_var_14) * wesm_rate_var_14)/(total_ssload_var_14 - contestable_energy_var_14)
        
        dest_sheet['W15'] = current_rate_formula_14
        dest_wb.save(destination_file_path)

    elif hour_now == '15':
        # HOUR 15
        scpc_value_15 = 25000
        scpc_fixed_cost_15 = dest_sheet['G16'].value
        bcq_scpc_15 = dest_sheet['B16'].value
        scpc_variable_cost_15 = dest_sheet['H16'].value

        kspc_value_15 = 20000
        kspc_fixed_cost_15 = dest_sheet['N16'].value
        bcq_kspc_15 = dest_sheet['I16'].value
        kspc_variable_cost_15 = dest_sheet['O16'].value

        edc_value_15 = 20000
        edc_fixed_cost_15 = dest_sheet['Q16'].value
        bcq_edc_15 = dest_sheet['P16'].value
        edc_variable_cost_15 = dest_sheet['R16'].value

        total_ssload_var_15 = dest_sheet['S16'].value
        contestable_energy_var_15 = dest_sheet['T16'].value
        total_bcq_var_15 = dest_sheet['U16'].value
        wesm_rate_var_15 = (dest_sheet['V16'].value)/1000

        current_rate_formula_15 = (((scpc_value_15 * scpc_fixed_cost_15) + (bcq_scpc_15 * scpc_variable_cost_15)) + ((kspc_value_15 * kspc_fixed_cost_15) + (bcq_kspc_15 * kspc_variable_cost_15)) + ((edc_value_15 * edc_fixed_cost_15) + (bcq_edc_15 * edc_variable_cost_15)) + (total_ssload_var_15 - contestable_energy_var_15 - total_bcq_var_15) * wesm_rate_var_15)/(total_ssload_var_15 - contestable_energy_var_15)
        
        dest_sheet['W16'] = current_rate_formula_15
        dest_wb.save(destination_file_path)

    elif hour_now == '16':
        # HOUR 16
        scpc_value_16 = 25000
        scpc_fixed_cost_16 = dest_sheet['G17'].value
        bcq_scpc_16 = dest_sheet['B17'].value
        scpc_variable_cost_16 = dest_sheet['H17'].value

        kspc_value_16 = 20000
        kspc_fixed_cost_16 = dest_sheet['N17'].value
        bcq_kspc_16 = dest_sheet['I17'].value
        kspc_variable_cost_16 = dest_sheet['O17'].value

        edc_value_16 = 20000
        edc_fixed_cost_16 = dest_sheet['Q17'].value
        bcq_edc_16 = dest_sheet['P17'].value
        edc_variable_cost_16 = dest_sheet['R17'].value

        total_ssload_var_16 = dest_sheet['S17'].value
        contestable_energy_var_16 = dest_sheet['T17'].value
        total_bcq_var_16 = dest_sheet['U17'].value
        wesm_rate_var_16 = (dest_sheet['V17'].value)/1000

        current_rate_formula_16 = (((scpc_value_16 * scpc_fixed_cost_16) + (bcq_scpc_16 * scpc_variable_cost_16)) + ((kspc_value_16 * kspc_fixed_cost_16) + (bcq_kspc_16 * kspc_variable_cost_16)) + ((edc_value_16 * edc_fixed_cost_16) + (bcq_edc_16 * edc_variable_cost_16)) + (total_ssload_var_16 - contestable_energy_var_16 - total_bcq_var_16) * wesm_rate_var_16)/(total_ssload_var_16 - contestable_energy_var_16)
        
        dest_sheet['W17'] = current_rate_formula_16
        dest_wb.save(destination_file_path)

    elif hour_now == '17':
        # HOUR 17
        scpc_value_17 = 25000
        scpc_fixed_cost_17 = dest_sheet['G18'].value
        bcq_scpc_17 = dest_sheet['B18'].value
        scpc_variable_cost_17 = dest_sheet['H18'].value

        kspc_value_17 = 20000
        kspc_fixed_cost_17 = dest_sheet['N18'].value
        bcq_kspc_17 = dest_sheet['I18'].value
        kspc_variable_cost_17 = dest_sheet['O18'].value

        edc_value_17 = 20000
        edc_fixed_cost_17 = dest_sheet['Q18'].value
        bcq_edc_17 = dest_sheet['P18'].value
        edc_variable_cost_17 = dest_sheet['R18'].value

        total_ssload_var_17 = dest_sheet['S18'].value
        contestable_energy_var_17 = dest_sheet['T18'].value
        total_bcq_var_17 = dest_sheet['U18'].value
        wesm_rate_var_17 = (dest_sheet['V18'].value)/1000

        current_rate_formula_17 = (((scpc_value_17 * scpc_fixed_cost_17) + (bcq_scpc_17 * scpc_variable_cost_17)) + ((kspc_value_17 * kspc_fixed_cost_17) + (bcq_kspc_17 * kspc_variable_cost_17)) + ((edc_value_17 * edc_fixed_cost_17) + (bcq_edc_17 * edc_variable_cost_17)) + (total_ssload_var_17 - contestable_energy_var_17 - total_bcq_var_17) * wesm_rate_var_17)/(total_ssload_var_17 - contestable_energy_var_17)
        
        dest_sheet['W18'] = current_rate_formula_17
        dest_wb.save(destination_file_path)

    elif hour_now == '18':
        # HOUR 18
        scpc_value_18 = 25000
        scpc_fixed_cost_18 = dest_sheet['G19'].value
        bcq_scpc_18 = dest_sheet['B19'].value
        scpc_variable_cost_18 = dest_sheet['H19'].value

        kspc_value_18 = 20000
        kspc_fixed_cost_18 = dest_sheet['N19'].value
        bcq_kspc_18 = dest_sheet['I19'].value
        kspc_variable_cost_18 = dest_sheet['O19'].value

        edc_value_18 = 20000
        edc_fixed_cost_18 = dest_sheet['Q19'].value
        bcq_edc_18 = dest_sheet['P19'].value
        edc_variable_cost_18 = dest_sheet['R19'].value

        total_ssload_var_18 = dest_sheet['S19'].value
        contestable_energy_var_18 = dest_sheet['T19'].value
        total_bcq_var_18 = dest_sheet['U19'].value
        wesm_rate_var_18 = (dest_sheet['V19'].value)/1000

        current_rate_formula_18 = (((scpc_value_18 * scpc_fixed_cost_18) + (bcq_scpc_18 * scpc_variable_cost_18)) + ((kspc_value_18 * kspc_fixed_cost_18) + (bcq_kspc_18 * kspc_variable_cost_18)) + ((edc_value_18 * edc_fixed_cost_18) + (bcq_edc_18 * edc_variable_cost_18)) + (total_ssload_var_18 - contestable_energy_var_18 - total_bcq_var_18) * wesm_rate_var_18)/(total_ssload_var_18 - contestable_energy_var_18)
        
        dest_sheet['W19'] = current_rate_formula_18
        dest_wb.save(destination_file_path)

    elif hour_now == '19':
        # HOUR 19 
        scpc_value_19 = 25000
        scpc_fixed_cost_19 = dest_sheet['G20'].value
        bcq_scpc_19 = dest_sheet['B20'].value
        scpc_variable_cost_19 = dest_sheet['H20'].value

        kspc_value_19 = 20000
        kspc_fixed_cost_19 = dest_sheet['N20'].value
        bcq_kspc_19 = dest_sheet['I20'].value
        kspc_variable_cost_19 = dest_sheet['O20'].value

        edc_value_19 = 20000
        edc_fixed_cost_19 = dest_sheet['Q20'].value
        bcq_edc_19 = dest_sheet['P20'].value
        edc_variable_cost_19 = dest_sheet['R20'].value

        total_ssload_var_19 = dest_sheet['S20'].value
        contestable_energy_var_19 = dest_sheet['T20'].value
        total_bcq_var_19 = dest_sheet['U20'].value
        wesm_rate_var_19 = (dest_sheet['V20'].value)/1000

        current_rate_formula_19 = (((scpc_value_19 * scpc_fixed_cost_19) + (bcq_scpc_19 * scpc_variable_cost_19)) + ((kspc_value_19 * kspc_fixed_cost_19) + (bcq_kspc_19 * kspc_variable_cost_19)) + ((edc_value_19 * edc_fixed_cost_19) + (bcq_edc_19 * edc_variable_cost_19)) + (total_ssload_var_19 - contestable_energy_var_19 - total_bcq_var_19) * wesm_rate_var_19)/(total_ssload_var_19 - contestable_energy_var_19)
        
        dest_sheet['W20'] = current_rate_formula_19
        dest_wb.save(destination_file_path)    

    elif hour_now == '20':
        # HOUR 20
        scpc_value_20 = 25000
        scpc_fixed_cost_20 = dest_sheet['G21'].value
        bcq_scpc_20 = dest_sheet['B21'].value
        scpc_variable_cost_20 = dest_sheet['H21'].value

        kspc_value_20 = 20000
        kspc_fixed_cost_20 = dest_sheet['N21'].value
        bcq_kspc_20 = dest_sheet['I21'].value
        kspc_variable_cost_20 = dest_sheet['O21'].value

        edc_value_20 = 20000
        edc_fixed_cost_20 = dest_sheet['Q21'].value
        bcq_edc_20 = dest_sheet['P21'].value
        edc_variable_cost_20 = dest_sheet['R21'].value

        total_ssload_var_20 = dest_sheet['S21'].value
        contestable_energy_var_20 = dest_sheet['T21'].value
        total_bcq_var_20 = dest_sheet['U21'].value
        wesm_rate_var_20 = (dest_sheet['V21'].value)/1000

        current_rate_formula_20 = (((scpc_value_20 * scpc_fixed_cost_20) + (bcq_scpc_20 * scpc_variable_cost_20)) + ((kspc_value_20 * kspc_fixed_cost_20) + (bcq_kspc_20 * kspc_variable_cost_20)) + ((edc_value_20 * edc_fixed_cost_20) + (bcq_edc_20 * edc_variable_cost_20)) + (total_ssload_var_20 - contestable_energy_var_20 - total_bcq_var_20) * wesm_rate_var_20)/(total_ssload_var_20 - contestable_energy_var_20)
        
        dest_sheet['W21'] = current_rate_formula_20
        dest_wb.save(destination_file_path)

    elif hour_now == '21':
        # HOUR 21
        scpc_value_21 = 25000
        scpc_fixed_cost_21 = dest_sheet['G22'].value
        bcq_scpc_21 = dest_sheet['B22'].value
        scpc_variable_cost_21 = dest_sheet['H22'].value

        kspc_value_21 = 20000
        kspc_fixed_cost_21 = dest_sheet['N22'].value
        bcq_kspc_21 = dest_sheet['I22'].value
        kspc_variable_cost_21 = dest_sheet['O22'].value

        edc_value_21 = 20000
        edc_fixed_cost_21 = dest_sheet['Q22'].value
        bcq_edc_21 = dest_sheet['P22'].value
        edc_variable_cost_21 = dest_sheet['R22'].value

        total_ssload_var_21 = dest_sheet['S22'].value
        contestable_energy_var_21 = dest_sheet['T22'].value
        total_bcq_var_21 = dest_sheet['U22'].value
        wesm_rate_var_21 = (dest_sheet['V22'].value)/1000

        current_rate_formula_21 = (((scpc_value_21 * scpc_fixed_cost_21) + (bcq_scpc_21 * scpc_variable_cost_21)) + ((kspc_value_21 * kspc_fixed_cost_21) + (bcq_kspc_21 * kspc_variable_cost_21)) + ((edc_value_21 * edc_fixed_cost_21) + (bcq_edc_21 * edc_variable_cost_21)) + (total_ssload_var_21 - contestable_energy_var_21 - total_bcq_var_21) * wesm_rate_var_21)/(total_ssload_var_21 - contestable_energy_var_21)
        
        dest_sheet['W22'] = current_rate_formula_21
        dest_wb.save(destination_file_path)
    
    elif hour_now == '22':
        # HOUR 22
        scpc_value_22 = 25000
        scpc_fixed_cost_22 = dest_sheet['G23'].value
        bcq_scpc_22 = dest_sheet['B23'].value
        scpc_variable_cost_22 = dest_sheet['H23'].value

        kspc_value_22 = 20000
        kspc_fixed_cost_22 = dest_sheet['N23'].value
        bcq_kspc_22 = dest_sheet['I23'].value
        kspc_variable_cost_22 = dest_sheet['O23'].value

        edc_value_22 = 20000
        edc_fixed_cost_22 = dest_sheet['Q23'].value
        bcq_edc_22 = dest_sheet['P23'].value
        edc_variable_cost_22 = dest_sheet['R23'].value

        total_ssload_var_22 = dest_sheet['S23'].value
        contestable_energy_var_22 = dest_sheet['T23'].value
        total_bcq_var_22 = dest_sheet['U23'].value
        wesm_rate_var_22 = (dest_sheet['V23'].value)/1000

        current_rate_formula_22 = (((scpc_value_22 * scpc_fixed_cost_22) + (bcq_scpc_22 * scpc_variable_cost_22)) + ((kspc_value_22 * kspc_fixed_cost_22) + (bcq_kspc_22 * kspc_variable_cost_22)) + ((edc_value_22 * edc_fixed_cost_22) + (bcq_edc_22 * edc_variable_cost_22)) + (total_ssload_var_22 - contestable_energy_var_22 - total_bcq_var_22) * wesm_rate_var_22)/(total_ssload_var_22 - contestable_energy_var_22)
        
        dest_sheet['W23'] = current_rate_formula_22
        dest_wb.save(destination_file_path)

    elif hour_now == '23':
        # HOUR 23
        scpc_value_23 = 25000
        scpc_fixed_cost_23 = dest_sheet['G24'].value
        bcq_scpc_23 = dest_sheet['B24'].value
        scpc_variable_cost_23 = dest_sheet['H24'].value

        kspc_value_23 = 20000
        kspc_fixed_cost_23 = dest_sheet['N24'].value
        bcq_kspc_23 = dest_sheet['I24'].value
        kspc_variable_cost_23 = dest_sheet['O24'].value

        edc_value_23 = 20000
        edc_fixed_cost_23 = dest_sheet['Q24'].value
        bcq_edc_23 = dest_sheet['P24'].value
        edc_variable_cost_23 = dest_sheet['R24'].value

        total_ssload_var_23 = dest_sheet['S24'].value
        contestable_energy_var_23 = dest_sheet['T24'].value
        total_bcq_var_23 = dest_sheet['U24'].value
        wesm_rate_var_23 = (dest_sheet['V24'].value)/1000

        current_rate_formula_23 = (((scpc_value_23 * scpc_fixed_cost_23) + (bcq_scpc_23 * scpc_variable_cost_23)) + ((kspc_value_23 * kspc_fixed_cost_23) + (bcq_kspc_23 * kspc_variable_cost_23)) + ((edc_value_23 * edc_fixed_cost_23) + (bcq_edc_23 * edc_variable_cost_23)) + (total_ssload_var_23 - contestable_energy_var_23 - total_bcq_var_23) * wesm_rate_var_23)/(total_ssload_var_23 - contestable_energy_var_23)
        
        dest_sheet['W24'] = current_rate_formula_23
        dest_wb.save(destination_file_path)

    elif hour_now == '00':
        # HOUR 24
        scpc_value_24 = 25000
        scpc_fixed_cost_24 = dest_sheet['G25'].value
        bcq_scpc_24 = dest_sheet['B25'].value
        scpc_variable_cost_24 = dest_sheet['H25'].value

        kspc_value_24 = 20000
        kspc_fixed_cost_24 = dest_sheet['N25'].value
        bcq_kspc_24 = dest_sheet['I25'].value
        kspc_variable_cost_24 = dest_sheet['O25'].value

        edc_value_24 = 20000
        edc_fixed_cost_24 = dest_sheet['Q25'].value
        bcq_edc_24 = dest_sheet['P25'].value
        edc_variable_cost_24 = dest_sheet['R25'].value

        total_ssload_var_24 = dest_sheet['S25'].value
        contestable_energy_var_24 = dest_sheet['T25'].value
        total_bcq_var_24 = dest_sheet['U25'].value
        wesm_rate_var_24 = (dest_sheet['V25'].value)/1000

        current_rate_formula_24 = 0
        
        dest_sheet['W25'] = current_rate_formula_24
        dest_wb.save(destination_file_path)   

# Function to retrieve the current hour in 24-hour format
def get_current_hour():
    now = datetime.now()
    current_hour = now.hour
    if current_hour == 0:
        return 24
    return current_hour

# Function to get current rate value for the current hour (1 to 24) from Excel
def get_current_rate_for_current_hour(excel_file, current_hour, default_value=0.0):
    # Read Excel file into a Pandas DataFrame
    df = pd.read_excel(excel_file, header=None, names=['HOUR', 'BCQ_SCPC', 'RATE_SCPC_B1_FIXED_FEE', 
                                                        'RATE_SCPC_B1_VARIABLE_FEE', 'RATE_SCPC_B2_FIXED_FEE', 
                                                        'RATE_SCPC_B2_VARIABLE_FEE', 'RATE_SCPC_AVE_FIXED_FEE', 
                                                        'RATE_SCPC_AVE_VARIABLE_FEE', 'BCQ_KSPC', 'RATE_KSPC_B1_FIXED_FEE', 
                                                        'RATE_KSPC_B1_VARIABLE_FEE', 'RATE_KSPC_B2_FIXED_FEE', 
                                                        'RATE_KSPC_B2_VARIABLE_FEE', 'RATE_KSPC_AVE_FIXED_FEE', 
                                                        'RATE_KSPC_AVE_VARIABLE_FEE', 'BCQ_EDC', 'RATE_EDC_FIXED_FEE', 
                                                        'RATE_EDC_VARIABLE_FEE', 'TOTAL_SS_LOAD', 'CONTESTABLE_ENERGY', 
                                                        'TOTAL_BCQ', 'WESM_RATE', 'CURRENT_RATE'], skiprows=1)

    # Filter the row where 'HOUR' matches the current_hour
    current_rate_value = df.loc[df['HOUR'] == current_hour, 'CURRENT_RATE']

    if current_rate_value.empty or pd.isna(current_rate_value.iloc[0]):
        # Get today's date and time
        now = datetime.now()
        # Subtract one hour to get the previous hour
        previous_hour = now - timedelta(hours=1)
        # Extract the hour part
        previous_hour_only = previous_hour.hour
        # Filter the row where 'HOUR' matches the previous_hour
        previous_rate_value = df.loc[df['HOUR'] == previous_hour_only, 'CURRENT_RATE']
        # Check if the previous rate value is not empty or NaN, and convert it to float
        if not previous_rate_value.empty and not pd.isna(previous_rate_value.iloc[0]):
            float_current_rate_value = float(previous_rate_value.iloc[0])
        else:
            float_current_rate_value = float(default_value)
    else:
        float_current_rate_value = float(current_rate_value.iloc[0])

    return float_current_rate_value

# Function to insert value into MySQL database
def insert_into_mysql(conn, float_current_rate_value):
    # Create a cursor object
    cursor = conn.cursor()
    # SQL Insert Statement
    sql = "INSERT INTO current_rate (rate) VALUES (%s)"
    # Execute the SQL Query
    cursor.execute(sql, (float_current_rate_value,))
    # Commit the transaction
    conn.commit()
    # Close the cursor
    cursor.close()
    # Print a confirmation message
    print(f"Value {float_current_rate_value} inserted successfully into MySQL.")

def get_last_value_from_mysql(conn):
    cursor = conn.cursor()
    sql = "SELECT rate FROM current_rate ORDER BY id DESC LIMIT 1"
    cursor.execute(sql)
    result = cursor.fetchone()
    last_result = result[0]
    cursor.close()
    print(f"Value {last_result} inserted successfully into MySQL.")
    return last_result

# Main function to run the script
def main(excel_file, conn):
    initial_function()
    total_ss_load()
    contestable_energy()
    find_total()
    find_total_2()
    find_total_3()
    current_hour = get_current_hour()
    float_current_rate_value = get_current_rate_for_current_hour(excel_file, current_hour)
    if float_current_rate_value is not None:
        # Insert into MySQL database
        insert_into_mysql(conn, float_current_rate_value)
    else: 
        get_last_value_from_mysql(conn)
    # Wait for 60 seconds before next update
    time.sleep(60)

if __name__ == "__main__":
    while True:
        excel_file = 'current_rate.xlsx'
        
        conn = mysql.connector.connect(
            host='localhost',
            database='myDb',
            user='root',
            password=''
        )

        # Create MySQL table if it doesn't exist
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS current_rate (
                id INT AUTO_INCREMENT PRIMARY KEY,
                rate FLOAT(10, 4) NOT NULL,
                insert_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cursor.close()

        # Run main function to continuously check and update database
        main(excel_file, conn)