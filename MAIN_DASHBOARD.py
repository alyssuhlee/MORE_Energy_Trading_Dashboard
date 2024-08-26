# -- NECESSARY IMPORTS -- 
from datetime import datetime
from mysql.connector import Error
from openpyxl import load_workbook
from pandas.errors import EmptyDataError
import altair as alt
import mysql.connector
import openpyxl
import os
import pandas as pd
import plotly.express as px
import plotly.graph_objs as go
import streamlit as st
import time

# Authenticator Imports
import pickle
from pathlib import Path
import streamlit_authenticator as stauth


REFRESH_INTERVAL = 900 # Refresh Interval Value (900 SECONDS)

# -- START OF FUNCTIONS --
def view_all_data():
    c.execute('select * from timeintervals order by id asc')
    data=c.fetchall()
    return data

def view_all_data2():
    c.execute('select * from forecasted_energy order by id asc')
    data=c.fetchall()
    return data

def view_all_data3():
    c.execute('select * from total_bcq_nomination order by id asc')
    data=c.fetchall()
    return data

def view_all_data4():
    c.execute('select * from total_substation_load order by id asc')
    data=c.fetchall()
    return data

def view_all_data5():
    c.execute('select * from contestable_energy order by id asc')
    data=c.fetchall()
    return data

def view_all_data6():
    c.execute('select * from actual_energy order by id asc')
    data = c.fetchall()
    return data

def view_all_data7():
    c.execute('select * from wesm_exposure order by id asc')
    data = c.fetchall()
    return data

def view_all_data8():
    c.execute('select * from more_trading_node order by id asc')
    data = c.fetchall()
    return data

def view_all_data9():
    c.execute('select * from current_rate order by id asc')
    data = c.fetchall()
    return data

def view_all_data10():
    c.execute('select * from temp_data order by id asc')
    data = c.fetchall()
    return data

# Function to fetch latest 10 entries
def fetch_latest_10_entries():
    latest_10_entries = df.head(10)
    return latest_10_entries

# Function to fetch latest entry
def fetch_latest_entry():
    latest_entry = df.head(1)
    return latest_entry

def safe_get(df, row, column):
    try:
        return df.at[row, column]
    except KeyError:
        return 0

# Function to copy values from source range to destination range
def copy_values_slg(source_sheet_slg, dest_sheet_slg, source_range_slg, dest_range_slg):
    source_start_slg, source_end_slg = source_range_slg
    dest_start_slg, dest_end_slg = dest_range_slg

    # Copy values from source to destination
    source_values_slg = []
    for row in source_sheet_slg[source_start_slg:source_end_slg]:
        for cell in row:
            source_values_slg.append(cell.value)
    
    dest_index_slg = 0
    for row_slg in dest_sheet_slg[dest_start_slg:dest_end_slg]:
        for cell_slg in row_slg:
            if dest_index_slg < len(source_values_slg):
                cell_slg.value = source_values_slg[dest_index_slg]
                dest_index_slg += 1
            else:
                break

def load_data_from_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Hour'
    sheet['B1'] = 'Lapaz'
    sheet['C1'] = 'Jaro'
    sheet['D1'] = 'Mandurriao'
    sheet['E1'] = 'Molo'
    sheet['F1'] = 'Diversion'
    sheet['G1'] = 'Mobile SS 1'
    sheet['H1'] = 'Mobile SS 2'
    sheet['I1'] = 'Megaworld'
    workbook.save('station_load_graph.xlsx')
    workbook.close()
    print("Excel file 'station_load_graph.xlsx' created successfully.")

    # File path for destination Excel file
    destination_file_path_slg = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load_graph.xlsx'
    source_file_path_slg = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load.xlsx'

    # Load source workbook
    source_wb_slg = load_workbook(source_file_path_slg)  
    # Sheet name of the source workbook
    source_sheet_slg = source_wb_slg['Sheet1']

    # Load destination workbook
    dest_wb_slg = load_workbook(destination_file_path_slg)

    # Sheet name of the destination workbook
    dest_sheet_slg = dest_wb_slg['Sheet']

    # Define source and destination ranges
    source_ranges_slg = [('B2', 'B25'), ('C2', 'C25'), ('D2', 'D25'), ('E2', 'E25'), ('F2', 'F25'), ('G2', 'G25'), ('H2', 'H25'), ('I2', 'I25'), ('J2', 'J25')]
    dest_ranges_slg = [('A2', 'A25'), ('B2', 'B25'), ('C2', 'C25'), ('D2', 'D25'), ('E2', 'E25'), ('F2', 'F25'), ('G2', 'G25'), ('H2', 'H25'), ('I2', 'I25')]
    
    # Copy data from source to destination based on specified ranges
    for i in range(len(source_ranges_slg)):
        copy_values_slg(source_sheet_slg, dest_sheet_slg, source_ranges_slg[i], dest_ranges_slg[i])
    
    # Save the destination workbook
    dest_wb_slg.save(destination_file_path_slg)

    # Read the Excel file into a Pandas Dataframe
    file_path = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load_graph.xlsx"
    df = pd.read_excel(file_path)

    # Connect to the MySQL Database
    try:
        connection = mysql.connector.connect(
            host='localhost',
            database='myDb',
            user='root',
            password=''
        )

        if connection.is_connected():
            print("Successfully connected to the database")
            cursor = connection.cursor()

            # Create a table
            table_name = 'station_load_graph'
            create_table_query = f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                id INT AUTO_INCREMENT PRIMARY KEY,
                Hour INT(11) NOT NULL,
                Lapaz INT(11) NOT NULL,
                Jaro INT(11) NOT NULL,
                Mandurriao INT(11) NOT NULL,
                Molo INT(11) NOT NULL,
                Diversion INT(11) NOT NULL,
                `Mobile SS 1` INT(11) NOT NULL,
                `Mobile SS 2` INT(11) NOT NULL,
                Megaworld INT(11) NOT NULL
            )
            """

            cursor.execute(create_table_query)
            print("Table created successfully or table already exists")

            # Insert data from dataframe into the MySQL table
            for i, row in df.iterrows():
                sql = f"INSERT INTO {table_name} (Hour, Lapaz, Jaro, Mandurriao, Molo, Diversion, `Mobile SS 1`, `Mobile SS 2`, Megaworld) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"
                cursor.execute(sql, tuple(row))
                print(f"Inserted row {i + 1}")

            # Commit the transaction
            connection.commit()
            print("Data committed to the database")

    except Error as e:
        print(f"Error: {e}")

    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            print("MySQL connection is closed.")

    # Read the Excel file
    file_path = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load_graph.xlsx"
    df = pd.read_excel(file_path, sheet_name='Sheet')
    return df

# Calculates the average of every 12 values in a list and stores them
def average_every_12(data):
    averages = []
    for i in range(0, len(data), 12):
        window = data[i:i+12]
        if window:
            averages.append(sum(window)/len(window))
    # Returns a list containing the averages for each 12-element window
    return averages

# Function to copy values from source range to destination range
def copy_values(source_sheet, dest_sheet, source_range, dest_range):
    source_start, source_end = source_range
    dest_start, dest_end = dest_range
    
    # Copy values from source to destination
    source_values = []
    for row in source_sheet[source_start:source_end]:
        for cell in row:
            source_values.append(cell.value)
    
    dest_index = 0
    for row in dest_sheet[dest_start:dest_end]:
        for cell in row:
            if dest_index < len(source_values):
                cell.value = source_values[dest_index]
                dest_index += 1
            else:
                break

# Helper function to format numbers in 'k' or 'M' style with two decimal places
def format_value(value):
    if value >= 1_000_000:
        return f'{value / 1_000_000:.2f}M'  # Format to millions with two decimal places
    elif value >= 1_000:
        return f'{value / 1_000:.2f}k'  # Format to thousands with two decimal places
    else:
        return f'{value:.2f}'  # Format with two decimal places for values less than 1,000

def sub_load_func():
    # -- START OF DISPLAYING THE SUBSTATION LOAD (KW) BAR CHART -- 

        # Load data
        df_excel = load_data_from_excel()

        # For getting the current hour
        now = datetime.now()
        now_hour = now.hour
        if now_hour == 0:
            now_hour = 24

        # Find the row where the hour matches the current hour
        matching_row = df_excel[df_excel['Hour'] == now_hour]

        # Check if there's no matching data for the current hour
        if matching_row.empty:
            # If no data for the current hour, get the latest available data
            matching_row = df_excel.iloc[-1]  # Get the last row in the DataFrame
        else:
            matching_row = matching_row.iloc[0]  # Get the first row matching the current hour

        # Extract the values
        hour_value = matching_row['Hour']
        lapaz_value = matching_row['Lapaz']
        jaro_value = matching_row['Jaro']
        mandurriao_value = matching_row['Mandurriao']
        molo_value = matching_row['Molo']
        diversion_value = matching_row['Diversion']
        mobile_ss1_value = matching_row['Mobile SS 1']
        mobile_ss2_value = matching_row['Mobile SS 2']
        megaworld_value = matching_row['Megaworld']

        # Prepare data for the bar chart
        chart_data = pd.DataFrame({
            'Substation': ['Lapaz', 'Jaro', 'Mandurriao', 'Molo', 'Diversion', 'Mobile SS 1', 'Mobile SS 2', 'Megaworld'],
            'kW': [lapaz_value, jaro_value, mandurriao_value, molo_value, diversion_value, mobile_ss1_value, mobile_ss2_value, megaworld_value]
        })

        # Create the horizontal bar chart with Plotly
        fig_ss_load = px.bar(chart_data, y='Substation', x='kW', text='kW', orientation='h')

        # Customize the bar colors
        colors = ['orangered', 'green', 'cornflowerblue', 'darkblue', 'yellow', 'lightpink', 'mediumpurple', 'lime']
        fig_ss_load.update_traces(marker_color=colors)

        # Customize the layout
        fig_ss_load.update_traces(
            texttemplate='%{x:,.0f}',  # Format numbers with thousands separators and no decimal places
            textposition='outside',
            textfont_size=12  # Reduce text size if needed
        )

        # Add padding by adjusting the layout
        fig_ss_load.update_layout(
            title='Substation Load (kW)',
            margin=dict(r=30, t=30, b=15),  # Increase margins further
            xaxis=dict(
                title='kW',
                tickformat=',',  # Format axis ticks with commas as thousands separators
                showgrid=False,
                zeroline=False,
                range=[0, max(chart_data['kW']) * 1.2]  # Extend the x-axis range further
            ),
            uniformtext_minsize=8,
            uniformtext_mode='hide',
            height=260  # Set the height of the bar chart
        )

        fig_ss_load.update_layout(
            plot_bgcolor='black',
            paper_bgcolor='black'
        )

        return fig_ss_load

        # -- END OF DISPLAYING THE SUBSTATION LOAD (KW) BAR CHART -- 

def actual_vs_forecasted():
    # -- START OF DISPLAYING THE ACTUAL VS FORECASTED ENERGY CHART --

        # File paths
        station_load_destination_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load_graph.xlsx'
        bcq_nomination_destination_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\total_bcq_nomination.xlsx'
        forecasted_energy_destination_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\forecasted_energy.xlsx'
        actual_energy_destination_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\actual_energy.xlsx'

        # Load Actual Energy data
        actual_energy_wb = openpyxl.load_workbook(actual_energy_destination_path)
        actual_energy_sheet = actual_energy_wb.active
        actual_energy_values = []  # Initialize Actual Energy list

        now = datetime.now()
        current_hour = now.hour
        if current_hour == 0:
            current_hour = 24

        # Create a list of actual energy values for current hour backwards
        for i in range(2, current_hour + 2):
            ae_value = actual_energy_sheet[f'D{i}'].value
            actual_energy_values.append(ae_value)

        # Read the station load data into a DataFrame
        df = pd.read_excel(station_load_destination_path)

        # Initialize a dictionary to store the results
        data = {'Hour': list(range(1, 25))}

        # List of stations and their corresponding colors
        stations = {
            'Lapaz': 'red',
            'Jaro': 'orange',
            'Mandurriao': 'yellow',
            'Molo': 'lime',
            'Diversion': 'green',
            'Mobile SS 1': 'cornflowerblue',
            'Mobile SS 2': 'darkblue',
            'Megaworld': 'mediumpurple'
        }

        # Loop through each station and add their hourly values to the dictionary
        for station in stations.keys():
            data[station] = [safe_get(df, hour - 1, station) for hour in data['Hour']]

        # Calculate total values for each hour
        data['Total'] = [sum(data[station][hour - 1] for station in stations.keys()) for hour in data['Hour']]

        # Convert the dictionary to a DataFrame
        result_df = pd.DataFrame(data)

        # Create a Figure object for the plot
        fig = go.Figure()

        # Add each station as a bar trace with specified colors
        for station, color in stations.items():
            fig.add_trace(go.Bar(
                x=result_df['Hour'],
                y=[value * 0.9 for value in result_df[station]],  # Lower the bar heights slightly
                name=station,
                marker_color=color
            ))

        # Add total as a scatter trace with a straight line
        df_2 = pd.read_excel(forecasted_energy_destination_path)
        fig.add_trace(go.Scatter(
            x=df_2['HOUR'],
            y=df_2['NET'],
            mode='lines',
            name='Forecasted Energy',
            line=dict(color='white')
        ))

        # Add a crooked line for demonstration
        df_3 = pd.read_excel(bcq_nomination_destination_path)
        fig.add_trace(go.Scatter(
            x=df_3['HOUR'],
            y=df_3['TOTAL_BCQ'],
            mode='lines+markers',
            name='BCQ Nomination',
            line=dict(color='silver')
        ))

        # Update layout with x and y axis labels and custom y-axis formatting
        fig.update_layout(
            title='Actual vs Forecasted Energy (kWh)',
            margin=dict(l=0, r=0, t=30, b=0),  # Adjust margins to fit small screens
            xaxis_title='Hour',
            yaxis_title='kWh',
            xaxis=dict(
                tickmode='array',
                tickvals=list(range(1, 25)),
                ticktext=[str(i) for i in range(1, 25)]
            ),
            yaxis_tickformat=',',  # Format y-axis to use commas for thousands
            plot_bgcolor='black',
            paper_bgcolor='black',
            barmode='stack',
            height=260  # Maintain the height at 260
        )

        # Add actual energy values as text annotations on top of the bars
        for i, hour in enumerate(result_df['Hour']):
            if i < len(actual_energy_values) and hour <= current_hour:
                formatted_value = format_value(int(actual_energy_values[i]))
                fig.add_trace(go.Scatter(
                    x=[hour],
                    y=[result_df['Total'][i] * 0.92],  # Position text slightly above the adjusted bar height
                    mode='text',
                    text=[formatted_value],
                    textposition='top center',
                    showlegend=False,
                    textfont=dict(size=10, color='white')  # Adjust font size and color for better visibility
                ))

        return fig

        # -- END OF DISPLAYING THE ACTUAL VS FORECASTED ENERGY CHART --

def tipc_func():
            # -- START OF TRADING INTERVAL PRICE CALCULATION LINE CHART --

        directory = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\MORE Trading Node'
        now = datetime.now()
        date_today = now.strftime("%Y%m%d")
        contents = os.listdir(directory)
        for item in contents:
            if item == date_today:
                folder_path = os.path.join(directory, item) 
                if os.path.isdir(folder_path):
                    files_in_folder = os.listdir(folder_path) # MPI_LMP_DIPC_202407110000, etc.
                    
                    last_known_value = None

                    list_botoca_g01 = []
                    list_calaca_g01 = []
                    list_calaca_g02 = []
                    list_kal_g01 = []
                    list_kal_g02 = []
                    list_kal_g03 = []
                    list_kal_g04 = []
                    list_leyte_a = []
                    list_kspc_g01 = []
                    list_kspc_g02 = []
                    list_bantap_l01 = []
                    list_pedc_t1l1 = []
                    list_pedc_t1l2 = []
                    list_stbar_t1l1 = []

                    results_botoca_g01 = []
                    results_calaca_g01 = []
                    results_calaca_g02 = []
                    results_kal_g01 = []
                    results_kal_g02 = []
                    results_kal_g03 = []
                    results_kal_g04 = []
                    results_leyte_a = []
                    results_kspc_g01 = []
                    results_kspc_g02 = []
                    results_bantap_l01 = []
                    results_pedc_t1l1 = []
                    results_pedc_t1l2 = []
                    results_stbar_t1l1 = []
                    
                    # 03BOTOCA_G01
                    for i in range(1, len(files_in_folder)):
                        
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_botoca_g01.append(last_known_value)
                            continue
                        except IndexError:
                            break
                            
                        try: 
                            search_value = '03BOTOCA_G01'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:

                                file_03BOTOCA_G01 = matching_row['DIPC_PRICE'].values[0]
                                list_botoca_g01.append(file_03BOTOCA_G01)
                                last_known_value = file_03BOTOCA_G01
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_botoca_g01.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_botoca_g01.append(last_known_value)
                            continue
                            
                    # 03CALACA_G01
                    for i in range(1, len(files_in_folder)):

                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_botoca_g01.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '03CALACA_G01'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:

                                file_03CALACA_G01 = matching_row['DIPC_PRICE'].values[0]
                                list_calaca_g01.append(file_03CALACA_G01)
                                last_known_value = file_03CALACA_G01
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_calaca_g01.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_calaca_g01.append(last_known_value)
                            continue
                        
                    # 03CALACA_G02
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_calaca_g02.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '03CALACA_G02'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_03CALACA_G02 = matching_row['DIPC_PRICE'].values[0]
                                list_calaca_g02.append(file_03CALACA_G02)
                                last_known_value = file_03CALACA_G02
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_calaca_g02.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_calaca_g02.append(last_known_value)
                            continue
                    
                    # 03KAL_G01
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_kal_g01.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '03KAL_G01'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_03KAL_G01 = matching_row['DIPC_PRICE'].values[0]
                                list_kal_g01.append(file_03KAL_G01)
                                last_known_value = file_03KAL_G01
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_kal_g01.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_kal_g01.append(last_known_value)
                            continue
                    
                    # 03KAL_G02
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_kal_g02.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '03KAL_G02'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_03KAL_G02 = matching_row['DIPC_PRICE'].values[0]
                                list_kal_g02.append(file_03KAL_G02)
                                last_known_value = file_03KAL_G02
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_kal_g02.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_kal_g02.append(last_known_value)
                            continue
                    
                    # 03KAL_G03
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_kal_g03.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '03KAL_G03'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_03KAL_G03 = matching_row['DIPC_PRICE'].values[0]
                                list_kal_g03.append(file_03KAL_G03)
                                last_known_value = file_03KAL_G03
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_kal_g03.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_kal_g03.append(last_known_value)
                            continue
                    
                    # 03KAL_G04
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_kal_g04.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '03KAL_G04'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_03KAL_G04 = matching_row['DIPC_PRICE'].values[0]
                                list_kal_g04.append(file_03KAL_G04)
                                last_known_value = file_03KAL_G04
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_kal_g04.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_kal_g04.append(last_known_value)
                            continue
                    
                    # 04LEYTE_A
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_leyte_a.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '04LEYTE_A'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_04LEYTE_A = matching_row['DIPC_PRICE'].values[0]
                                list_leyte_a.append(file_04LEYTE_A)
                                last_known_value = file_04LEYTE_A
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_leyte_a.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_leyte_a.append(last_known_value)
                            continue
                    
                    # 05KSPC_G01
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_kspc_g01.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '05KSPC_G01'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_05KSPC_G01 = matching_row['DIPC_PRICE'].values[0]
                                list_kspc_g01.append(file_05KSPC_G01)
                                last_known_value = file_05KSPC_G01
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_kspc_g01.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_kspc_g01.append(last_known_value)
                            continue
                    
                    # 05KSPC_G02
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_kspc_g02.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '05KSPC_G02'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_05KSPC_G02 = matching_row['DIPC_PRICE'].values[0]
                                list_kspc_g02.append(file_05KSPC_G02)
                                last_known_value = file_05KSPC_G02
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_kspc_g02.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_kspc_g02.append(last_known_value)
                            continue
                    
                    # 08BANTAP_L01
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_bantap_l01.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '08BANTAP_L01'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_08BANTAP_L01 = matching_row['DIPC_PRICE'].values[0]
                                list_bantap_l01.append(file_08BANTAP_L01)
                                last_known_value = file_08BANTAP_L01
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_bantap_l01.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_bantap_l01.append(last_known_value)
                            continue

                    # 08PEDC_T1L1
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_pedc_t1l1.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '08PEDC_T1L1'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_08PEDC_T1L1 = matching_row['DIPC_PRICE'].values[0]
                                list_pedc_t1l1.append(file_08PEDC_T1L1)
                                last_known_value = file_08PEDC_T1L1
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_pedc_t1l1.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_pedc_t1l1.append(last_known_value)
                            continue

                    # 08PEDC_T1L2
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_pedc_t1l1.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '08PEDC_T1L2'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_08PEDC_T1L2 = matching_row['DIPC_PRICE'].values[0]
                                list_pedc_t1l2.append(file_08PEDC_T1L2)
                                last_known_value = file_08PEDC_T1L2
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_pedc_t1l2.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_pedc_t1l2.append(last_known_value)
                            continue
                    
                    # 08STBAR_T1L1
                    for i in range(1, len(files_in_folder)):
                        filepath = os.path.join(folder_path, files_in_folder[i]) 

                        try: 
                            df = pd.read_csv(filepath)
                        except pd.errors.EmptyDataError:
                            if last_known_value is not None:
                                list_stbar_t1l1.append(last_known_value)
                            continue
                        except IndexError:
                            break
                        try: 
                            search_value = '08STBAR_T1L1'
                            matching_row = df[df['RESOURCE_NAME'] == search_value]
                            if not matching_row.empty:
                                file_08STBAR_T1L1 = matching_row['DIPC_PRICE'].values[0]
                                list_stbar_t1l1.append(file_08STBAR_T1L1)
                                last_known_value = file_08STBAR_T1L1
                            else:
                                # If no matching row is found, use the last known value
                                if last_known_value is not None:
                                    list_stbar_t1l1.append(last_known_value)
                        except KeyError:
                            # If RESOURCE_NAME column is missing, use the last known value
                            if last_known_value is not None:
                                list_stbar_t1l1.append(last_known_value)
                            continue

                    filepath_0 = os.path.join(folder_path, files_in_folder[0]) 
                    try:
                        df_0 = pd.read_csv(filepath_0)
                    except IndexError:
                        break
                    
                    search_value_0 = '03BOTOCA_G01'
                    matching_row_0 = df_0[df_0['RESOURCE_NAME'] == search_value_0]
                    file_03BOTOCA_G01_0 = matching_row_0['DIPC_PRICE'].values[0]
                    list_botoca_g01[-1] = file_03BOTOCA_G01_0

                    search_value_1 = '03CALACA_G01'
                    matching_row_1 = df_0[df_0['RESOURCE_NAME'] == search_value_1]
                    file_03CALACA_G01_0 = matching_row_1['DIPC_PRICE'].values[0]
                    list_calaca_g01[-1] = file_03CALACA_G01_0

                    search_value_2 = '03CALACA_G02'
                    matching_row_2 = df_0[df_0['RESOURCE_NAME'] == search_value_2]
                    file_03CALACA_G02_0 = matching_row_2['DIPC_PRICE'].values[0]
                    list_calaca_g02[-1] = file_03CALACA_G02_0

                    search_value_3 = '03KAL_G01'
                    matching_row_3 = df_0[df_0['RESOURCE_NAME'] == search_value_3]
                    file_03KAL_G01_0 = matching_row_3['DIPC_PRICE'].values[0]
                    list_kal_g01[-1] = file_03KAL_G01_0

                    search_value_4 = '03KAL_G02'
                    matching_row_4 = df_0[df_0['RESOURCE_NAME'] == search_value_4]
                    file_03KAL_G02_0 = matching_row_4['DIPC_PRICE'].values[0]
                    list_kal_g02[-1] = file_03KAL_G02_0

                    search_value_5 = '03KAL_G03'
                    matching_row_5 = df_0[df_0['RESOURCE_NAME'] == search_value_5]
                    file_03KAL_G03_0 = matching_row_5['DIPC_PRICE'].values[0]
                    list_kal_g03[-1] = file_03KAL_G03_0

                    search_value_6 = '03KAL_G04'
                    matching_row_6 = df_0[df_0['RESOURCE_NAME'] == search_value_6]
                    file_03KAL_G04_0 = matching_row_6['DIPC_PRICE'].values[0]
                    list_kal_g04[-1] = file_03KAL_G04_0

                    search_value_7 = '04LEYTE_A'
                    matching_row_7 = df_0[df_0['RESOURCE_NAME'] == search_value_7]
                    file_04LEYTE_A_0 = matching_row_7['DIPC_PRICE'].values[0]
                    list_leyte_a[-1] = file_04LEYTE_A_0

                    search_value_8 = '05KSPC_G01'
                    matching_row_8 = df_0[df_0['RESOURCE_NAME'] == search_value_8]
                    file_05KSPC_G01_0 = matching_row_8['DIPC_PRICE'].values[0]
                    list_kspc_g01[-1] = file_05KSPC_G01_0

                    search_value_9 = '05KSPC_G02'
                    matching_row_9 = df_0[df_0['RESOURCE_NAME'] == search_value_9]
                    file_05KSPC_G02_0 = matching_row_9['DIPC_PRICE'].values[0]
                    list_kspc_g02[-1] = file_05KSPC_G02_0

                    search_value_10 = '08BANTAP_L01'
                    matching_row_10 = df_0[df_0['RESOURCE_NAME'] == search_value_10]
                    file_08BANTAP_L01_0 = matching_row_10['DIPC_PRICE'].values[0]
                    list_bantap_l01[-1] = file_08BANTAP_L01_0

                    search_value_11 = '08PEDC_T1L1'
                    matching_row_11 = df_0[df_0['RESOURCE_NAME'] == search_value_11]
                    file_08PEDC_T1L1_0 = matching_row_11['DIPC_PRICE'].values[0]
                    list_pedc_t1l1[-1] = file_08PEDC_T1L1_0

                    search_value_12 = '08PEDC_T1L2'
                    matching_row_12 = df_0[df_0['RESOURCE_NAME'] ==  search_value_12]
                    file_08PEDC_T1L2_0 = matching_row_12['DIPC_PRICE'].values[0]
                    list_pedc_t1l2[-1] = file_08PEDC_T1L2_0

                    search_value_13 = '08STBAR_T1L1'
                    matching_row_13 = df_0[df_0['RESOURCE_NAME'] ==  search_value_13]
                    file_08STBAR_T1L1_0 = matching_row_13['DIPC_PRICE'].values[0]
                    list_stbar_t1l1[-1] = file_08STBAR_T1L1_0

                    average_botoca_g01 = average_every_12(list_botoca_g01)
                    average_calaca_g01 = average_every_12(list_calaca_g01)
                    average_calaca_g02 = average_every_12(list_calaca_g02)
                    average_kal_g01 = average_every_12(list_kal_g01)
                    average_kal_g02 = average_every_12(list_kal_g02)
                    average_kal_g03 = average_every_12(list_kal_g03)
                    average_kal_g04 = average_every_12(list_kal_g04)
                    average_leyte_a = average_every_12(list_leyte_a)
                    average_kspc_g01 = average_every_12(list_kspc_g01)
                    average_kspc_g02 = average_every_12(list_kspc_g02)
                    average_bantap_l01 = average_every_12(list_bantap_l01)
                    average_pedc_t1l1 = average_every_12(list_pedc_t1l1)
                    average_pedc_t1l2 = average_every_12(list_pedc_t1l2)
                    average_stbar_t1l1 = average_every_12(list_stbar_t1l1)

        current_time = datetime.now()
        hour_value = current_time.strftime("%H")

        if hour_value == "01":
            data = {
                "HOUR": ["01"],
                "03BOTOCA_G01": [average_botoca_g01[0]],
                "03CALACA_G01": [average_calaca_g01[0]],
                "03CALACA_G02": [average_calaca_g02[0]],
                "03KAL_G01": [average_kal_g01[0]], 
                "03KAL_G02": [average_kal_g02[0]], 
                "03KAL_G03": [average_kal_g03[0]],
                "03KAL_G04": [average_kal_g04[0]],
                "04LEYTE_A": [average_leyte_a[0]],
                "05KSPC_G01": [average_kspc_g01[0]],
                "05KSPC_G02": [average_kspc_g02[0]], 
                "08BANTAP_L01": [average_bantap_l01[0]],
                "08PEDC_T1L1": [average_pedc_t1l1[0]],
                "08PEDC_T1L2": [average_pedc_t1l2[0]],
                "08STBAR_T1L1": [average_stbar_t1l1[0]]
            }
        elif hour_value == "02":
            data = {
                "HOUR": ["01", "02"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1]]
            }
        elif hour_value == '03':
            data = {
                "HOUR": ["01", "02", "03"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2]]
            }
        elif hour_value == '04':
            data = {
                "HOUR": ["01", "02", "03", "04"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3]]
            }
        elif hour_value == '05':
            data = {
                "HOUR": ["01", "02", "03", "04", "05"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4]]
            }
        elif hour_value == '06':
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5]]
            }
        elif hour_value == '07':
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6]]
            }
        elif hour_value == '08':
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7]]
            }
        elif hour_value == "09":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8]]
            }
        elif hour_value == "10":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9]]
            }
        elif hour_value == "11":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10]]
            }
        elif hour_value == "12":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11]]
            }
        elif hour_value == "13":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12]]
            }
        elif hour_value == "14":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13]]
            }

        elif hour_value == "15":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14]]
            }

        elif hour_value == "16":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14], average_botoca_g01[15]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14], average_calaca_g01[15]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14], average_calaca_g02[15]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14], average_kal_g01[15]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14], average_kal_g02[15]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14], average_kal_g03[15]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14], average_kal_g04[15]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14], average_leyte_a[15]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14], average_kspc_g01[15]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14], average_kspc_g02[15]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14], average_bantap_l01[15]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14], average_pedc_t1l1[15]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14], average_pedc_t1l2[15]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14], average_stbar_t1l1[15]]
            }
        elif hour_value == "17":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14], average_botoca_g01[15], average_botoca_g01[16]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14], average_calaca_g01[15], average_calaca_g01[16]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14], average_calaca_g02[15], average_calaca_g02[16]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14], average_kal_g01[15], average_kal_g01[16]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14], average_kal_g02[15], average_kal_g02[16]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14], average_kal_g03[15], average_kal_g03[16]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14], average_kal_g04[15], average_kal_g04[16]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14], average_leyte_a[15], average_leyte_a[16]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14], average_kspc_g01[15], average_kspc_g01[16]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14], average_kspc_g02[15], average_kspc_g02[16]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14], average_bantap_l01[15], average_bantap_l01[16]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14], average_pedc_t1l1[15], average_pedc_t1l1[16]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14], average_pedc_t1l2[15], average_pedc_t1l2[16]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14], average_stbar_t1l1[15], average_stbar_t1l1[16]]
            }
        elif hour_value == "18":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14], average_botoca_g01[15], average_botoca_g01[16], average_botoca_g01[17]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14], average_calaca_g01[15], average_calaca_g01[16], average_calaca_g01[17]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14], average_calaca_g02[15], average_calaca_g02[16], average_calaca_g02[17]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14], average_kal_g01[15], average_kal_g01[16], average_kal_g01[17]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14], average_kal_g02[15], average_kal_g02[16], average_kal_g02[17]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14], average_kal_g03[15], average_kal_g03[16], average_kal_g03[17]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14], average_kal_g04[15], average_kal_g04[16], average_kal_g04[17]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14], average_leyte_a[15], average_leyte_a[16], average_leyte_a[17]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14], average_kspc_g01[15], average_kspc_g01[16], average_kspc_g01[17]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14], average_kspc_g02[15], average_kspc_g02[16], average_kspc_g02[17]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14], average_bantap_l01[15], average_bantap_l01[16], average_bantap_l01[17]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14], average_pedc_t1l1[15], average_pedc_t1l1[16], average_pedc_t1l1[17]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14], average_pedc_t1l2[15], average_pedc_t1l2[16], average_pedc_t1l2[17]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14], average_stbar_t1l1[15], average_stbar_t1l1[16], average_stbar_t1l1[17]]
            }
        elif hour_value == "19":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14], average_botoca_g01[15], average_botoca_g01[16], average_botoca_g01[17], average_botoca_g01[18]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14], average_calaca_g01[15], average_calaca_g01[16], average_calaca_g01[17], average_calaca_g01[18]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14], average_calaca_g02[15], average_calaca_g02[16], average_calaca_g02[17], average_calaca_g02[18]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14], average_kal_g01[15], average_kal_g01[16], average_kal_g01[17], average_kal_g01[18]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14], average_kal_g02[15], average_kal_g02[16], average_kal_g02[17], average_kal_g02[18]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14], average_kal_g03[15], average_kal_g03[16], average_kal_g03[17], average_kal_g03[18]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14], average_kal_g04[15], average_kal_g04[16], average_kal_g04[17], average_kal_g04[18]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14], average_leyte_a[15], average_leyte_a[16], average_leyte_a[17], average_leyte_a[18]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14], average_kspc_g01[15], average_kspc_g01[16], average_kspc_g01[17], average_kspc_g01[18]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14], average_kspc_g02[15], average_kspc_g02[16], average_kspc_g02[17], average_kspc_g02[18]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14], average_bantap_l01[15], average_bantap_l01[16], average_bantap_l01[17], average_bantap_l01[18]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14], average_pedc_t1l1[15], average_pedc_t1l1[16], average_pedc_t1l1[17], average_pedc_t1l1[18]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14], average_pedc_t1l2[15], average_pedc_t1l2[16], average_pedc_t1l2[17], average_pedc_t1l2[18]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14], average_stbar_t1l1[15], average_stbar_t1l1[16], average_stbar_t1l1[17], average_stbar_t1l1[18]]
            }
        elif hour_value == "20":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14], average_botoca_g01[15], average_botoca_g01[16], average_botoca_g01[17], average_botoca_g01[18], average_botoca_g01[19]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14], average_calaca_g01[15], average_calaca_g01[16], average_calaca_g01[17], average_calaca_g01[18], average_calaca_g01[19]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14], average_calaca_g02[15], average_calaca_g02[16], average_calaca_g02[17], average_calaca_g02[18], average_calaca_g02[19]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14], average_kal_g01[15], average_kal_g01[16], average_kal_g01[17], average_kal_g01[18], average_kal_g01[19]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14], average_kal_g02[15], average_kal_g02[16], average_kal_g02[17], average_kal_g02[18], average_kal_g02[19]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14], average_kal_g03[15], average_kal_g03[16], average_kal_g03[17], average_kal_g03[18], average_kal_g03[19]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14], average_kal_g04[15], average_kal_g04[16], average_kal_g04[17], average_kal_g04[18], average_kal_g04[19]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14], average_leyte_a[15], average_leyte_a[16], average_leyte_a[17], average_leyte_a[18], average_leyte_a[19]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14], average_kspc_g01[15], average_kspc_g01[16], average_kspc_g01[17], average_kspc_g01[18], average_kspc_g01[19]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14], average_kspc_g02[15], average_kspc_g02[16], average_kspc_g02[17], average_kspc_g02[18], average_kspc_g02[19]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14], average_bantap_l01[15], average_bantap_l01[16], average_bantap_l01[17], average_bantap_l01[18], average_bantap_l01[19]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14], average_pedc_t1l1[15], average_pedc_t1l1[16], average_pedc_t1l1[17], average_pedc_t1l1[18], average_pedc_t1l1[19]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14], average_pedc_t1l2[15], average_pedc_t1l2[16], average_pedc_t1l2[17], average_pedc_t1l2[18], average_pedc_t1l2[19]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14], average_stbar_t1l1[15], average_stbar_t1l1[16], average_stbar_t1l1[17], average_stbar_t1l1[18], average_stbar_t1l1[19]]
            }
        elif hour_value == "21":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14], average_botoca_g01[15], average_botoca_g01[16], average_botoca_g01[17], average_botoca_g01[18], average_botoca_g01[19], average_botoca_g01[20]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14], average_calaca_g01[15], average_calaca_g01[16], average_calaca_g01[17], average_calaca_g01[18], average_calaca_g01[19], average_calaca_g01[20]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14], average_calaca_g02[15], average_calaca_g02[16], average_calaca_g02[17], average_calaca_g02[18], average_calaca_g02[19], average_calaca_g02[20]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14], average_kal_g01[15], average_kal_g01[16], average_kal_g01[17], average_kal_g01[18], average_kal_g01[19], average_kal_g01[20]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14], average_kal_g02[15], average_kal_g02[16], average_kal_g02[17], average_kal_g02[18], average_kal_g02[19], average_kal_g02[20]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14], average_kal_g03[15], average_kal_g03[16], average_kal_g03[17], average_kal_g03[18], average_kal_g03[19], average_kal_g03[20]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14], average_kal_g04[15], average_kal_g04[16], average_kal_g04[17], average_kal_g04[18], average_kal_g04[19], average_kal_g04[20]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14], average_leyte_a[15], average_leyte_a[16], average_leyte_a[17], average_leyte_a[18], average_leyte_a[19], average_leyte_a[20]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14], average_kspc_g01[15], average_kspc_g01[16], average_kspc_g01[17], average_kspc_g01[18], average_kspc_g01[19], average_kspc_g01[20]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14], average_kspc_g02[15], average_kspc_g02[16], average_kspc_g02[17], average_kspc_g02[18], average_kspc_g02[19], average_kspc_g02[20]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14], average_bantap_l01[15], average_bantap_l01[16], average_bantap_l01[17], average_bantap_l01[18], average_bantap_l01[19], average_bantap_l01[20]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14], average_pedc_t1l1[15], average_pedc_t1l1[16], average_pedc_t1l1[17], average_pedc_t1l1[18], average_pedc_t1l1[19], average_pedc_t1l1[20]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14], average_pedc_t1l2[15], average_pedc_t1l2[16], average_pedc_t1l2[17], average_pedc_t1l2[18], average_pedc_t1l2[19], average_pedc_t1l2[20]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14], average_stbar_t1l1[15], average_stbar_t1l1[16], average_stbar_t1l1[17], average_stbar_t1l1[18], average_stbar_t1l1[19], average_stbar_t1l1[20]]
            }
        elif hour_value == "22":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14], average_botoca_g01[15], average_botoca_g01[16], average_botoca_g01[17], average_botoca_g01[18], average_botoca_g01[19], average_botoca_g01[20], average_botoca_g01[21]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14], average_calaca_g01[15], average_calaca_g01[16], average_calaca_g01[17], average_calaca_g01[18], average_calaca_g01[19], average_calaca_g01[20], average_calaca_g01[21]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14], average_calaca_g02[15], average_calaca_g02[16], average_calaca_g02[17], average_calaca_g02[18], average_calaca_g02[19], average_calaca_g02[20], average_calaca_g02[21]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14], average_kal_g01[15], average_kal_g01[16], average_kal_g01[17], average_kal_g01[18], average_kal_g01[19], average_kal_g01[20], average_kal_g01[21]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14], average_kal_g02[15], average_kal_g02[16], average_kal_g02[17], average_kal_g02[18], average_kal_g02[19], average_kal_g02[20], average_kal_g02[21]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14], average_kal_g03[15], average_kal_g03[16], average_kal_g03[17], average_kal_g03[18], average_kal_g03[19], average_kal_g03[20], average_kal_g03[21]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14], average_kal_g04[15], average_kal_g04[16], average_kal_g04[17], average_kal_g04[18], average_kal_g04[19], average_kal_g04[20], average_kal_g04[21]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14], average_leyte_a[15], average_leyte_a[16], average_leyte_a[17], average_leyte_a[18], average_leyte_a[19], average_leyte_a[20], average_leyte_a[21]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14], average_kspc_g01[15], average_kspc_g01[16], average_kspc_g01[17], average_kspc_g01[18], average_kspc_g01[19], average_kspc_g01[20], average_kspc_g01[21]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14], average_kspc_g02[15], average_kspc_g02[16], average_kspc_g02[17], average_kspc_g02[18], average_kspc_g02[19], average_kspc_g02[20], average_kspc_g02[21]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14], average_bantap_l01[15], average_bantap_l01[16], average_bantap_l01[17], average_bantap_l01[18], average_bantap_l01[19], average_bantap_l01[20], average_bantap_l01[21]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14], average_pedc_t1l1[15], average_pedc_t1l1[16], average_pedc_t1l1[17], average_pedc_t1l1[18], average_pedc_t1l1[19], average_pedc_t1l1[20], average_pedc_t1l1[21]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14], average_pedc_t1l2[15], average_pedc_t1l2[16], average_pedc_t1l2[17], average_pedc_t1l2[18], average_pedc_t1l2[19], average_pedc_t1l2[20], average_pedc_t1l2[21]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14], average_stbar_t1l1[15], average_stbar_t1l1[16], average_stbar_t1l1[17], average_stbar_t1l1[18], average_stbar_t1l1[19], average_stbar_t1l1[20], average_stbar_t1l1[21]]
            }
        elif hour_value == "23":
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14], average_botoca_g01[15], average_botoca_g01[16], average_botoca_g01[17], average_botoca_g01[18], average_botoca_g01[19], average_botoca_g01[20], average_botoca_g01[21], average_botoca_g01[22]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14], average_calaca_g01[15], average_calaca_g01[16], average_calaca_g01[17], average_calaca_g01[18], average_calaca_g01[19], average_calaca_g01[20], average_calaca_g01[21], average_calaca_g01[22]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14], average_calaca_g02[15], average_calaca_g02[16], average_calaca_g02[17], average_calaca_g02[18], average_calaca_g02[19], average_calaca_g02[20], average_calaca_g02[21], average_calaca_g02[22]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14], average_kal_g01[15], average_kal_g01[16], average_kal_g01[17], average_kal_g01[18], average_kal_g01[19], average_kal_g01[20], average_kal_g01[21], average_kal_g01[22]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14], average_kal_g02[15], average_kal_g02[16], average_kal_g02[17], average_kal_g02[18], average_kal_g02[19], average_kal_g02[20], average_kal_g02[21], average_kal_g02[22]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14], average_kal_g03[15], average_kal_g03[16], average_kal_g03[17], average_kal_g03[18], average_kal_g03[19], average_kal_g03[20], average_kal_g03[21], average_kal_g03[22]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14], average_kal_g04[15], average_kal_g04[16], average_kal_g04[17], average_kal_g04[18], average_kal_g04[19], average_kal_g04[20], average_kal_g04[21], average_kal_g04[22]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14], average_leyte_a[15], average_leyte_a[16], average_leyte_a[17], average_leyte_a[18], average_leyte_a[19], average_leyte_a[20], average_leyte_a[21], average_leyte_a[22]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14], average_kspc_g01[15], average_kspc_g01[16], average_kspc_g01[17], average_kspc_g01[18], average_kspc_g01[19], average_kspc_g01[20], average_kspc_g01[21], average_kspc_g01[22]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14], average_kspc_g02[15], average_kspc_g02[16], average_kspc_g02[17], average_kspc_g02[18], average_kspc_g02[19], average_kspc_g02[20], average_kspc_g02[21], average_kspc_g02[22]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14], average_bantap_l01[15], average_bantap_l01[16], average_bantap_l01[17], average_bantap_l01[18], average_bantap_l01[19], average_bantap_l01[20], average_bantap_l01[21], average_bantap_l01[22]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14], average_pedc_t1l1[15], average_pedc_t1l1[16], average_pedc_t1l1[17], average_pedc_t1l1[18], average_pedc_t1l1[19], average_pedc_t1l1[20], average_pedc_t1l1[21], average_pedc_t1l1[22]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14], average_pedc_t1l2[15], average_pedc_t1l2[16], average_pedc_t1l2[17], average_pedc_t1l2[18], average_pedc_t1l2[19], average_pedc_t1l2[20], average_pedc_t1l2[21], average_pedc_t1l2[22]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14], average_stbar_t1l1[15], average_stbar_t1l1[16], average_stbar_t1l1[17], average_stbar_t1l1[18], average_stbar_t1l1[19], average_stbar_t1l1[20], average_stbar_t1l1[21], average_stbar_t1l1[22]]
            }

        else:
            data = {
                "HOUR": ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24"],
                "03BOTOCA_G01": [average_botoca_g01[0], average_botoca_g01[1], average_botoca_g01[2], average_botoca_g01[3], average_botoca_g01[4], average_botoca_g01[5], average_botoca_g01[6], average_botoca_g01[7], average_botoca_g01[8], average_botoca_g01[9], average_botoca_g01[10], average_botoca_g01[11], average_botoca_g01[12], average_botoca_g01[13], average_botoca_g01[14], average_botoca_g01[15], average_botoca_g01[16], average_botoca_g01[17], average_botoca_g01[18], average_botoca_g01[19], average_botoca_g01[20], average_botoca_g01[21], average_botoca_g01[22], average_botoca_g01[23]],
                "03CALACA_G01": [average_calaca_g01[0], average_calaca_g01[1], average_calaca_g01[2], average_calaca_g01[3], average_calaca_g01[4], average_calaca_g01[5], average_calaca_g01[6], average_calaca_g01[7], average_calaca_g01[8], average_calaca_g01[9], average_calaca_g01[10], average_calaca_g01[11], average_calaca_g01[12], average_calaca_g01[13], average_calaca_g01[14], average_calaca_g01[15], average_calaca_g01[16], average_calaca_g01[17], average_calaca_g01[18], average_calaca_g01[19], average_calaca_g01[20], average_calaca_g01[21], average_calaca_g01[22], average_calaca_g01[23]],
                "03CALACA_G02": [average_calaca_g02[0], average_calaca_g02[1], average_calaca_g02[2], average_calaca_g02[3], average_calaca_g02[4], average_calaca_g02[5], average_calaca_g02[6], average_calaca_g02[7], average_calaca_g02[8], average_calaca_g02[9], average_calaca_g02[10], average_calaca_g02[11], average_calaca_g02[12], average_calaca_g02[13], average_calaca_g02[14], average_calaca_g02[15], average_calaca_g02[16], average_calaca_g02[17], average_calaca_g02[18], average_calaca_g02[19], average_calaca_g02[20], average_calaca_g02[21], average_calaca_g02[22], average_calaca_g02[23]],
                "03KAL_G01": [average_kal_g01[0], average_kal_g01[1], average_kal_g01[2], average_kal_g01[3], average_kal_g01[4], average_kal_g01[5], average_kal_g01[6], average_kal_g01[7], average_kal_g01[8], average_kal_g01[9], average_kal_g01[10], average_kal_g01[11], average_kal_g01[12], average_kal_g01[13], average_kal_g01[14], average_kal_g01[15], average_kal_g01[16], average_kal_g01[17], average_kal_g01[18], average_kal_g01[19], average_kal_g01[20], average_kal_g01[21], average_kal_g01[22], average_kal_g01[23]], 
                "03KAL_G02": [average_kal_g02[0], average_kal_g02[1], average_kal_g02[2], average_kal_g02[3], average_kal_g02[4], average_kal_g02[5], average_kal_g02[6], average_kal_g02[7], average_kal_g02[8], average_kal_g02[9], average_kal_g02[10], average_kal_g02[11], average_kal_g02[12], average_kal_g02[13], average_kal_g02[14], average_kal_g02[15], average_kal_g02[16], average_kal_g02[17], average_kal_g02[18], average_kal_g02[19], average_kal_g02[20], average_kal_g02[21], average_kal_g02[22], average_kal_g02[23]], 
                "03KAL_G03": [average_kal_g03[0], average_kal_g03[1], average_kal_g03[2], average_kal_g03[3], average_kal_g03[4], average_kal_g03[5], average_kal_g03[6], average_kal_g03[7], average_kal_g03[8], average_kal_g03[9], average_kal_g03[10], average_kal_g03[11], average_kal_g03[12], average_kal_g03[13], average_kal_g03[14], average_kal_g03[15], average_kal_g03[16], average_kal_g03[17], average_kal_g03[18], average_kal_g03[19], average_kal_g03[20], average_kal_g03[21], average_kal_g03[22], average_kal_g03[23]],
                "03KAL_G04": [average_kal_g04[0], average_kal_g04[1], average_kal_g04[2], average_kal_g04[3], average_kal_g04[4], average_kal_g04[5], average_kal_g04[6], average_kal_g04[7], average_kal_g04[8], average_kal_g04[9], average_kal_g04[10], average_kal_g04[11], average_kal_g04[12], average_kal_g04[13], average_kal_g04[14], average_kal_g04[15], average_kal_g04[16], average_kal_g04[17], average_kal_g04[18], average_kal_g04[19], average_kal_g04[20], average_kal_g04[21], average_kal_g04[22], average_kal_g04[23]],
                "04LEYTE_A": [average_leyte_a[0], average_leyte_a[1], average_leyte_a[2], average_leyte_a[3], average_leyte_a[4], average_leyte_a[5], average_leyte_a[6], average_leyte_a[7], average_leyte_a[8], average_leyte_a[9], average_leyte_a[10], average_leyte_a[11], average_leyte_a[12], average_leyte_a[13], average_leyte_a[14], average_leyte_a[15], average_leyte_a[16], average_leyte_a[17], average_leyte_a[18], average_leyte_a[19], average_leyte_a[20], average_leyte_a[21], average_leyte_a[22], average_leyte_a[23]],
                "05KSPC_G01": [average_kspc_g01[0], average_kspc_g01[1], average_kspc_g01[2], average_kspc_g01[3], average_kspc_g01[4], average_kspc_g01[5], average_kspc_g01[6], average_kspc_g01[7], average_kspc_g01[8], average_kspc_g01[9], average_kspc_g01[10], average_kspc_g01[11], average_kspc_g01[12], average_kspc_g01[13], average_kspc_g01[14], average_kspc_g01[15], average_kspc_g01[16], average_kspc_g01[17], average_kspc_g01[18], average_kspc_g01[19], average_kspc_g01[20], average_kspc_g01[21], average_kspc_g01[22], average_kspc_g01[23]],
                "05KSPC_G02": [average_kspc_g02[0], average_kspc_g02[1], average_kspc_g02[2], average_kspc_g02[3], average_kspc_g02[4], average_kspc_g02[5], average_kspc_g02[6], average_kspc_g02[7], average_kspc_g02[8], average_kspc_g02[9], average_kspc_g02[10], average_kspc_g02[11], average_kspc_g02[12], average_kspc_g02[13], average_kspc_g02[14], average_kspc_g02[15], average_kspc_g02[16], average_kspc_g02[17], average_kspc_g02[18], average_kspc_g02[19], average_kspc_g02[20], average_kspc_g02[21], average_kspc_g02[22], average_kspc_g02[23]], 
                "08BANTAP_L01": [average_bantap_l01[0], average_bantap_l01[1], average_bantap_l01[2], average_bantap_l01[3], average_bantap_l01[4], average_bantap_l01[5], average_bantap_l01[6], average_bantap_l01[7], average_bantap_l01[8], average_bantap_l01[9], average_bantap_l01[10], average_bantap_l01[11], average_bantap_l01[12], average_bantap_l01[13], average_bantap_l01[14], average_bantap_l01[15], average_bantap_l01[16], average_bantap_l01[17], average_bantap_l01[18], average_bantap_l01[19], average_bantap_l01[20], average_bantap_l01[21], average_bantap_l01[22], average_bantap_l01[23]],
                "08PEDC_T1L1": [average_pedc_t1l1[0], average_pedc_t1l1[1], average_pedc_t1l1[2], average_pedc_t1l1[3], average_pedc_t1l1[4], average_pedc_t1l1[5], average_pedc_t1l1[6], average_pedc_t1l1[7], average_pedc_t1l1[8], average_pedc_t1l1[9], average_pedc_t1l1[10], average_pedc_t1l1[11], average_pedc_t1l1[12], average_pedc_t1l1[13], average_pedc_t1l1[14], average_pedc_t1l1[15], average_pedc_t1l1[16], average_pedc_t1l1[17], average_pedc_t1l1[18], average_pedc_t1l1[19], average_pedc_t1l1[20], average_pedc_t1l1[21], average_pedc_t1l1[22], average_pedc_t1l1[23]],
                "08PEDC_T1L2": [average_pedc_t1l2[0], average_pedc_t1l2[1], average_pedc_t1l2[2], average_pedc_t1l2[3], average_pedc_t1l2[4], average_pedc_t1l2[5], average_pedc_t1l2[6], average_pedc_t1l2[7], average_pedc_t1l2[8], average_pedc_t1l2[9], average_pedc_t1l2[10], average_pedc_t1l2[11], average_pedc_t1l2[12], average_pedc_t1l2[13], average_pedc_t1l2[14], average_pedc_t1l2[15], average_pedc_t1l2[16], average_pedc_t1l2[17], average_pedc_t1l2[18], average_pedc_t1l2[19], average_pedc_t1l2[20], average_pedc_t1l2[21], average_pedc_t1l2[22], average_pedc_t1l2[23]],
                "08STBAR_T1L1": [average_stbar_t1l1[0], average_stbar_t1l1[1], average_stbar_t1l1[2], average_stbar_t1l1[3], average_stbar_t1l1[4], average_stbar_t1l1[5], average_stbar_t1l1[6], average_stbar_t1l1[7], average_stbar_t1l1[8], average_stbar_t1l1[9], average_stbar_t1l1[10], average_stbar_t1l1[11], average_stbar_t1l1[12], average_stbar_t1l1[13], average_stbar_t1l1[14], average_stbar_t1l1[15], average_stbar_t1l1[16], average_stbar_t1l1[17], average_stbar_t1l1[18], average_stbar_t1l1[19], average_stbar_t1l1[20], average_stbar_t1l1[21], average_stbar_t1l1[22], average_stbar_t1l1[23]]
            }
            
        # Create a DataFrame
        df_tipc = pd.DataFrame(data)

        # Melt the DataFrame to a long format
        df_long = df_tipc.melt(id_vars=['HOUR'], var_name='Resource', value_name='Value')

        # Pivot the DataFrame to prepare for hover information
        df_pivot = df_long.pivot(index='HOUR', columns='Resource', values='Value').fillna(0)
        df_pivot.reset_index(inplace=True)

        # Generate hover information
        hover_info = df_pivot.set_index('HOUR').apply(lambda x: '<br>'.join([f'{col}: {x[col]}' for col in x.index]), axis=1).to_dict()

        # Merge hover information into long DataFrame
        df_long['Hover'] = df_long['HOUR'].map(hover_info)

        # Create the interactive line chart with Plotly
        fig_tipc = px.line(df_long, x='HOUR', y='Value', color='Resource', height=200)

        # Customize hover information
        fig_tipc.update_traces(hovertemplate='%{customdata[0]}', customdata=df_long[['Hover']].values)

        fig_tipc.update_layout(
            title='Trading Interval Price Calculation',
            margin=dict(r=30, t=30, b=30),
            plot_bgcolor='black',
            paper_bgcolor='black',
            xaxis=dict(
                tickmode='array',
                tickvals=list(range(1, 25)),
                ticktext=[str(i) for i in range(1, 25)]
            ),
        )
        return fig_tipc

        # -- END OF TRADING INTERVAL PRICE CALCULATION LINE CHART --

def bcq_func():
    # Load the workbook and select the active sheet
    filepath_bcq = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\total_bcq_nomination.xlsx"
    wb_bcq = openpyxl.load_workbook(filepath_bcq)
    sheet_bcq = wb_bcq.active

    # Create lists to store the data
    hours = []
    scpc_values = []
    kspc1_values = []
    kspc2_values = []
    edc_values = []

    # Loop through column 'A' to get values for each hour, skip the header
    for row in sheet_bcq.iter_rows(min_row=2, max_row=25, min_col=1, max_col=4):
        hour = int(row[0].value)
        scpc = int(row[1].value)
        kspc = int(row[2].value)
        kspc1 = int(kspc / 2)
        kspc2 = int(kspc / 2)
        edc = int(row[3].value)

        # Store the values in the lists
        hours.append(hour)
        scpc_values.append(scpc)
        kspc1_values.append(kspc1)
        kspc2_values.append(kspc2)
        edc_values.append(edc)

    # Create a DataFrame with proper structure
    data_bcq = {
        'Hour': hours,
        'SEMCAL': scpc_values,
        'KSPC1': kspc1_values,
        'KSPC2': kspc2_values,
        'EDC': edc_values
    }

    df_bcq = pd.DataFrame(data_bcq)

    # Melt the DataFrame for Plotly
    df_bcq_melted = df_bcq.melt('Hour', var_name='Contracted Supplier', value_name='BCQ')

    # Define the color map
    color_map = {
        'SEMCAL': 'orange', 
        'KSPC1': 'red', 
        'KSPC2': 'darkblue', 
        'EDC': 'green'
    }

    # Create a Plotly area chart
    fig_bcq = px.area(df_bcq_melted, x='Hour', y='BCQ', color='Contracted Supplier', 
                labels={'Hour': 'Hour', 'BCQ': 'BCQ'}, color_discrete_map=color_map, # Apply the color map
                height=200)

    # Update layout to match the Altair configuration
    fig_bcq.update_layout(
        title='BCQ Nominations per Supplier',
        margin=dict(r=30, t=30, b=30),
        plot_bgcolor='black',  # Set background color to black
        paper_bgcolor='black',  # Set paper background color to black
        xaxis_title='Interval',
        yaxis_title='kW',
        legend_title='Contracted Supplier'
    )

    # Update y-axis to show full number format
    fig_bcq.update_yaxes(tickformat=',')  # Use ',' to show numbers without abbreviating

    return fig_bcq 

        # -- END OF DISPLAYING THE BCQ NOMINATIONS PER SUPPLIER AREA CHART --

def genmix_func():
    # -- START OF DISPLAYING THE GENERATION MIX DONUT CHART --
        
        actual_energy_excel = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\actual_energy.xlsx"
        total_bcq_nomination_excel = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\total_bcq_nomination.xlsx"
        wesm_exposure_excel = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\wesm_exposure.xlsx"

        # Generation Mix Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'ACTUAL_ENERGY'
        sheet['B1'] = 'TOTAL_BCQ'
        sheet['C1'] = 'SCPC'
        sheet['D1'] = 'KSPC'
        sheet['E1'] = 'KSPC1'
        sheet['F1'] = 'KSPC2'
        sheet['G1'] = 'EDC'
        sheet['H1'] = 'WESM'
        sheet['I1'] = 'HOUR'
        sheet['I2'] = 1
        sheet['I3'] = 2
        sheet['I4'] = 3
        sheet['I5'] = 4
        sheet['I6'] = 5
        sheet['I7'] = 6
        sheet['I8'] = 7
        sheet['I9'] = 8
        sheet['I10'] = 9
        sheet['I11'] = 10
        sheet['I12'] = 11
        sheet['I13'] = 12
        sheet['I14'] = 13
        sheet['I15'] = 14
        sheet['I16'] = 15
        sheet['I17'] = 16
        sheet['I18'] = 17
        sheet['I19'] = 18
        sheet['I20'] = 19
        sheet['I21'] = 20
        sheet['I22'] = 21
        sheet['I23'] = 22
        sheet['I24'] = 23
        sheet['I25'] = 24
        workbook.save('generation_mix.xlsx')
        workbook.close()
        print("Excel file 'generation_mix.xlsx' created successfully.")

        source_wb_genmix_1 = load_workbook(actual_energy_excel) 
        source_sheet_genmix_1 = source_wb_genmix_1['Sheet']
        
        genmix_filepath = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\generation_mix.xlsx"
        dest_wb_genmix = load_workbook(genmix_filepath)

        dest_sheet_genmix = dest_wb_genmix['Sheet']

        source_ranges_genmix_1 = [('D2', 'D25')]
        dest_ranges_genmix_1 = [('A2', 'A25')]

        for i in range(len(source_ranges_genmix_1)):
            copy_values(source_sheet_genmix_1, dest_sheet_genmix, source_ranges_genmix_1[i], dest_ranges_genmix_1[i])

        # Save the destination workbook
        dest_wb_genmix.save(genmix_filepath)

        source_wb_genmix_2 = load_workbook(total_bcq_nomination_excel)
        source_sheet_genmix_2 = source_wb_genmix_2['Sheet']

        source_ranges_genmix_2 = [('E2', 'E25'), ('B2', 'B25'), ('C2', 'C25'), ('D2', 'D25')]
        dest_ranges_genmix_2 = [('B2', 'B25'), ('C2', 'C25'), ('D2', 'D25'), ('G2', 'G25')]

        for i in range(len(source_ranges_genmix_2)):
            copy_values(source_sheet_genmix_2, dest_sheet_genmix, source_ranges_genmix_2[i], dest_ranges_genmix_2[i])

        # Save the destination workbook
        dest_wb_genmix.save(genmix_filepath)

        kspc_value_1 = source_sheet_genmix_2['C2'].value
        half_value_1 = int(kspc_value_1 / 2)
        dest_sheet_genmix['E2'] = half_value_1
        dest_sheet_genmix['F2'] = half_value_1

        kspc_value_2 = source_sheet_genmix_2['C3'].value
        half_value_2 = int(kspc_value_2 / 2)
        dest_sheet_genmix['E3'] = half_value_2
        dest_sheet_genmix['F3'] = half_value_2

        kspc_value_3 = source_sheet_genmix_2['C4'].value
        half_value_3 = int(kspc_value_3 / 2)
        dest_sheet_genmix['E4'] = half_value_3
        dest_sheet_genmix['F4'] = half_value_3

        kspc_value_4 = source_sheet_genmix_2['C5'].value
        half_value_4 = int(kspc_value_4 / 2)
        dest_sheet_genmix['E5'] = half_value_4
        dest_sheet_genmix['F5'] = half_value_4

        kspc_value_5 = source_sheet_genmix_2['C6'].value
        half_value_5 = int(kspc_value_5 / 2)
        dest_sheet_genmix['E6'] = half_value_5
        dest_sheet_genmix['F6'] = half_value_5

        kspc_value_6 = source_sheet_genmix_2['C7'].value
        half_value_6 = int(kspc_value_6 / 2)
        dest_sheet_genmix['E7'] = half_value_6
        dest_sheet_genmix['F7'] = half_value_6

        kspc_value_7 = source_sheet_genmix_2['C8'].value
        half_value_7 = int(kspc_value_7 / 2)
        dest_sheet_genmix['E8'] = half_value_7
        dest_sheet_genmix['F8'] = half_value_7

        kspc_value_8 = source_sheet_genmix_2['C9'].value
        half_value_8 = int(kspc_value_8 / 2)
        dest_sheet_genmix['E9'] = half_value_8
        dest_sheet_genmix['F9'] = half_value_8

        kspc_value_9 = source_sheet_genmix_2['C10'].value
        half_value_9 = int(kspc_value_9 / 2)
        dest_sheet_genmix['E10'] = half_value_9
        dest_sheet_genmix['F10'] = half_value_9

        kspc_value_10 = source_sheet_genmix_2['C11'].value
        half_value_10 = int(kspc_value_10 / 2)
        dest_sheet_genmix['E11'] = half_value_10
        dest_sheet_genmix['F11'] = half_value_10

        kspc_value_11 = source_sheet_genmix_2['C12'].value
        half_value_11 = int(kspc_value_11 / 2)
        dest_sheet_genmix['E12'] = half_value_11
        dest_sheet_genmix['F12'] = half_value_11

        kspc_value_12 = source_sheet_genmix_2['C13'].value
        half_value_12 = int(kspc_value_12 / 2)
        dest_sheet_genmix['E13'] = half_value_12
        dest_sheet_genmix['F13'] = half_value_12

        kspc_value_13 = source_sheet_genmix_2['C14'].value
        half_value_13 = int(kspc_value_13 / 2)
        dest_sheet_genmix['E14'] = half_value_13
        dest_sheet_genmix['F14'] = half_value_13

        kspc_value_14 = source_sheet_genmix_2['C15'].value
        half_value_14 = int(kspc_value_14 / 2)
        dest_sheet_genmix['E15'] = half_value_14
        dest_sheet_genmix['F15'] = half_value_14

        kspc_value_15 = source_sheet_genmix_2['C16'].value
        half_value_15 = int(kspc_value_15 / 2)
        dest_sheet_genmix['E16'] = half_value_15
        dest_sheet_genmix['F16'] = half_value_15

        kspc_value_16 = source_sheet_genmix_2['C17'].value
        half_value_16 = int(kspc_value_16 / 2)
        dest_sheet_genmix['E17'] = half_value_16
        dest_sheet_genmix['F17'] = half_value_16

        kspc_value_17 = source_sheet_genmix_2['C18'].value
        half_value_17 = int(kspc_value_17 / 2)
        dest_sheet_genmix['E18'] = half_value_17
        dest_sheet_genmix['F18'] = half_value_17

        kspc_value_18 = source_sheet_genmix_2['C19'].value
        half_value_18 = int(kspc_value_18 / 2)
        dest_sheet_genmix['E19'] = half_value_18
        dest_sheet_genmix['F19'] = half_value_18

        kspc_value_19 = source_sheet_genmix_2['C20'].value
        half_value_19 = int(kspc_value_19 / 2)
        dest_sheet_genmix['E20'] = half_value_19
        dest_sheet_genmix['F20'] = half_value_19

        kspc_value_20 = source_sheet_genmix_2['C21'].value
        half_value_20 = int(kspc_value_20 / 2)
        dest_sheet_genmix['E21'] = half_value_20
        dest_sheet_genmix['F21'] = half_value_20

        kspc_value_21 = source_sheet_genmix_2['C22'].value
        half_value_21 = int(kspc_value_21 / 2)
        dest_sheet_genmix['E22'] = half_value_21
        dest_sheet_genmix['F22'] = half_value_21

        kspc_value_22 = source_sheet_genmix_2['C23'].value
        half_value_22 = int(kspc_value_22 / 2)
        dest_sheet_genmix['E23'] = half_value_22
        dest_sheet_genmix['F23'] = half_value_22

        kspc_value_23 = source_sheet_genmix_2['C24'].value
        half_value_23 = int(kspc_value_23 / 2)
        dest_sheet_genmix['E24'] = half_value_23
        dest_sheet_genmix['F24'] = half_value_23

        kspc_value_24 = source_sheet_genmix_2['C25'].value
        half_value_24 = int(kspc_value_24 / 2)
        dest_sheet_genmix['E25'] = half_value_24
        dest_sheet_genmix['F25'] = half_value_24

        # Save the destination workbook
        dest_wb_genmix.save(genmix_filepath)

        source_wb_genmix_3 = load_workbook(wesm_exposure_excel)
        source_sheet_genmix_3 = source_wb_genmix_3['Sheet']
        
        source_ranges_genmix_3 = [('E2', 'E25')]
        dest_ranges_genmix_3 = [('H2', 'H25')]

        for i in range(len(source_ranges_genmix_3)):
            copy_values(source_sheet_genmix_3, dest_sheet_genmix, source_ranges_genmix_3[i], dest_ranges_genmix_3[i])

        # Save the destination workbook
        dest_wb_genmix.save(genmix_filepath)

        # CASE A. ACTUAL ENERGY <= BCQ
        # SCPC, KSPC1, KSPC2, EDC

        # CASE B. ACTUAL ENERGY > BCQ
        # SCPC, KSPC1, KSPC2, EDC, WESM

        now = datetime.now()
        current_hour = now.hour
        if current_hour == 0:
            current_hour = 24

        # Initialize empty list containing Generation Mix values
        generation_mix_list = []

        # HOUR 1
        if current_hour == dest_sheet_genmix['I2'].value:
            if dest_sheet_genmix['A2'].value <= dest_sheet_genmix['B2'].value:
                generation_mix_list.append(dest_sheet_genmix['C2'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E2'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F2'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G2'].value) #EDC

            elif dest_sheet_genmix['A2'].value > dest_sheet_genmix['B2'].value:
                generation_mix_list.append(dest_sheet_genmix['C2'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E2'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F2'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G2'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H2'].value) #WESM

        # HOUR 2
        elif current_hour == dest_sheet_genmix['I3'].value:
            if dest_sheet_genmix['A3'].value <= dest_sheet_genmix['B3'].value:
                generation_mix_list.append(dest_sheet_genmix['C3'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E3'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F3'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G3'].value) #EDC

            elif dest_sheet_genmix['A3'].value > dest_sheet_genmix['B3'].value:
                generation_mix_list.append(dest_sheet_genmix['C3'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E3'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F3'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G3'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H3'].value) #WESM

        # HOUR 3
        elif current_hour == dest_sheet_genmix['I4'].value:
            if dest_sheet_genmix['A4'].value <= dest_sheet_genmix['B4'].value:
                generation_mix_list.append(dest_sheet_genmix['C4'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E4'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F4'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G4'].value) #EDC

            elif dest_sheet_genmix['A4'].value > dest_sheet_genmix['B4'].value:
                generation_mix_list.append(dest_sheet_genmix['C4'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E4'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F4'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G4'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H4'].value) #WESM

        # HOUR 4
        elif current_hour == dest_sheet_genmix['I5'].value:
            if dest_sheet_genmix['A5'].value <= dest_sheet_genmix['B5'].value:
                generation_mix_list.append(dest_sheet_genmix['C5'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E5'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F5'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G5'].value) #EDC

            elif dest_sheet_genmix['A5'].value > dest_sheet_genmix['B5'].value:
                generation_mix_list.append(dest_sheet_genmix['C5'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E5'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F5'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G5'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H5'].value) #WESM
        
        # HOUR 5
        elif current_hour == dest_sheet_genmix['I6'].value:
            if dest_sheet_genmix['A6'].value <= dest_sheet_genmix['B6'].value:
                generation_mix_list.append(dest_sheet_genmix['C6'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E6'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F6'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G6'].value) #EDC

            elif dest_sheet_genmix['A6'].value > dest_sheet_genmix['B6'].value:
                generation_mix_list.append(dest_sheet_genmix['C6'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E6'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F6'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G6'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H6'].value) #WESM

        # HOUR 6  
        elif current_hour == dest_sheet_genmix['I7'].value:
            if dest_sheet_genmix['A7'].value <= dest_sheet_genmix['B7'].value:
                generation_mix_list.append(dest_sheet_genmix['C7'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E7'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F7'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G7'].value) #EDC

            elif dest_sheet_genmix['A7'].value > dest_sheet_genmix['B7'].value:
                generation_mix_list.append(dest_sheet_genmix['C7'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E7'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F7'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G7'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H7'].value) #WESM
        
        # HOUR 7
        elif current_hour == dest_sheet_genmix['I8'].value:
            if dest_sheet_genmix['A8'].value <= dest_sheet_genmix['B8'].value:
                generation_mix_list.append(dest_sheet_genmix['C8'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E8'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F8'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G8'].value) #EDC

            elif dest_sheet_genmix['A8'].value > dest_sheet_genmix['B8'].value:
                generation_mix_list.append(dest_sheet_genmix['C8'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E8'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F8'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G8'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H8'].value) #WESM
        
        # HOUR 8
        elif current_hour == dest_sheet_genmix['I9'].value:
            if dest_sheet_genmix['A9'].value <= dest_sheet_genmix['B9'].value:
                generation_mix_list.append(dest_sheet_genmix['C9'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E9'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F9'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G9'].value) #EDC

            elif dest_sheet_genmix['A9'].value > dest_sheet_genmix['B9'].value:
                generation_mix_list.append(dest_sheet_genmix['C9'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E9'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F9'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G9'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H9'].value) #WESM

        # HOUR 9
        elif current_hour == dest_sheet_genmix['I10'].value:
            if dest_sheet_genmix['A10'].value <= dest_sheet_genmix['B10'].value:
                generation_mix_list.append(dest_sheet_genmix['C10'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E10'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F10'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G10'].value) #EDC

            elif dest_sheet_genmix['A10'].value > dest_sheet_genmix['B10'].value:
                generation_mix_list.append(dest_sheet_genmix['C10'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E10'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F10'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G10'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H10'].value) #WESM
        
        # HOUR 10
        elif current_hour == dest_sheet_genmix['I11'].value:
            if dest_sheet_genmix['A11'].value <= dest_sheet_genmix['B11'].value:
                generation_mix_list.append(dest_sheet_genmix['C11'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E11'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F11'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G11'].value) #EDC

            elif dest_sheet_genmix['A11'].value > dest_sheet_genmix['B11'].value:
                generation_mix_list.append(dest_sheet_genmix['C11'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E11'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F11'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G11'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H11'].value) #WESM

        # HOUR 11
        elif current_hour == dest_sheet_genmix['I12'].value:
            if dest_sheet_genmix['A12'].value <= dest_sheet_genmix['B12'].value:
                generation_mix_list.append(dest_sheet_genmix['C12'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E12'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F12'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G12'].value) #EDC

            elif dest_sheet_genmix['A12'].value > dest_sheet_genmix['B12'].value:
                generation_mix_list.append(dest_sheet_genmix['C12'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E12'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F12'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G12'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H12'].value) #WESM

        # HOUR 12
        elif current_hour == dest_sheet_genmix['I13'].value:
            if dest_sheet_genmix['A13'].value <= dest_sheet_genmix['B13'].value:
                generation_mix_list.append(dest_sheet_genmix['C13'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E13'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F13'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G13'].value) #EDC

            elif dest_sheet_genmix['A13'].value > dest_sheet_genmix['B13'].value:
                generation_mix_list.append(dest_sheet_genmix['C13'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E13'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F13'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G13'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H13'].value) #WESM
        
        # HOUR 13
        elif current_hour == dest_sheet_genmix['I14'].value:
            if dest_sheet_genmix['A14'].value <= dest_sheet_genmix['B14'].value:
                generation_mix_list.append(dest_sheet_genmix['C14'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E14'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F14'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G14'].value) #EDC

            elif dest_sheet_genmix['A14'].value > dest_sheet_genmix['B14'].value:
                generation_mix_list.append(dest_sheet_genmix['C14'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E14'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F14'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G14'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H14'].value) #WESM
        
        # HOUR 14
        elif current_hour == dest_sheet_genmix['I15'].value:
            if dest_sheet_genmix['A15'].value <= dest_sheet_genmix['B15'].value:
                generation_mix_list.append(dest_sheet_genmix['C15'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E15'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F15'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G15'].value) #EDC

            elif dest_sheet_genmix['A15'].value > dest_sheet_genmix['B15'].value:
                generation_mix_list.append(dest_sheet_genmix['C15'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E15'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F15'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G15'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H15'].value) #WESM
        
        # HOUR 15
        elif current_hour == dest_sheet_genmix['I16'].value:
            if dest_sheet_genmix['A16'].value <= dest_sheet_genmix['B16'].value:
                generation_mix_list.append(dest_sheet_genmix['C16'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E16'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F16'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G16'].value) #EDC

            elif dest_sheet_genmix['A16'].value > dest_sheet_genmix['B16'].value:
                generation_mix_list.append(dest_sheet_genmix['C16'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E16'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F16'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G16'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H16'].value) #WESM
        
        # HOUR 16
        elif current_hour == dest_sheet_genmix['I17'].value:
            if dest_sheet_genmix['A17'].value <= dest_sheet_genmix['B17'].value:
                generation_mix_list.append(dest_sheet_genmix['C17'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E17'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F17'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G17'].value) #EDC

            elif dest_sheet_genmix['A17'].value > dest_sheet_genmix['B17'].value:
                generation_mix_list.append(dest_sheet_genmix['C17'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E17'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F17'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G17'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H17'].value) #WESM
        
        # HOUR 17
        elif current_hour == dest_sheet_genmix['I18'].value:
            if dest_sheet_genmix['A18'].value <= dest_sheet_genmix['B18'].value:
                generation_mix_list.append(dest_sheet_genmix['C18'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E18'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F18'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G18'].value) #EDC

            elif dest_sheet_genmix['A18'].value > dest_sheet_genmix['B18'].value:
                generation_mix_list.append(dest_sheet_genmix['C18'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E18'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F18'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G18'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H18'].value) #WESM
        
        # HOUR 18
        elif current_hour == dest_sheet_genmix['I19'].value:
            if dest_sheet_genmix['A19'].value <= dest_sheet_genmix['B19'].value:
                generation_mix_list.append(dest_sheet_genmix['C19'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E19'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F19'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G19'].value) #EDC

            elif dest_sheet_genmix['A19'].value > dest_sheet_genmix['B19'].value:
                generation_mix_list.append(dest_sheet_genmix['C19'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E19'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F19'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G19'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H19'].value) #WESM
        
        # HOUR 19
        elif current_hour == dest_sheet_genmix['I20'].value:
            if dest_sheet_genmix['A20'].value <= dest_sheet_genmix['B20'].value:
                generation_mix_list.append(dest_sheet_genmix['C20'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E20'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F20'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G20'].value) #EDC

            elif dest_sheet_genmix['A20'].value > dest_sheet_genmix['B20'].value:
                generation_mix_list.append(dest_sheet_genmix['C20'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E20'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F20'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G20'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H20'].value) #WESM

        # HOUR 20
        elif current_hour == dest_sheet_genmix['I21'].value:
            if dest_sheet_genmix['A21'].value <= dest_sheet_genmix['B21'].value:
                generation_mix_list.append(dest_sheet_genmix['C21'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E21'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F21'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G21'].value) #EDC

            elif dest_sheet_genmix['A21'].value > dest_sheet_genmix['B21'].value:
                generation_mix_list.append(dest_sheet_genmix['C21'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E21'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F21'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G21'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H21'].value) #WESM
        
        # HOUR 21
        elif current_hour == dest_sheet_genmix['I22'].value:
            if dest_sheet_genmix['A22'].value <= dest_sheet_genmix['B22'].value:
                generation_mix_list.append(dest_sheet_genmix['C22'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E22'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F22'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G22'].value) #EDC

            elif dest_sheet_genmix['A22'].value > dest_sheet_genmix['B22'].value:
                generation_mix_list.append(dest_sheet_genmix['C22'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E22'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F22'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G22'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H22'].value) #WESM
        
        # HOUR 22
        elif current_hour == dest_sheet_genmix['I23'].value:
            if dest_sheet_genmix['A23'].value <= dest_sheet_genmix['B23'].value:
                generation_mix_list.append(dest_sheet_genmix['C23'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E23'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F23'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G23'].value) #EDC

            elif dest_sheet_genmix['A23'].value > dest_sheet_genmix['B23'].value:
                generation_mix_list.append(dest_sheet_genmix['C23'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E23'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F23'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G23'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H23'].value) #WESM
        
        # HOUR 23
        elif current_hour == dest_sheet_genmix['I24'].value:
            if dest_sheet_genmix['A24'].value <= dest_sheet_genmix['B24'].value:
                generation_mix_list.append(dest_sheet_genmix['C24'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E24'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F24'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G24'].value) #EDC

            elif dest_sheet_genmix['A24'].value > dest_sheet_genmix['B24'].value:
                generation_mix_list.append(dest_sheet_genmix['C24'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E24'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F24'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G24'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H24'].value) #WESM

        # HOUR 24
        elif current_hour == dest_sheet_genmix['I25'].value:
            if dest_sheet_genmix['A25'].value <= dest_sheet_genmix['B25'].value:
                generation_mix_list.append(dest_sheet_genmix['C25'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E25'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F25'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G25'].value) #EDC

            elif dest_sheet_genmix['A25'].value > dest_sheet_genmix['B25'].value:
                generation_mix_list.append(dest_sheet_genmix['C25'].value) #SCPC
                generation_mix_list.append(dest_sheet_genmix['E25'].value) #KSPC1
                generation_mix_list.append(dest_sheet_genmix['F25'].value) #KSPC2
                generation_mix_list.append(dest_sheet_genmix['G25'].value) #EDC
                generation_mix_list.append(dest_sheet_genmix['H25'].value) #WESM
        
        if len(generation_mix_list) == 4:
            genmix_data = {
            'Names': ['SEMCAL', 'KSPC1', 'KSPC2', 'EDC'],
            'Values': generation_mix_list
            }

        elif len(generation_mix_list) == 5:
            genmix_data = {
            'Names': ['SEMCAL', 'KSPC1', 'KSPC2', 'EDC', 'WESM'],
            'Values': generation_mix_list
            }

        genmix_df = pd.DataFrame(genmix_data)

        # Creating the donut chart
        fig_genmix = px.pie(genmix_df, values='Values', names='Names', hole=0.5)

        # Customizing the labels to appear on the slices
        fig_genmix.update_traces(
            textposition='outside',    # Places the text inside the slices
            textinfo='label+percent', # Shows both the label and percentage inside the slices
            insidetextorientation='horizontal',  # Keeps text horizontal inside the slices
            marker=dict(
                colors=['orangered', 'red', 'darkblue', 'green', 'yellow']
            )
        )

        fig_genmix.update_layout(
            height=200
        )

        fig_genmix.update_layout(
            title='Generation Mix',
            margin=dict(r=40, t=40, b=40),
            plot_bgcolor='black',
            paper_bgcolor='black'
        )

        return fig_genmix

        # -- END OF DISPLAYING THE GENERATION MIX DONUT CHART --

# -- END OF FUNCTIONS --
# Set the page layout to wide
st.set_page_config(
    page_title="MEPC Energy Trading Dashboard",
    page_icon="data/logo_only.png",
    layout="wide"
)

# -- USER AUTHENTICATION --
names = ["MORE Electric and Power Corporation", "Niel Parcon", "Roel Castro"]
usernames = ["more", "nparcon", "rcastro"]

file_path = Path(__file__).parent / "hashed_pw.pkl"
if file_path.exists():
    with file_path.open("rb") as file:
        hashed_passwords = pickle.load(file)
else:
    st.error("Hashed passwords file not found.")
    st.stop()

credentials = {
    "usernames": {
        usernames[i]: {"name": names[i], "password": hashed_passwords[i]}
        for i in range(len(usernames))
    }
}

authenticator = stauth.Authenticate(
    credentials,
    "mepc_energy_trading_dashboard",
    "morepower",
    cookie_expiry_days=0
)


name, authentication_status, username = authenticator.login("main")

if authentication_status == False:
    st.error("Wrong username/password!")

if authentication_status is None:
    st.warning("Please enter your username and password.")

if authentication_status == True:
    while True: 
        # Logout Button
        authenticator.logout("Logout", "main")

        # Add custom CSS
        st.markdown(
            """
            <style>
            /* This targets the main block content within a container */
            .main .block-container {
                padding-left: 13px;    /* Adjust left padding */
                padding-right: 13px;   /* Adjust right padding */
                padding-top: 0px;
                padding-bottom: 13px;
            }
            /* Additional styles to target specific elements within the container if needed */
            .stContainer {
                padding-left: 13px;    /* Adjust left padding */
                padding-right: 13px;   /* Adjust right padding */
                padding-top: 0rem;
                padding-bottom: 0rem;
            }
            </style>
            """,
            unsafe_allow_html=True
        )

        margins_css = """
        <style>
            .main > div {
                padding-left: 0rem;
                padding-right: 0rem;
            }
        </style>
    """

        st.markdown(margins_css, unsafe_allow_html=True)

        # Path to your image
        image_path = 'data/MORE_Power_Logo.png'

        # Hide the Streamlit header and footer
        st.markdown("""
            <style>
            /* Hide Streamlit header */
            header {visibility: hidden;}
            /* Hide Streamlit footer */
            footer {visibility: hidden;}
            </style>
            """, unsafe_allow_html=True)

        # Connection
        conn=mysql.connector.connect(
            host = "localhost",
            port = "3306",
            user = "root",
            passwd = "",
            db = "myDb"
        )

        c=conn.cursor()

        # -- START OF SO ADVISORIES HEADER --

        # Load your Excel file
        excel_file = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\SO_ADVISORIES.xlsx"
        df = pd.read_excel(excel_file)

        # Convert TIME_MESSAGE to datetime format (if it's not already)
        df['TIME_MESSAGE'] = pd.to_datetime(df['TIME_MESSAGE'], errors='coerce')

        # Drop rows where TIME_MESSAGE couldn't be converted to datetime (if needed)
        df = df.dropna(subset=['TIME_MESSAGE'])

        # Sort dataframe by TIME_MESSAGE in descending order
        df = df.sort_values(by='TIME_MESSAGE', ascending=False)

        # Prepare the initial data for scrolling
        scrollable_content = []
        latest_10_entries = fetch_latest_10_entries()
        for index, row in latest_10_entries.iterrows():
            time_message = row['TIME_MESSAGE'].strftime('%Y/%m/%d %H:%M:%S')
            message = row['MESSAGE']
            entry_text = f"{time_message}: {message}"
            scrollable_content.append(entry_text)       

        # -- END OF SO ADVISORIES HEADER --

        # -- START OF MO ADVISORIES HEADER --

        # Load your Excel file
        excel_file = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\MO_ADVISORIES.xlsx"
        df = pd.read_excel(excel_file)

        # Sort dataframe by TIMESTAMP in descending order
        df = df.sort_values(by='TIMESTAMP', ascending=False)

        # Prepare the initial data for scrolling
        scrollable_content_2 = []
        latest_entry = fetch_latest_entry()
        for index, row in latest_entry.iterrows():
            time_message_2 = row['TIMESTAMP']
            message_2 = row['ADVISORY']
            entry_text_2 = f"{time_message_2} - {message_2}"
            scrollable_content_2.append(entry_text_2)

        # -- END OF MO ADVISORIES HEADER --

        # # HTML and CSS for the tickers
        # html_code = f"""
        # <style>
        # * {{
        #     box-sizing: border-box;
        # }}
        # @-webkit-keyframes ticker {{
        #     0% {{
        #         -webkit-transform: translate3d(100%, 0, 0);
        #         transform: translate3d(100%, 0, 0);
        #     }}
        #     100% {{
        #         -webkit-transform: translate3d(-100%, 0, 0);
        #         transform: translate3d(-100%, 0, 0);
        #     }}
        # }}
        # @keyframes ticker {{
        #     0% {{
        #         -webkit-transform: translate3d(100%, 0, 0);
        #         transform: translate3d(100%, 0, 0);
        #     }}
        #     100% {{
        #         -webkit-transform: translate3d(-100%, 0, 0);
        #         transform: translate3d(-100%, 0, 0);
        #     }}
        # }}
        # .ticker-wrap {{
        #     position: fixed;
        #     width: calc(100% - 2px);  /* Adjust width to account for border */
        #     overflow: hidden;
        #     height: 2rem;
        #     background-color: #000000;
        #     box-sizing: border-box;
        #     border: 2px solid white;  /* White border around the entire box */
        #     border-radius: 5px;  /* Optional: Add rounded corners */
        #     padding: 0;  /* Remove padding to avoid clipping */
        #     margin: 0;  /* Remove margin to avoid shifting */
        # }}
        # .ticker1 {{
        #     display: inline-block;
        #     height: 1.8rem;
        #     line-height: 1.8rem;
        #     white-space: nowrap;
        #     box-sizing: content-box;
        #     -webkit-animation-iteration-count: infinite;
        #     animation-iteration-count: infinite;
        #     -webkit-animation-timing-function: linear;
        #     animation-timing-function: linear;
        #     -webkit-animation-name: ticker;
        #     animation-name: ticker;
        #     -webkit-animation-duration: 200s;
        #     animation-duration: 200s;
        # }}
        # .ticker2 {{
        #     display: inline-block;
        #     height: 1.8rem;
        #     line-height: 1.8rem;
        #     white-space: nowrap;
        #     box-sizing: content-box;
        #     -webkit-animation-iteration-count: infinite;
        #     animation-iteration-count: infinite;
        #     -webkit-animation-timing-function: linear;
        #     animation-timing-function: linear;
        #     -webkit-animation-name: ticker;
        #     animation-name: ticker;
        #     -webkit-animation-duration: 40s;
        #     animation-duration: 40s;
        # }}
        # .ticker__item {{
        #     display: inline-block;
        #     padding: 0 2rem; /* Increased padding for more space between entries */
        #     font-size: 0.8rem;
        #     font-family: Helvetica;
        #     font-weight: bold;
        #     color: #FFFFFF;
        # }}
        # .ticker__item:first-child {{
        #     margin-top: 0;
        # }}
        # body {{ margin: 0; padding-bottom: 0; }}
        # h1, h2, p {{ padding: 0 5%; }}
        # </style>

        # <div class="ticker-wrap" style="bottom: 2.5rem;">  <!-- Bottom ticker -->
        #     <div class="ticker ticker1">
        # """

        # for entry in scrollable_content:
        #     html_code += f'<div class="ticker__item">{entry}</div>'

        # html_code += """
        #     </div>
        # </div>

        # <div class="ticker-wrap" style="bottom: 0;">  <!-- Top ticker -->
        #     <div class="ticker ticker2">
        # """

        # for entry in scrollable_content_2:
        #     html_code += f'<div class="ticker__item">{entry}</div>'

        # html_code += """
        #     </div>
        # </div>
        # """

        # HTML and CSS for the tickers
        html_code = f"""
        <style>
        * {{
            box-sizing: border-box;
        }}
        @-webkit-keyframes ticker {{
            0% {{
                -webkit-transform: translate3d(100%, 0, 0);
                transform: translate3d(100%, 0, 0);
            }}
            100% {{
                -webkit-transform: translate3d(-100%, 0, 0);
                transform: translate3d(-100%, 0, 0);
            }}
        }}
        @keyframes ticker {{
            0% {{
                -webkit-transform: translate3d(100%, 0, 0);
                transform: translate3d(100%, 0, 0);
            }}
            100% {{
                -webkit-transform: translate3d(-100%, 0, 0);
                transform: translate3d(-100%, 0, 0);
            }}
        }}
        .ticker-wrap {{
            position: fixed;
            width: calc(100% - 2px);  /* Adjust width to account for border */
            overflow: hidden;
            height: 2rem;
            background-color: #000000;
            box-sizing: border-box;
            border: 2px solid white;  /* White border around the entire box */
            border-radius: 5px;  /* Optional: Add rounded corners */
            padding: 0;  /* Remove padding to avoid clipping */
            margin: 0;  /* Remove margin to avoid shifting */
        }}
        .ticker1 {{
            display: inline-block;
            height: 1.8rem;
            line-height: 1.8rem;
            white-space: nowrap;
            box-sizing: content-box;
            -webkit-animation-iteration-count: infinite;
            animation-iteration-count: infinite;
            -webkit-animation-timing-function: linear;
            animation-timing-function: linear;
            -webkit-animation-name: ticker;
            animation-name: ticker;
            -webkit-animation-duration: 280s;
            animation-duration: 280s;
            transform: translate3d(100%, 0, 0); /* Start position to ensure it's visible immediately */
        }}
        .ticker2 {{
            display: inline-block;
            height: 1.8rem;
            line-height: 1.8rem;
            white-space: nowrap;
            box-sizing: content-box;
            -webkit-animation-iteration-count: infinite;
            animation-iteration-count: infinite;
            -webkit-animation-timing-function: linear;
            animation-timing-function: linear;
            -webkit-animation-name: ticker;
            animation-name: ticker;
            -webkit-animation-duration: 40s;
            animation-duration: 40s;
        }}
        .ticker__item {{
            display: inline-block;
            padding: 0 2rem; /* Increased padding for more space between entries */
            font-size: 0.8rem;
            font-family: Helvetica;
            font-weight: bold;
            color: #FFFFFF;
        }}
        .ticker__item:first-child {{
            margin-top: 0;
        }}
        body {{ margin: 0; padding-bottom: 0; }}
        h1, h2, p {{ padding: 0 5%; }}
        </style>

        <div class="ticker-wrap" style="bottom: 2.5rem;">  <!-- Bottom ticker -->
            <div class="ticker ticker1">
        """

        for entry in scrollable_content:
            html_code += f'<div class="ticker__item">{entry}</div>'

        html_code += """
            </div>
        </div>

        <div class="ticker-wrap" style="bottom: 0;">  <!-- Top ticker -->
            <div class="ticker ticker2">
        """

        for entry in scrollable_content_2:
            html_code += f'<div class="ticker__item">{entry}</div>'

        html_code += """
            </div>
        </div>
        """

        # -- START OF CURRENT INTERVAL SUMMARY --

        # Fetch Data
        result = view_all_data()
        result2 = view_all_data2()
        result3 = view_all_data3()
        result4 = view_all_data4()
        result5 = view_all_data5()
        result6 = view_all_data6()
        result7 = view_all_data7()
        result8 = view_all_data8()
        result9 = view_all_data9()
        result10 = view_all_data10()

        df = pd.DataFrame(result,columns=["id", "time_interval", "timestamp"])
        df2 = pd.DataFrame(result2, columns=["id", "net", "insert_time"])
        df3 = pd.DataFrame(result3, columns=["id", "total_bcq", "insert_time"])
        df4 = pd.DataFrame(result4, columns=["id", "substation_load", "insert_time"])
        df5 = pd.DataFrame(result5, columns=["id", "cc", "insert_time"])
        df6 = pd.DataFrame(result6, columns=["id", "actual_energy", "insert_time"])
        df7 = pd.DataFrame(result7, columns=["id", "wesm_exposure", "insert_time"])
        df8 = pd.DataFrame(result8, columns=["id", "final_total", "insert_time"])
        df9 = pd.DataFrame(result9, columns=["id", "rate", "insert_time"])
        df10 = pd.DataFrame(result10, columns=["id", "temp", "weather", "insert_time"])

        last_interval= df['time_interval'].iloc[-1]
        forecasted_energy = df2['net'].iloc[-1]
        total_bcq = df3['total_bcq'].iloc[-1]
        total_substation_load = df4['substation_load'].iloc[-1]
        contestable_energy = df5['cc'].iloc[-1]
        actual_energy= df6['actual_energy'].iloc[-1]
        wesm_exposure = df7['wesm_exposure'].iloc[-1]
        final_total = df8['final_total'].iloc[-1]
        current_rate = df9['rate'].iloc[-1]
        temperature = df10['temp'].iloc[-1]
        weather_condition = df10['weather'].iloc[-1]

        # Define the current date
        c = datetime.now()
        formatted_date = c.strftime("%Y-%m-%d")

        # Current Interval Aesthetics
        st.markdown("""
        <style>
        /* Styling for equal-sized boxes with minimal padding */
        .custom-box {
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 0;  
            margin: 0;
            background-color: #000000;
            height: 13vh;  
            max-width: 100%;  
            box-sizing: border-box;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center; 
        }
                    
        .custom-box-2 {
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 0;  
            margin: 0;
            background-color: #000000;
            height: 13vh;  
            max-width: 100%;  
            box-sizing: border-box;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center; 
        }
        
        @media (max-width: 768px) {
            .custom-box {
                justify-content: center;
                align-items: center;
                text-align: center; 
                height: 17.33vh;  
            }
        }

        @media (max-width: 480px) {
            .custom-box {
                justify-content: center;
                align-items: center;
                text-align: center; 
                height: 21.67vh;  
            }
        }

        /* Smaller font sizes with minimal padding and overflow control */
        .custom-box h4, .custom-box p {
            margin: 0;
            padding: 0;
            font-size: 0.7rem;
            font-family: Helvetica;
            font-weight: normal;
            color: #FFFFFF;
            max-height: 100%;  
            overflow: hidden;  
            text-overflow: ellipsis;  
            justify-content: center;
            align-items: center;
            text-align: center; 
        }
                    
        .custom-box-2 h4, .custom-box-2 p {
            margin: 0;
            padding: 0;
            font-size: 0.55rem;
            font-family: Helvetica;
            font-weight: normal;
            color: #FFFFFF;
            max-height: 100%;  
            overflow: hidden;  
            text-overflow: ellipsis;  
            justify-content: center;
            align-items: center;
            text-align: center; 
        }
                
        .custom-box p {
            font-size: 1.1rem;
            font-family: Helvetica;
            font-weight: bold;
            color: #FFFFFF;
            max-height: 100%;  
            max-width: 100%;
            overflow: hidden;  
            text-overflow: ellipsis;  
            justify-content: center;
            align-items: center;
            text-align: center; 
        }
        
        .custom-box-2 p {
            font-size: 0.8rem;
            font-family: Helvetica;
            font-weight: bold;
            color: #FFFFFF;
            max-height: 100%;  
            max-width: 100%;
            overflow: hidden;  
            text-overflow: ellipsis;  
            justify-content: center;
            align-items: center;
            text-align: center; 
        }
        </style>
        """, unsafe_allow_html=True)


        # Define your custom CSS
        custom_css = """
        <style>
        .container-class {
            min-height: 100%;
        }
        </style>
        """

        # Apply the custom CSS
        st.markdown(custom_css, unsafe_allow_html=True)

        with st.container():

            # Set the image's width to the column width
            st.image(image_path, use_column_width=True)

            # Display the HTML and CSS in Streamlit
            st.components.v1.html(html_code, height=80)

            # Create columns with small gap
            card1, card2, card3, card4, card5, card6, card7, card8, card9, card10 = st.columns(10, gap='small')

            with card1:
                st.markdown('<div class="custom-box"><h4>Current Date</h4><p>{}</p></div>'.format(formatted_date), unsafe_allow_html=True)

            with card2:
                #st.markdown('<div class="custom-box"><h4>Current Interval</h4><p>{}</p></div>'.format(last_interval), unsafe_allow_html=True)
                st.markdown('<div class="custom-box-2"><h4>Current Interval</h4><p>{}</p></div>'.format(last_interval), unsafe_allow_html=True)
            with card3:
                st.markdown('<div class="custom-box"><h4>Temperature</h4><p>{}°C</p><h4>Weather Condition</h4><p>{}</p></div>'.format(temperature, weather_condition), unsafe_allow_html=True)
        
            with card4:
                st.markdown('<div class="custom-box"><h4>Total Substation Load (kW)</h4><p>{}</p></div>'.format(total_substation_load), unsafe_allow_html=True)

            with card5:
                #st.markdown('<div class="custom-box"><h4>Actual Energy (kWh)</h4><p>{}</p><h4>Forecasted Energy (kWh)</h4><p>{}</p></div>'.format(actual_energy, forecasted_energy), unsafe_allow_html=True)
                st.markdown('<div class="custom-box-2"><h4>Actual Energy (kWh)</h4><p>{}</p><h4>Forecasted Energy (kWh)</h4><p>{}</p></div>'.format(actual_energy, forecasted_energy), unsafe_allow_html=True)
            with card6:
                #st.markdown('<div class="custom-box"><h4>WESM Exposure (kWh)</h4><p>{}</p><h4>Contestable Energy (kWh)</h4><p>{}</p></div>'.format(wesm_exposure, contestable_energy), unsafe_allow_html=True)
                st.markdown('<div class="custom-box-2"><h4>WESM Exposure (kWh)</h4><p>{}</p><h4>Contestable Energy (kWh)</h4><p>{}</p></div>'.format(wesm_exposure, contestable_energy), unsafe_allow_html=True)
            with card7:
                st.markdown('<div class="custom-box"><h4>Total BCQ Nomination (kW)</h4><p>{}</p></div>'.format(total_bcq), unsafe_allow_html=True)

            with card8:
                st.markdown('<div class="custom-box"><h4>MORE Trading Node (PhP/kWh)</h4><p>{}</p></div>'.format(final_total), unsafe_allow_html=True)

            with card9:
                st.markdown('<div class="custom-box"><h4>PEDC Trading Node (PhP/kWh)</h4><p>Pending</p></div>', unsafe_allow_html=True)

            with card10:
                st.markdown('<div class="custom-box"><h4>Current Rate (PhP/kWh)</h4><p>{}</p></div>'.format(current_rate), unsafe_allow_html=True)

        st.markdown("""
        <style>
        .small-space {
            margin-bottom: 0.1px;
        }
        </style>
        """, unsafe_allow_html=True)

        # Layout with smaller padding
        col1, col2 = st.columns([1, 1.75])  
        latest_chart_total_sub_load = None # Initialize latest chart
        with col1:
            try:
                latest_chart_total_sub_load = sub_load_func()
                st.plotly_chart(latest_chart_total_sub_load)
            except:
                if latest_chart_total_sub_load:
                    st.plotly_chart(latest_chart_total_sub_load)
        latest_chart_actual_vs_forecasted = None
        with col2:
            try:
                latest_chart_actual_vs_forecasted = actual_vs_forecasted()
                st.plotly_chart(latest_chart_actual_vs_forecasted)
            except:
                if latest_chart_actual_vs_forecasted:
                    st.plotly_chart(latest_chart_actual_vs_forecasted)
                # time.sleep(1)
                # st.rerun()

        col4, col5, col6 = st.columns([1, 1, 0.65])
        latest_chart_bcq = None
        with col4:
            try:
                latest_chart_bcq = bcq_func()
                st.plotly_chart(latest_chart_bcq)
            except:
                if latest_chart_bcq:
                    st.plotly_chart(latest_chart_bcq)
        latest_chart_tipc = None
        with col5:
            try:
                latest_chart_tipc = tipc_func()
                st.plotly_chart(latest_chart_tipc)
            except:
                if latest_chart_tipc:
                    st.plotly_chart(latest_chart_tipc)

        latest_chart_genmix = None
        with col6:
            try:
                latest_chart_genmix = genmix_func()
                st.plotly_chart(latest_chart_genmix)
            except:
                if latest_chart_genmix:
                    st.plotly_chart(latest_chart_genmix)

        # For rerunning the script every 900 seconds (15 minutes)
        time.sleep(REFRESH_INTERVAL)
        st.rerun()