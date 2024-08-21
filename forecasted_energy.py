# Necessary imports
from datetime import datetime
from openpyxl import load_workbook
import mysql.connector
import os
import openpyxl
import pandas as pd
import time

# Function to copy values from source range to destination range
def copy_values(source_sheet, dest_sheet, source_range, dest_range):
    source_start, source_end = source_range
    dest_start, dest_end = dest_range
    
    # Copy values from source to destination
    source_values = []
    for row in source_sheet[source_start:source_end]:
        for cell in row:
            source_values.append(cell.value)
    
    dest_index = 0
    for row in dest_sheet[dest_start:dest_end]:
        for cell in row:
            if dest_index < len(source_values):
                cell.value = source_values[dest_index]
                dest_index += 1
            else:
                break

def forecasted_energy():
    # Defining the base directory
    # base_directory = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\MORE Energy Sourcing\013. Rate Analysis\005. 2024 Rate Analysis"

    # Getting the current month and year
    # current_date = datetime.now()
    # current_month_numeric = current_date.strftime("%m")  # Month in numeric format (e.g., '01' for January)
    # current_month_name = current_date.strftime("%B")   # Full month name (e.g., 'January')
    # current_day = current_date.strftime("%d") # Day in numeric format (e.g., '26')
    # current_year = current_date.strftime("%Y")   # Year in 4-digit format (e.g., '2024')

    # #List of all month names
    # months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    # #Find index of current month
    # current_month_index = months.index(current_month_name)

    # if (current_day>'25') and (current_date.strftime("%m")==current_month_numeric):
    #     #Increment current month by converting to integer, adding 1, and formatting back to zero-padded string
    #     current_month_numeric_increment = '{:02d}'.format((int(current_month_numeric) % 12) + 1)
    #     #Calculate index of next month
    #     next_month_index = (current_month_index + 1) % 12
    #     current_month_name = months[next_month_index]
    # elif(current_day<='25') and (current_date.strftime("%m")==current_month_numeric):
    #     current_month_numeric_increment = current_month_numeric

    # # Constructing the folder name
    # folder_prefix = current_month_numeric_increment.zfill(3)  # Zero-padded three-digit month number (e.g., '006' for June)
    # folder_name = f"{folder_prefix}. {current_month_name} {current_year}"

    # # Constructing the directory path
    # folder_path = os.path.join(base_directory, folder_name)

    # # Additional needed variables for the folder name
    # drs = "Daily Rate Simulation_"
    # extension = ".xlsx"

    # # Folder Name (Example: Daily Rate Simulation_06252024) 
    # folder_name_2 = f"{drs}{current_month_numeric}{current_day}{current_year}{extension}"
    # folder_path_2 = os.path.join(folder_path, folder_name_2)

    base_directory = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\01_Energy Sourcing Files\03_Daily Reports\01_Day Ahead Projections"
    current_date = datetime.now()
    current_month_numeric = current_date.strftime("%m")
    current_month_name = current_date.strftime("%B")
    current_day = current_date.strftime("%d")
    current_year = current_date.strftime("%Y")

    folder_path = os.path.join(base_directory, current_year)

    #List of all month names
    months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    #Find index of current month
    current_month_index = months.index(current_month_name)

    if (current_day>'25') and (current_date.strftime("%m")==current_month_numeric):
        #Increment current month by converting to integer, adding 1, and formatting back to zero-padded string
        current_month_numeric_increment = '{:02d}'.format((int(current_month_numeric) % 12) + 1)
        #Calculate index of next month
        next_month_index = (current_month_index + 1) % 12
        current_month_name = months[next_month_index]
    elif(current_day<='25') and (current_date.strftime("%m")==current_month_numeric):
        current_month_numeric_increment = current_month_numeric

    # Constructing the folder name
    folder_prefix = current_month_numeric_increment.zfill(3)  # Zero-padded three-digit month number (e.g., '006' for June)
    folder_name = f"{folder_prefix}. {current_month_name} {current_year}"

    # Constructing the directory path
    folder_path_2 = os.path.join(folder_path, folder_name)

    # Additional needed variables for the folder name
    drs = "Daily Rate Simulation_"
    extension = ".xlsx"

    # Folder Name (Example: Daily Rate Simulation_06252024) 
    folder_name_2 = f"{drs}{current_month_numeric}{current_day}{current_year}{extension}"
    folder_path_3 = os.path.join(folder_path_2, folder_name_2)

    # For creating a new Excel file with the HOUR and NET columns 
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'HOUR'
    sheet['B1'] = 'NET'
    workbook.save('forecasted_energy.xlsx')
    workbook.close()
    print("Excel file 'forecasted_energy.xlsx' created successfully.")

    # File path for destination Excel file
    destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\forecasted_energy.xlsx'

    # Load source workbook
    source_wb = load_workbook(folder_path_3, data_only=True)  # Use data_only=True to get values instead of formulas
    # Sheet name of the source workbook
    source_sheet = source_wb['6. DAP Report']

    # Load destination workbook
    dest_wb = load_workbook(destination_file_path)

    # Sheet name of the destination workbook
    dest_sheet = dest_wb['Sheet']

    # Define source and destination ranges
    source_ranges = [('C11', 'C34'), ('F11', 'F34')]
    dest_ranges = [('A2', 'A25'), ('B2', 'B25')]

    # Copy data from source to destination based on specified ranges
    for i in range(len(source_ranges)):
        copy_values(source_sheet, dest_sheet, source_ranges[i], dest_ranges[i])

    # Save the destination workbook
    dest_wb.save(destination_file_path)

    # Close workbooks
    source_wb.close()
    dest_wb.close()
    print("Data transfer completed.")

# Function to retrieve the current hour in 24-hour format
def get_current_hour():
    now = datetime.now()
    current_hour = now.hour
    if current_hour == 0:
        return 24
    return current_hour

# Function to get net value for the current hour (1 to 24) from Excel
def get_net_for_current_hour(excel_file, current_hour):
    try:
        # Read Excel file into a Pandas DataFrame, assuming 'HOUR' column has numbers 1 to 24
        df = pd.read_excel(excel_file, header=None, names=['HOUR', 'NET'], skiprows=1)
        
        # Filter the row where 'HOUR' matches the current_hour
        net_value = df.loc[df['HOUR'] == current_hour, 'NET'].values[0]
        return float(net_value) if pd.notnull(net_value) else None
    except Exception as e:
        print(f"Error retrieving net value from Excel: {e}")
        return None

# Function to insert value into MySQL database
def insert_into_mysql(conn, net_value):
    try:
        cursor = conn.cursor()
        sql = "INSERT INTO forecasted_energy (net) VALUES (%s)"
        cursor.execute(sql, (net_value,))
        conn.commit()  # Commit the transaction
        cursor.close()
        print(f"Value {net_value} inserted successfully into MySQL.")
    except mysql.connector.Error as e:
        print(f"Error inserting into MySQL: {e}")
    except Exception as e:
        print(f"Error: {e}")

# Main function to run the script
def main(excel_file, conn):
    while True:
        try:
            forecasted_energy()
            current_hour = get_current_hour()
            net_value = get_net_for_current_hour(excel_file, current_hour)

            if net_value is not None:
                insert_into_mysql(conn, net_value)
            else:
                print(f"Invalid net_value: {net_value}")

            # Delay for 800 seconds before checking again
            time.sleep(800)
        
        except KeyboardInterrupt:
            print("\nTerminating the script.")
            break
        except Exception as e:
            print(f"Error in main loop: {e}")
            time.sleep(60)  # Delay before retrying

if __name__ == "__main__":
    while True:
        excel_file = 'forecasted_energy.xlsx'

        conn = mysql.connector.connect(
            host='localhost',
            database='myDb',
            user='root',
            password=''
        )

        # Create MySQL table if it doesn't exist
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS forecasted_energy (
                id INT AUTO_INCREMENT PRIMARY KEY,
                net FLOAT(10, 2) NOT NULL,
                insert_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cursor.close()

        # Run main function to continuously check and update database
        main(excel_file, conn)