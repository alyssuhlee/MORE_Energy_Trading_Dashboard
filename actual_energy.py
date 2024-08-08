# Necessary imports
from datetime import datetime
from openpyxl import load_workbook
import mysql.connector
import pandas as pd
import time

# File path for destination Excel file
destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\actual_energy.xlsx'

# Load source workbooks
def load_source_workbooks():
    source_wb = load_workbook('station_load.xlsx')
    source_sheet = source_wb['Sheet1']
    source_wb2 = load_workbook('contestable_energy.xlsx')
    source_sheet2 = source_wb2['Sheet']
    return source_wb, source_sheet, source_wb2, source_sheet2

# Load destination workbook
def load_destination_workbook():
    dest_wb = load_workbook(destination_file_path)
    dest_sheet = dest_wb['Sheet']
    return dest_wb, dest_sheet

# Function to copy values from source to destination
def copy_values(source_sheet, dest_sheet, source_range, dest_range):
    source_start, source_end = source_range
    dest_start, dest_end = dest_range
    source_values = [cell.value for row in source_sheet[source_start:source_end] for cell in row]
    dest_index = 0
    for row in dest_sheet[dest_start:dest_end]:
        for cell in row:
            if dest_index < len(source_values):
                cell.value = source_values[dest_index]
                dest_index += 1
            else:
                break

# Function to copy values from source to destination (for contestable energy)
def copy_values2(source_sheet2, dest_sheet, source_range2, dest_range2):
    source_start2, source_end2 = source_range2
    dest_start2, dest_end2 = dest_range2
    source_values2 = [cell2.value for row2 in source_sheet2[source_start2:source_end2] for cell2 in row2]
    dest_index2 = 0
    for row2 in dest_sheet[dest_start2:dest_end2]:
        for cell2 in row2:
            if dest_index2 < len(source_values2):
                cell2.value = source_values2[dest_index2]
                dest_index2 += 1
            else:
                break

# Retrieve the previous Actual Energy value
def get_previous_ae_value(dest_sheet, current_hour):
    for row in dest_sheet.iter_rows(min_row=2, max_row=25, min_col=1, max_col=4):
        if row[0].value == current_hour:
            return row[3].value
    return None

# Calculate Actual Energy and update the destination sheet
def calculate_actual_energy(dest_sheet):
    for row in dest_sheet.iter_rows(min_row=2, max_row=25, min_col=2, max_col=4):
        total_substation_load = row[0].value
        contestable_energy = row[1].value
        previous_actual_energy = row[2].value  # This was the previous Actual Energy

        if total_substation_load is not None and contestable_energy is not None:
            actual_energy = total_substation_load - contestable_energy 
            if actual_energy < 0:
                actual_energy = 0
        else:
            actual_energy = previous_actual_energy  # Use the previous Actual Energy if no new data

        dest_sheet.cell(row=row[0].row, column=4).value = actual_energy

# Function to retrieve the current hour in 24-hour format
def get_current_hour():
    now = datetime.now()
    current_hour = now.hour
    if current_hour == 0:
        return 24
    return current_hour

def get_tsl_for_current_hour(excel_file, current_hour):
    try:
        df = pd.read_excel(excel_file, header=None, names=['HOUR', 'TOTAL_SUBSTATION_LOAD', 'CONTESTABLE_ENERGY', 'ACTUAL_ENERGY'], skiprows=1)
        tsl_value = df.loc[df['HOUR'] == current_hour, 'TOTAL_SUBSTATION_LOAD'].values[0]
        return float(tsl_value) if pd.notnull(tsl_value) else None
    except Exception as e:
        print(f"Error retrieving tsl value from Excel: {e}")
        return None

def get_ce_for_current_hour(excel_file, current_hour):
    try:
        df = pd.read_excel(excel_file, header=None, names=['HOUR', 'TOTAL_SUBSTATION_LOAD', 'CONTESTABLE_ENERGY', 'ACTUAL_ENERGY'], skiprows=1)
        ce_value = df.loc[df['HOUR'] == current_hour, 'CONTESTABLE_ENERGY'].values[0]
        return float(ce_value) if pd.notnull(ce_value) else None
    except Exception as e:
        print(f"Error retrieving ce value from Excel: {e}")
        return None

def get_ae_for_current_hour(excel_file, current_hour):
    try:
        tsl_value = get_tsl_for_current_hour(excel_file, current_hour)
        ce_value = get_ce_for_current_hour(excel_file, current_hour)
        if tsl_value is not None and ce_value is not None:
            ae_value = tsl_value - ce_value 
            if ae_value < 0:
                ae_value = 0
            return ae_value
        else:
            return None
    except Exception as e:
        print(f"Error calculating ae value: {e}")
        return None

# Function to insert value into MySQL database
def insert_into_mysql(conn, ae_value):
    try:
        cursor = conn.cursor()
        sql = "INSERT INTO actual_energy (actual_energy) VALUES (%s)"
        cursor.execute(sql, (ae_value,))
        conn.commit()
        cursor.close()
        print(f"Value {ae_value} inserted successfully into MySQL.")
    except mysql.connector.Error as e:
        print(f"Error inserting into MySQL: {e}")
    except Exception as e:
        print(f"Error: {e}")

# Main function to run the script
def main(excel_file, conn):
    previous_ae_value = None
    while True:
        try:
            # Load workbooks and sheets
            source_wb, source_sheet, source_wb2, source_sheet2 = load_source_workbooks()
            dest_wb, dest_sheet = load_destination_workbook()

            # Define ranges and copy data
            source_ranges = [('K2', 'K25')]
            dest_ranges = [('B2', 'B25')]
            source_ranges2 = [('A2', 'A25'), ('B2', 'B25')]
            dest_ranges2 = [('A2', 'A25'), ('C2', 'C25')]

            for i in range(len(source_ranges)):
                copy_values(source_sheet, dest_sheet, source_ranges[i], dest_ranges[i])

            for i in range(len(source_ranges2)):
                copy_values2(source_sheet2, dest_sheet, source_ranges2[i], dest_ranges2[i])

            # Calculate Actual Energy
            calculate_actual_energy(dest_sheet)

            # Save the updated destination workbook
            dest_wb.save(destination_file_path)

            # Retrieve and insert the current hour's data into MySQL
            current_hour = get_current_hour()
            ae_value = get_ae_for_current_hour(destination_file_path, current_hour)
            if ae_value is not None:
                insert_into_mysql(conn, ae_value)
                previous_ae_value = ae_value
            else:
                # Insert previous value if current AE value is None
                if previous_ae_value is not None:
                    insert_into_mysql(conn, previous_ae_value)

            # Close workbooks
            source_wb.close()
            dest_wb.close()
            source_wb2.close()

            # Delay for 1 minute (60 seconds) before checking again
            time.sleep(60)
        except KeyboardInterrupt:
            print("\nTerminating the script.")
            break
        except Exception as e:
            print(f"Error in main loop: {e}")
            time.sleep(60)

if __name__ == "__main__":
    while True:
        excel_file = destination_file_path

        conn = mysql.connector.connect(
            host='localhost',
            database='myDb',
            user='root',
            password=''
        )

        # Create MySQL table if it doesn't exist
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS actual_energy (
                id INT AUTO_INCREMENT PRIMARY KEY,
                actual_energy FLOAT(10, 2) NOT NULL,
                insert_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cursor.close()

        # Run main function to continuously check and update database
        main(excel_file, conn)