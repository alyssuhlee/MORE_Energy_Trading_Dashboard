# Necessary imports
from datetime import datetime
from openpyxl import load_workbook
import mysql.connector
import openpyxl
import pandas as pd
import time

# Function to copy values from source range to destination range
def copy_values(source_sheet, dest_sheet, source_range, dest_range):
    source_start, source_end = source_range
    dest_start, dest_end = dest_range

    # Copy values from source to destination
    source_values = []
    for row in source_sheet[source_start:source_end]:
        for cell in row:
            source_values.append(cell.value)
    dest_index = 0
    for row in dest_sheet[dest_start:dest_end]:
        for cell in row:
            if dest_index < len(source_values):
                cell.value = source_values[dest_index]
                dest_index += 1
            else:
                break

# Function to copy values from source range to destination range (for contestable energy)
def copy_values2(source_sheet2, dest_sheet, source_range2, dest_range2):
    source_start2, source_end2 = source_range2
    dest_start2, dest_end2 = dest_range2

    # Copy values from source to destination
    source_values2 = []
    for row2 in source_sheet2[source_start2:source_end2]:
        for cell2 in row2:
            source_values2.append(cell2.value)
    dest_index2 = 0
    for row2 in dest_sheet[dest_start2:dest_end2]:
        for cell2 in row2:
            if dest_index2 < len(source_values2):
                cell2.value = source_values2[dest_index2]
                dest_index2 += 1
            else:
                break

# Function to copy values from source range to destination range (for contestable energy)
def copy_values3(source_sheet3, dest_sheet, source_range3, dest_range3):
    source_start3, source_end3 = source_range3
    dest_start3, dest_end3 = dest_range3

    # Copy values from source to destination
    source_values3 = []
    for row3 in source_sheet3[source_start3:source_end3]:
        for cell3 in row3:
            source_values3.append(cell3.value)
    dest_index3 = 0
    for row3 in dest_sheet[dest_start3:dest_end3]:
        for cell3 in row3:
            if dest_index3 < len(source_values3):
                cell3.value = source_values3[dest_index3]
                dest_index3 += 1
            else:
                break

def wesm_exposure():
    # For creating an Excel file 
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'HOUR'
    sheet['B1'] = 'ACTUAL_ENERGY'
    sheet['C1'] = 'CONTESTABLE_ENERGY'
    sheet['D1'] = 'TOTAL_BCQ_NOMINATION'
    sheet['E1'] = 'WESM_EXPOSURE'  
    workbook.save('wesm_exposure.xlsx')
    workbook.close()
    print("Excel file 'wesm_exposure.xlsx' created successfully.")

    # File path for destination Excel file
    destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\wesm_exposure.xlsx'

    # Load source workbooks
    source_wb = load_workbook('actual_energy.xlsx')
    source_sheet = source_wb['Sheet']
    source_wb3 = load_workbook('contestable_energy.xlsx')
    source_sheet3 = source_wb3['Sheet']
    source_wb2 = load_workbook('total_bcq_nomination.xlsx')
    source_sheet2 = source_wb2['Sheet']

    # Load destination workbook
    dest_wb = load_workbook(destination_file_path)
    dest_sheet = dest_wb['Sheet']

    # Define source and destination ranges

    # Hour and Actual Energy
    source_ranges = [('A2', 'A25'), ('D2', 'D25')]
    dest_ranges = [('A2', 'A25'), ('B2', 'B25')]
    # Contestable Energy
    source_ranges3 = [('B2', 'B25')]
    dest_ranges3 = [('C2', 'C25')]
    # Total BCQ Nomination
    source_ranges2 = [('E2', 'E25')]
    dest_ranges2 = [('D2', 'D25')]

    # Copy data from source to destination based on specified ranges
    for i in range(len(source_ranges)):
        copy_values(source_sheet, dest_sheet, source_ranges[i], dest_ranges[i])

    for i in range(len(source_ranges2)):
        copy_values2(source_sheet2, dest_sheet, source_ranges2[i], dest_ranges2[i])

    for i in range(len(source_ranges3)):
        copy_values3(source_sheet3, dest_sheet, source_ranges3[i], dest_ranges3[i])

    for row in dest_sheet.iter_rows(min_row=2, max_row=25, min_col=2, max_col=4):
        actual_energy = row[0].value
        contestable_energy = row[1].value
        total_bcq_nomination = row[2].value
        wesm_exposure = actual_energy - contestable_energy - total_bcq_nomination if actual_energy and contestable_energy and total_bcq_nomination else None
        dest_sheet.cell(row=row[0].row, column=5).value = wesm_exposure

    # Save the destination workbook
    dest_wb.save(destination_file_path)
    # Close workbooks
    source_wb.close()
    dest_wb.close()
    source_wb3.close()
    source_wb2.close()
    print("Data transfer completed.")

# Function to retrieve the current hour in 24-hour format
def get_current_hour():
    now = datetime.now()
    current_hour = now.hour
    if current_hour == 0:
        return 24
    return current_hour

def get_actual_energy_for_current_hour(excel_file, current_hour):
    try:
        df = pd.read_excel(excel_file, header=None, names=['HOUR', 'ACTUAL_ENERGY', 'TOTAL_BCQ_NOMINATION', 'CONTESTABLE_ENERGY', 'WESM_EXPOSURE'], skiprows=1)
        actual_energy_value = df.loc[df['HOUR'] == current_hour, 'ACTUAL_ENERGY'].values[0]
        return float(actual_energy_value) if pd.notnull(actual_energy_value) else None
    except Exception as e:
        print(f"Error retrieving actual energy value from Excel: {e}")
        return None

def get_contestable_energy_for_current_hour(excel_file, current_hour):
    try:
        df = pd.read_excel(excel_file, header=None, names=['HOUR', 'ACTUAL_ENERGY', 'TOTAL_BCQ_NOMINATION', 'CONTESTABLE_ENERGY', 'WESM_EXPOSURE'], skiprows=1)
        contestable_energy_value = df.loc[df['HOUR'] == current_hour, 'CONTESTABLE_ENERGY'].values[0]
        return float(contestable_energy_value) if pd.notnull(contestable_energy_value) else None
    except Exception as e:
        print(f"Error retrieving contestable energy value from Excel: {e}")
        return None

def get_total_bcq_nomination_for_current_hour(excel_file, current_hour):
    try:
        df = pd.read_excel(excel_file, header=None, names=['HOUR', 'ACTUAL_ENERGY', 'TOTAL_BCQ_NOMINATION', 'CONTESTABLE_ENERGY', 'WESM_EXPOSURE'], skiprows=1)
        total_bcq_nomination_value = df.loc[df['HOUR'] == current_hour, 'TOTAL_BCQ_NOMINATION'].values[0]
        return float(total_bcq_nomination_value) if pd.notnull(total_bcq_nomination_value) else None
    except Exception as e:
        print(f"Error retrieving total bcq nomination value from Excel: {e}")
        return None

def get_wesm_exposure_for_current_hour(excel_file, current_hour):
    try:
        actual_energy_value = get_actual_energy_for_current_hour(excel_file, current_hour)
        contestable_energy_value = get_contestable_energy_for_current_hour(excel_file, current_hour)
        total_bcq_nomination_value = get_total_bcq_nomination_for_current_hour(excel_file, current_hour)

        wesm_exposure_value = actual_energy_value - contestable_energy_value - total_bcq_nomination_value if actual_energy_value is not None and contestable_energy_value is not None and total_bcq_nomination_value is not None else None
        return wesm_exposure_value
    except Exception as e:
        print(f"Error calculating wesm exposure value: {e}")
        return None

# Function to insert value into MySQL database
def insert_into_mysql(conn, wesm_exposure_value):
    try:
        cursor = conn.cursor()
        sql = "INSERT INTO wesm_exposure (wesm_exposure) VALUES (%s)"
        cursor.execute(sql, (wesm_exposure_value,))
        conn.commit()  # Commit the transaction
        cursor.close()
        print(f"Value {wesm_exposure_value} inserted successfully into MySQL.")
    except mysql.connector.Error as e:
        print(f"Error inserting into MySQL: {e}")
    except Exception as e:
        print(f"Error: {e}")

# Main function to run the script
def main(excel_file, conn):
    while True:
        try:
            wesm_exposure()
            current_hour = get_current_hour()
            wesm_exposure_value = get_wesm_exposure_for_current_hour(excel_file, current_hour)
            if wesm_exposure_value is not None:
                insert_into_mysql(conn, wesm_exposure_value)
            else:
                insert_into_mysql(conn, wesm_exposure_value)
            # Delay for 1 minute (60 seconds) before checking again
            time.sleep(60)
        except KeyboardInterrupt:
            print("\nTerminating the script.")
            break
        except Exception as e:
            print(f"Error in main loop: {e}")
            time.sleep(60)  # Delay before retrying

if __name__ == "__main__":
    while True:
        excel_file = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\wesm_exposure.xlsx'

        conn = mysql.connector.connect(
            host='localhost',
            database='myDb',
            user='root',
            password=''
        )

        # Create MySQL table if it doesn't exist
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS wesm_exposure (
                id INT AUTO_INCREMENT PRIMARY KEY,
                wesm_exposure FLOAT(10, 2) NOT NULL,
                insert_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cursor.close()

        # Run main function to continuously check and update database
        main(excel_file, conn)