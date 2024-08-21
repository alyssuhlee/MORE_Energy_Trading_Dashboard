# Necessary imports
from datetime import datetime
from openpyxl import load_workbook
import mysql.connector
import os
import openpyxl
import pandas as pd
import time

# Function to copy values from source range to destination range
def copy_values(source_sheet, dest_sheet, source_range, dest_range):
    source_start, source_end = source_range
    dest_start, dest_end = dest_range
    
    # Copy values from source to destination
    source_values = []
    for row in source_sheet[source_start:source_end]:
        for cell in row:
            source_values.append(cell.value)
    
    dest_index = 0
    for row in dest_sheet[dest_start:dest_end]:
        for cell in row:
            if dest_index < len(source_values):
                cell.value = source_values[dest_index]
                dest_index += 1
            else:
                break

def total_bcq_nomination():
    # # Defining the base directory
    # base_directory = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\MORE Energy Sourcing\013. Rate Analysis\005. 2024 Rate Analysis"

    # # Getting the current month and year
    # current_date = datetime.now()
    # current_month_numeric = current_date.strftime("%m")  # Month in numeric format (e.g., '01' for January)
    # current_month_name = current_date.strftime("%B")   # Full month name (e.g., 'January')
    # current_day = current_date.strftime("%d") # Day in numeric format (e.g., '26')
    # current_year = current_date.strftime("%Y")   # Year in 4-digit format (e.g., '2024')

    # #List of all month names
    # months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    # #Find index of current month
    # current_month_index = months.index(current_month_name)

    # # Initialize current_month_numeric_increment
    # current_month_numeric_increment = current_month_numeric # Default to current month if condition not met

    # if (current_day>'25') and (current_date.strftime("%m")==current_month_numeric):
    #     #Increment current month by converting to integer, adding 1, and formatting back to zero-padded string
    #     current_month_numeric_increment = '{:02d}'.format((int(current_month_numeric) % 12) + 1)
    #     #Calculate index of next month
    #     next_month_index = (current_month_index + 1) % 12
    #     current_month_name = months[next_month_index]

    # # Constructing the folder name
    # folder_prefix = current_month_numeric_increment.zfill(3)  # Zero-padded three-digit month number (e.g., '006' for June)
    # folder_name = f"{folder_prefix}. {current_month_name} {current_year}"

    # # Constructing the directory path
    # folder_path = os.path.join(base_directory, folder_name)

    # # Additional needed variables for the folder name
    # drs = "Daily Rate Simulation_"
    # extension = ".xlsx"

    # # Folder name (Example: Daily Rate Simulation_06252024) 
    # folder_name_2 = f"{drs}{current_month_numeric}{current_day}{current_year}{extension}"
    # folder_path_2 = os.path.join(folder_path, folder_name_2)

    base_directory = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\01_Energy Sourcing Files\03_Daily Reports\01_Day Ahead Projections"
    current_date = datetime.now()
    current_month_numeric = current_date.strftime("%m")
    current_month_name = current_date.strftime("%B")
    current_day = current_date.strftime("%d")
    current_year = current_date.strftime("%Y")

    folder_path = os.path.join(base_directory, current_year)

    #List of all month names
    months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    #Find index of current month
    current_month_index = months.index(current_month_name)

    if (current_day>'25') and (current_date.strftime("%m")==current_month_numeric):
        #Increment current month by converting to integer, adding 1, and formatting back to zero-padded string
        current_month_numeric_increment = '{:02d}'.format((int(current_month_numeric) % 12) + 1)
        #Calculate index of next month
        next_month_index = (current_month_index + 1) % 12
        current_month_name = months[next_month_index]
    elif(current_day<='25') and (current_date.strftime("%m")==current_month_numeric):
        current_month_numeric_increment = current_month_numeric

    # Constructing the folder name
    folder_prefix = current_month_numeric_increment.zfill(3)  # Zero-padded three-digit month number (e.g., '006' for June)
    folder_name = f"{folder_prefix}. {current_month_name} {current_year}"

    # Constructing the directory path
    folder_path_2 = os.path.join(folder_path, folder_name)

    # Additional needed variables for the folder name
    drs = "Daily Rate Simulation_"
    extension = ".xlsx"

    # Folder Name (Example: Daily Rate Simulation_06252024) 
    folder_name_2 = f"{drs}{current_month_numeric}{current_day}{current_year}{extension}"
    folder_path_3 = os.path.join(folder_path_2, folder_name_2)

    # For creating a new Excel file with the HOUR and TOTAL BCQ NOMINATION columns 
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'HOUR'
    sheet['B1'] = 'SCPC'
    sheet['C1'] = 'KSPC'
    sheet['D1'] = 'EDC'
    sheet['E1'] = 'TOTAL_BCQ'

    workbook.save('total_bcq_nomination.xlsx')
    workbook.close()
    print("Excel file 'total_bcq_nomination.xlsx' created successfully.")

    # File path for destination Excel file
    destination_file_path = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\total_bcq_nomination.xlsx'

    # Load source workbook
    source_wb = load_workbook(folder_path_3, data_only=True)  # Use data_only=True to get values instead of formulas
    # Sheet name of the source workbook
    source_sheet = source_wb['6. DAP Report']

    # Load destination workbook
    dest_wb = load_workbook(destination_file_path)

    # Sheet name of the destination workbook
    dest_sheet = dest_wb['Sheet']

    # Define source and destination ranges
    source_ranges = [('C11', 'C34'), ('G11', 'G34'), ('H11', 'H34'), ('I11', 'I34')]
    dest_ranges = [('A2', 'A25'), ('B2', 'B25'), ('C2', 'C25'), ('D2', 'D25')]

    # Copy data from source to destination based on specified ranges
    for i in range(len(source_ranges)):
        copy_values(source_sheet, dest_sheet, source_ranges[i], dest_ranges[i])

    # Save the destination workbook
    dest_wb.save(destination_file_path)

    # Close workbooks
    source_wb.close()
    dest_wb.close()
    print("Data transfer completed.")

    # Total BCQ of 1st Hour
    value_B2 = dest_sheet['B2'].value
    value_C2 = dest_sheet['C2'].value
    value_D2 = dest_sheet['D2'].value

    result2 = value_B2 + value_C2 + value_D2
    dest_sheet['E2'] = result2

    # Total BCQ of 2nd Hour
    value_B3 = dest_sheet['B3'].value
    value_C3 = dest_sheet['C3'].value
    value_D3 = dest_sheet['D3'].value

    result3 = value_B3 + value_C3 + value_D3
    dest_sheet['E3'] = result3

    # Total BCQ of 3rd Hour
    value_B4 = dest_sheet['B4'].value
    value_C4 = dest_sheet['C4'].value
    value_D4 = dest_sheet['D4'].value

    result4 = value_B4 + value_C4 + value_D4
    dest_sheet['E4'] = result4

    # Total BCQ of 4th Hour
    value_B5 = dest_sheet['B5'].value
    value_C5 = dest_sheet['C5'].value
    value_D5 = dest_sheet['D5'].value

    result5 = value_B5 + value_C5 + value_D5
    dest_sheet['E5'] = result5

    # Total BCQ of 5th Hour
    value_B6 = dest_sheet['B6'].value
    value_C6 = dest_sheet['C6'].value
    value_D6 = dest_sheet['D6'].value

    result6 = value_B6 + value_C6 + value_D6
    dest_sheet['E6'] = result6

    # Total BCQ of 6th Hour
    value_B7 = dest_sheet['B7'].value
    value_C7 = dest_sheet['C7'].value
    value_D7 = dest_sheet['D7'].value

    result7 = value_B7 + value_C7 + value_D7
    dest_sheet['E7'] = result7

    # Total BCQ of 7th Hour
    value_B8 = dest_sheet['B8'].value
    value_C8 = dest_sheet['C8'].value
    value_D8 = dest_sheet['D8'].value

    result8 = value_B8 + value_C8 + value_D8
    dest_sheet['E8'] = result8

    # Total BCQ of 8th Hour
    value_B9 = dest_sheet['B9'].value
    value_C9 = dest_sheet['C9'].value
    value_D9 = dest_sheet['D9'].value

    result9 = value_B9 + value_C9 + value_D9
    dest_sheet['E9'] = result9

    # Total BCQ of 9th Hour
    value_B10 = dest_sheet['B10'].value
    value_C10 = dest_sheet['C10'].value
    value_D10 = dest_sheet['D10'].value

    result10 = value_B10 + value_C10 + value_D10
    dest_sheet['E10'] = result10

    # Total BCQ of 10th Hour
    value_B11 = dest_sheet['B11'].value
    value_C11 = dest_sheet['C11'].value
    value_D11 = dest_sheet['D11'].value

    result11 = value_B11 + value_C11 + value_D11
    dest_sheet['E11'] = result11

    # Total BCQ of 11th Hour
    value_B12 = dest_sheet['B12'].value
    value_C12 = dest_sheet['C12'].value
    value_D12 = dest_sheet['D12'].value

    result12 = value_B12 + value_C12 + value_D12
    dest_sheet['E12'] = result12

    # Total BCQ of 12th Hour
    value_B13 = dest_sheet['B13'].value
    value_C13 = dest_sheet['C13'].value
    value_D13 = dest_sheet['D13'].value

    result13 = value_B13 + value_C13 + value_D13
    dest_sheet['E13'] = result13

    # Total BCQ of 13th Hour
    value_B14 = dest_sheet['B14'].value
    value_C14 = dest_sheet['C14'].value
    value_D14 = dest_sheet['D14'].value

    result14 = value_B14 + value_C14 + value_D14
    dest_sheet['E14'] = result14

    # Total BCQ of 14th Hour
    value_B15 = dest_sheet['B15'].value
    value_C15 = dest_sheet['C15'].value
    value_D15 = dest_sheet['D15'].value

    result15 = value_B15 + value_C15 + value_D15
    dest_sheet['E15'] = result15

    # Total BCQ of 15th Hour
    value_B16 = dest_sheet['B16'].value
    value_C16 = dest_sheet['C16'].value
    value_D16 = dest_sheet['D16'].value

    result16 = value_B16 + value_C16 + value_D16
    dest_sheet['E16'] = result16

    # Total BCQ of 16th Hour
    value_B17 = dest_sheet['B17'].value
    value_C17 = dest_sheet['C17'].value
    value_D17 = dest_sheet['D17'].value

    result17 = value_B17 + value_C17 + value_D17 
    dest_sheet['E17'] = result17

    # Total BCQ of 17th Hour
    value_B18 = dest_sheet['B18'].value
    value_C18 = dest_sheet['C18'].value
    value_D18 = dest_sheet['D18'].value

    result18 = value_B18 + value_C18 + value_D18
    dest_sheet['E18'] = result18

    # Total BCQ of 18th Hour
    value_B19 = dest_sheet['B19'].value
    value_C19 = dest_sheet['C19'].value
    value_D19 = dest_sheet['D19'].value

    result19 = value_B19 + value_C19 + value_D19
    dest_sheet['E19'] = result19

    # Total BCQ of 19th Hour
    value_B20 = dest_sheet['B20'].value
    value_C20 = dest_sheet['C20'].value
    value_D20 = dest_sheet['D20'].value

    result20 = value_B20 + value_C20 + value_D20
    dest_sheet['E20'] = result20

    # Total BCQ of 20th Hour
    value_B21 = dest_sheet['B21'].value
    value_C21 = dest_sheet['C21'].value
    value_D21 = dest_sheet['D21'].value

    result21 = value_B21 + value_C21 + value_D21
    dest_sheet['E21'] = result21

    # Total BCQ of 21st Hour
    value_B22 = dest_sheet['B22'].value
    value_C22 = dest_sheet['C22'].value
    value_D22 = dest_sheet['D22'].value

    result22 = value_B22 + value_C22 + value_D22
    dest_sheet['E22'] = result22

    # Total BCQ of 22nd Hour
    value_B23 = dest_sheet['B23'].value
    value_C23 = dest_sheet['C23'].value
    value_D23 = dest_sheet['D23'].value

    result23 = value_B23 + value_C23 + value_D23
    dest_sheet['E23'] = result23

    # Total BCQ of 23rd Hour
    value_B24 = dest_sheet['B24'].value
    value_C24 = dest_sheet['C24'].value
    value_D24 = dest_sheet['D24'].value

    result24 = value_B24 + value_C24 + value_D24
    dest_sheet['E24'] = result24

    # Total BCQ of 24th Hour
    value_B25 = dest_sheet['B25'].value
    value_C25 = dest_sheet['C25'].value
    value_D25 = dest_sheet['D25'].value

    result25 = value_B25 + value_C25 + value_D25
    dest_sheet['E25'] = result25

    # Save the workbook back to file
    dest_wb.save(destination_file_path)

# Function to retrieve the current hour in 24-hour format
def get_current_hour():
    now = datetime.now()
    current_hour = now.hour
    if current_hour == 0:
        return 24
    return current_hour

def get_total_bcq_for_current_hour(excel_file, current_hour):
    try:
        df = pd.read_excel(excel_file, header=None, names=['HOUR', 'SCPC', 'KSPC', 'EDC', 'TOTAL_BCQ'], skiprows=1)
        total_bcq_value = df.loc[df['HOUR'] == current_hour, 'TOTAL_BCQ'].values[0]
        return float(total_bcq_value) if pd.notnull(total_bcq_value) else None
    except Exception as e:
        print(f"Error retrieving total bcq value from Excel: {e}")
        return None

def insert_into_mysql(conn, total_bcq_value):
    try:
        cursor = conn.cursor()
        sql = "INSERT INTO total_bcq_nomination (total_bcq) VALUES (%s)"
        cursor.execute(sql, (total_bcq_value,))
        conn.commit()
        cursor.close()
        print(f"Value {total_bcq_value} inserted successfully into MySQL.")
    except mysql.connector.Error as e:
        print(f"Error inserting into MySQL: {e}")
    except Exception as e:
        print(f"Error: {e}")

def main(excel_file, conn):
    while True:
        try:
            total_bcq_nomination()
            current_hour = get_current_hour()
            total_bcq_value = get_total_bcq_for_current_hour(excel_file, current_hour)

            if total_bcq_value is not None:
                insert_into_mysql(conn, total_bcq_value)
            else:
                print(f"Invalid total_bcq_value: {total_bcq_value}")

            time.sleep(800)

        except KeyboardInterrupt:
            print("\nTerminating the script.")
            break
        except Exception as e:
            print(f"Error in main loop: {e}")
            time.sleep(60)  # Delay before retrying

if __name__ == "__main__":
    while True:
        excel_file = 'total_bcq_nomination.xlsx'

        conn = mysql.connector.connect(
            host='localhost',
            database='myDb',
            user='root',
            password=''
        )

        # Create MySQL table if it doesn't exist
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS total_bcq_nomination (
                id INT AUTO_INCREMENT PRIMARY KEY,
                total_bcq FLOAT(10, 2) NOT NULL,
                insert_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cursor.close()

        main(excel_file, conn)