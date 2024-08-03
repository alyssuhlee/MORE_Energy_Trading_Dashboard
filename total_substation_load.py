from mysql.connector import Error
from sqlalchemy import create_engine
from openpyxl import load_workbook
import pandas as pd
import mysql.connector
import time
import openpyxl
import datetime
import threading

file_loc = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\\'
current_day = datetime.datetime.today()

def query_substation_data():
    try:
        # Create SQLAlchemy engine
        engine = create_engine('mysql+mysqlconnector://mepcadmi_guest:hJER8pBv8WyS@192.185.52.175/mepcadmi_empower')

        query = (
            "SELECT substation, feeder_no, date_entered, time_entered, load_kw, meter_reading, main_meter "
            "FROM mepcadmi_empower.substation_load WHERE date_entered = '{}' "
        ).format(current_day.strftime('%Y-%m-%d'))

        substation_load_df = pd.read_sql(query, engine, parse_dates={"date_entered": {"format": "%d/%m/%y"}})
        substation_load_df = substation_load_df.sort_values(by=['date_entered', 'time_entered'])

        substation_load_df = pd.pivot_table(substation_load_df, values='load_kw',
                                            index=['date_entered', 'time_entered'], columns='substation',
                                            aggfunc='sum')
        substation_load_df['10 MVA Mobile SS - MS1'] = 0
        substation_load_df['MOLO SUBSTATION'] = 0
        substation_load_df = substation_load_df[
            ['LAPAZ SUBSTATION', 'JARO SUBSTATION', 'MANDURRIAO SUBSTATION', 'MOLO SUBSTATION',
             'DIVERSION SUBSTATION', '10 MVA Mobile SS - MS1', '36 MVA Mobile SS - MS2', '36 MVA Megaworld SS - MGW']]
        substation_load_df = substation_load_df.reset_index()
        substation_load_df.columns = ['Date', 'Hour', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8']
        substation_load_df.fillna(0, inplace=True)
        substation_load_df['Total'] = substation_load_df['S1'] + substation_load_df['S2'] + substation_load_df['S3'] + \
                                      substation_load_df['S4'] + substation_load_df['S5'] + substation_load_df['S6'] + \
                                      substation_load_df['S7'] + substation_load_df['S8']
        substation_load_df['eS1'] = substation_load_df['S1'] * 1.0001
        substation_load_df['eS2'] = substation_load_df['S2'] * 1.0001
        substation_load_df['eS3'] = substation_load_df['S3'] * 1.0001
        substation_load_df['eS4'] = substation_load_df['S4'] * 1.0001
        substation_load_df['eS5'] = substation_load_df['S5'] * 1.0001
        substation_load_df['eS6'] = substation_load_df['S6'] * 1.0001
        substation_load_df['eS7'] = substation_load_df['S7'] * 1.0001
        substation_load_df['eS8'] = substation_load_df['S8'] * 1.0001
        substation_load_df = substation_load_df[substation_load_df['Hour'] != 0]
        substation_load_df = substation_load_df.fillna(0)
        substation_load_df['eTotal'] = substation_load_df['eS1'] + substation_load_df['eS2'] + substation_load_df[
            'eS3'] + substation_load_df['eS4'] + substation_load_df['eS5'] + substation_load_df['eS6'] + \
                                       substation_load_df['eS7'] + substation_load_df['eS8']

        substation_load_df.to_csv(file_loc + 'station_load.csv', index=False)

    except Exception as e:
        print(e)

def get_current_hour():
    now = datetime.datetime.now()
    return now.hour

def get_total_for_current_hour(excel_file, current_hour):
    try:
        df = pd.read_excel(excel_file, header=None, names=['Date', 'Hour', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'Total', 'eS1', 'eS2', 'eS3', 'eS4', 'eS5', 'eS6', 'eS7', 'eS8', 'eTotal'], skiprows=1)
        
        # Check if DataFrame is empty
        if df.empty:
            print("Excel file is empty.")
            return None

        # Check if the DataFrame contains the current hour
        if current_hour not in df['Hour'].values:
            print(f"No data for the current hour: {current_hour}")
            return None

        total_substation_load_value = df.loc[df['Hour'] == current_hour, 'Total'].values[0]
        return float(total_substation_load_value) if pd.notnull(total_substation_load_value) else None
    except Exception as e:
        print(f"Error retrieving total substation load value from Excel: {e}")
        return None

def insert_into_mysql(conn, total_substation_load_value):
    try:
        cursor = conn.cursor()
        sql = "INSERT INTO total_substation_load (substation_load) VALUES (%s)"
        cursor.execute(sql, (total_substation_load_value,))
        conn.commit()  # Commit the transaction
        cursor.close()
        print(f"Value {total_substation_load_value} inserted successfully into MySQL.")
    except mysql.connector.Error as e:
        print(f"Error inserting into MySQL: {e}")
    except Exception as e:
        print(f"Error: {e}")

# Main function to run the script
def main(excel_file, conn):
    while True:
        try:
            current_hour = get_current_hour()
            total_substation_load_value = get_total_for_current_hour(excel_file, current_hour)
            if total_substation_load_value is not None:
                insert_into_mysql(conn, total_substation_load_value)
            else:
                print(f"Invalid total_substation_load_value: {total_substation_load_value}")
            time.sleep(300)
        except KeyboardInterrupt:
            print("\nTerminating the script.")
            break
        except Exception as e:
            print(f"Error in main loop: {e}")
            time.sleep(300)  

def query_run():
    while True:
        query_substation_data()
        print("Query executed")
        time.sleep(300)

if __name__ == "__main__":

    # Run query run in a separate thread
    query_thread = threading.Thread(target=query_run, daemon=True)
    query_thread.start()

    read_file = pd.read_csv(file_loc + 'station_load.csv')
    read_file.to_excel(file_loc + 'station_load.xlsx', index=False, header=True)
    print("Successfully created station_load.xlsx.")

    # station_load_graph.xlsx
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Hour'
    sheet['B1'] = 'Lapaz'
    sheet['C1'] = 'Jaro'
    sheet['D1'] = 'Mandurriao'
    sheet['E1'] = 'Molo'
    sheet['F1'] = 'Diversion'
    sheet['G1'] = 'Mobile SS 1'
    sheet['H1'] = 'Mobile SS 2'
    sheet['I1'] = 'Megaworld'
    workbook.save('station_load_graph.xlsx')
    workbook.close()
    print("Excel file 'station_load_graph.xlsx' created successfully.")

    # File path for destination Excel file
    destination_file_path_slg = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load_graph.xlsx'
    source_file_path_slg = r'C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load.xlsx'

    # Load source workbook
    source_wb_slg = load_workbook(source_file_path_slg)  
    # Sheet name of the source workbook
    source_sheet_slg = source_wb_slg['Sheet1']

    # Load destination workbook
    dest_wb_slg = load_workbook(destination_file_path_slg)

    # Sheet name of the destination workbook
    dest_sheet_slg = dest_wb_slg['Sheet']

    # Define source and destination ranges
    source_ranges_slg = [('B2', 'B25'), ('C2', 'C25'), ('D2', 'D25'), ('E2', 'E25'), ('F2', 'F25'), ('G2', 'G25'), ('H2', 'H25'), ('I2', 'I25'), ('J2', 'J25')]
    dest_ranges_slg = [('A2', 'A25'), ('B2', 'B25'), ('C2', 'C25'), ('D2', 'D25'), ('E2', 'E25'), ('F2', 'F25'), ('G2', 'G25'), ('H2', 'H25'), ('I2', 'I25')]

    # Function to copy values from source range to destination range
    def copy_values_slg(source_sheet_slg, dest_sheet_slg, source_range_slg, dest_range_slg):
        source_start_slg, source_end_slg = source_range_slg
        dest_start_slg, dest_end_slg = dest_range_slg

        # Copy values from source to destination
        source_values_slg = []
        for row in source_sheet_slg[source_start_slg:source_end_slg]:
            for cell in row:
                source_values_slg.append(cell.value)
        
        dest_index_slg = 0
        for row_slg in dest_sheet_slg[dest_start_slg:dest_end_slg]:
            for cell_slg in row_slg:
                if dest_index_slg < len(source_values_slg):
                    cell_slg.value = source_values_slg[dest_index_slg]
                    dest_index_slg += 1
                else:
                    break
    
    # Copy data from source to destination based on specified ranges
    for i in range(len(source_ranges_slg)):
        copy_values_slg(source_sheet_slg, dest_sheet_slg, source_ranges_slg[i], dest_ranges_slg[i])
    
    # Save the destination workbook
    dest_wb_slg.save(destination_file_path_slg)

    excel_file = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load.xlsx"

    try:
        conn = mysql.connector.connect(
            host='localhost',
            database='myDb',
            user='root',
            password=''
        )

        # Create MySQL table if it doesn't exist
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS total_substation_load (
                id INT AUTO_INCREMENT PRIMARY KEY,
                substation_load FLOAT(10, 2) NOT NULL,
                insert_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cursor.close()

        # Run main function to continuously check and update database
        main(excel_file, conn)

    except mysql.connector.Error as e:
        print(f"Error connecting to MySQL: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")
    finally:
        if 'conn' in locals() and conn.is_connected():
            conn.close()
            print("MySQL connection closed.")
    
    # Read the Excel file into a Pandas Dataframe
    file_path = r"C:\Users\aslee\OneDrive - MORE ELECTRIC AND POWER CORPORATION\Desktop\DASHBOARD_FINAL\station_load_graph.xlsx"
    df = pd.read_excel(file_path)

    # Connect to the MySQL Database
    try:
        connection = mysql.connector.connect(
            host='localhost',
            database='myDb',
            user='root',
            password=''
        )

        if connection.is_connected():
            print("Successfully connected to the database")
            cursor = connection.cursor()

            # Create a table
            table_name = 'station_load_graph'
            create_table_query = f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                id INT AUTO_INCREMENT PRIMARY KEY,
                Hour INT(11) NOT NULL,
                Lapaz INT(11) NOT NULL,
                Jaro INT(11) NOT NULL,
                Mandurriao INT(11) NOT NULL,
                Molo INT(11) NOT NULL,
                Diversion INT(11) NOT NULL,
                `Mobile SS 1` INT(11) NOT NULL,
                `Mobile SS 2` INT(11) NOT NULL,
                Megaworld INT(11) NOT NULL
            )
            """

            cursor.execute(create_table_query)
            print("Table created successfully or table already exists")

            # Insert data from dataframe into the MySQL table
            for i, row in df.iterrows():
                sql = f"INSERT INTO {table_name} (Hour, Lapaz, Jaro, Mandurriao, Molo, Diversion, `Mobile SS 1`, `Mobile SS 2`, Megaworld) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"
                cursor.execute(sql, tuple(row))
                print(f"Inserted row {i + 1}")

            # Commit the transaction
            connection.commit()
            print("Data committed to the database")

    except Error as e:
        print(f"Error: {e}")

    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            print("MySQL connection is closed.")